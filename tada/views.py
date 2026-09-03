"""TA/DA Portal API — OTP auth, request creation (3 types), workflow, policy validation."""
import io
import json
import secrets
from datetime import timedelta, datetime, date

import openpyxl
from django.conf import settings
from django.core.cache import cache
from django.core.mail import send_mail
from django.db import transaction
from django.utils import timezone

ADMIN_BOOTSTRAP_EMAIL = 'anshul@apisindia.com'
_ADMIN_OTP_KEY = 'tada_admin_otp'
_ADMIN_OTP_TTL = 300
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from config.tz import local_str
from .models import (TadaUser, TadaOTP, TravelRequest, TravelLeg, ExpenseItem,
                     LocalTravelItem, ApprovalLog, ReturnJourney, BookingOption, StayPlan)
from . import policy


# ── helpers ───────────────────────────────────────────────────────────────────
def _mask_email(email):
    if not email or '@' not in email:
        return '***@***.***'
    local, domain = email.split('@', 1)
    ml = local[0] + '*' * max(1, len(local) - 2) + local[-1] if len(local) > 2 else '*' * len(local)
    return f"{ml}@{domain}"


def _parse_date(v):
    if not v:
        return None
    if isinstance(v, (datetime,)):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except Exception:
            pass
    return None


def _parse_dt(v):
    """Datetime from the browser's <input type="datetime-local"> value."""
    if not v:
        return None
    if isinstance(v, datetime):
        return v
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(str(v).strip(), fmt)
        except Exception:
            pass
    d = _parse_date(v)
    return datetime.combine(d, datetime.min.time()) if d else None


def _sf(v, d=0):
    if v is None or str(v).strip() == '':
        return d
    try:
        return float(str(v).replace(',', ''))
    except Exception:
        return d


def serialize_user(u):
    return {
        'id': u.id, 'employee_id': u.employee_id, 'name': u.name, 'email': u.email,
        'designation': u.designation, 'department': u.department, 'level': u.level,
        'hq_city': u.hq_city, 'role': u.role, 'reporting_manager_id': u.reporting_manager_id,
        'caps': policy.get_caps(u.level),
    }


HEAD_LABELS = [('travel', 'Travel'), ('lodging', 'Lodging'), ('food', 'Food / DA'),
               ('local_transport', 'Conveyance'), ('misc', 'Miscellaneous')]


def _leg_nights(leg, is_last):
    """Nights follow where you sleep — the last stop drops one for the trip home."""
    days = leg.days or 0
    return max(0, days - 1) if is_last else days


def leg_policy_heads(level, city, days, nights):
    """The policy ceiling per head for one stop.

    Distinct from what was sanctioned: the sanction seeds from these but the
    employee can edit it, so a trip may be approved above or below entitlement.
    Both numbers have to be visible or an approver cannot tell which is which.
    Travel and misc have no ceiling — the policy sets an entitled class, not a
    fare, and misc is judged on its bills.
    """
    e = policy.leg_estimate(level, city, days, nights)
    return {'travel': None, 'lodging': e['caps']['lodging'], 'food': e['caps']['food'],
            'local_transport': e['caps']['local'], 'misc': None}


def build_settlement(r):
    """Policy vs sanctioned vs claimed, per stop and per head.

    Built here rather than in each UI so the employee filling the claim and
    every approver looking at it are reading the same arithmetic.
    """
    if r.request_type != 'travel_expense':
        return None
    s, level = r.sanction, r.user.level

    # Claimed totals keyed by (leg id, head), plus any policy flags raised.
    claimed, flags = {}, {}
    for it in r.expense_items.all():
        k = (it.leg_id, it.category)
        claimed[k] = claimed.get(k, 0.0) + float(it.claimed_amount)
        if it.policy_flag:
            flags.setdefault(k, []).extend(f for f in it.policy_flag.split(' | ') if f)

    legs = list(s.legs.all()) if s else []
    stops = []
    if legs:
        for i, l in enumerate(legs):
            nights = _leg_nights(l, i == len(legs) - 1)
            stops.append({
                'key': l.id, 'seq': l.seq, 'city': l.destination_city, 'grade': l.city_grade,
                'from_date': str(l.from_date) if l.from_date else None,
                'to_date': str(l.to_date) if l.to_date else None,
                'days': l.days, 'nights': nights,
                'sanctioned': l.estimate_heads,
                'policy': leg_policy_heads(level, l.destination_city, l.days or 0, nights),
            })
    else:
        # Single-destination trip, or a claim filed with no sanction at all.
        days = (s.number_of_days if s else r.number_of_days) or 0
        city = (s.destination_city if s else r.destination_city) or ''
        stops.append({
            'key': None, 'seq': None, 'city': city or 'This trip',
            'grade': policy.city_grade(city) if city else None,
            'from_date': str(s.from_date) if s and s.from_date else (str(r.from_date) if r.from_date else None),
            'to_date': str(s.to_date) if s and s.to_date else (str(r.to_date) if r.to_date else None),
            'days': days, 'nights': max(0, days - 1),
            'sanctioned': s.estimate_heads if s else {k: 0.0 for k, _ in HEAD_LABELS},
            # Policy needs a city (for the grade) and dates (for nights/days).
            # A standalone claim missing either can't be measured against it —
            # say so rather than quietly reporting a ceiling of zero.
            'policy': (leg_policy_heads(level, city, days, max(0, days - 1)) if city and days
                       else {k: None for k, _ in HEAD_LABELS}),
            'policy_unavailable': not (city and days),
        })

    tot = {'policy': 0.0, 'sanctioned': 0.0, 'claimed': 0.0}
    for st in stops:
        rows, sub = [], {'policy': 0.0, 'sanctioned': 0.0, 'claimed': 0.0}
        for key, label in HEAD_LABELS:
            pol = st['policy'].get(key)
            san = float(st['sanctioned'].get(key) or 0)
            act = claimed.get((st['key'], key), 0.0)
            # Every head is listed even at zero. A head that simply vanishes
            # reads as missing data; shown at zero it reads as nothing claimed.
            rows.append({
                'key': key, 'label': label, 'policy': pol, 'sanctioned': san, 'claimed': act,
                'over_policy': round(act - pol, 2) if pol is not None and act > pol else 0,
                'over_sanctioned': round(act - san, 2) if san and act > san else 0,
                'flags': flags.get((st['key'], key), []),
            })
            sub['policy'] += pol or 0
            sub['sanctioned'] += san
            sub['claimed'] += act
        st['rows'] = rows
        st['totals'] = {k: round(v, 2) for k, v in sub.items()}
        for k in tot:
            tot[k] += sub[k]

    # Bills filed against no recognised stop still have to be accounted for.
    loose = [{'key': k[1], 'label': dict(HEAD_LABELS).get(k[1], k[1]), 'claimed': v}
             for k, v in claimed.items()
             if k[0] not in {st['key'] for st in stops}]
    tot['claimed'] += sum(x['claimed'] for x in loose)

    advance = r.advance_adjusted
    return {
        'stops': stops, 'unattributed': loose,
        'totals': {k: round(v, 2) for k, v in tot.items()},
        'advance': advance,
        'net': round(tot['claimed'] - advance, 2),
        'over_policy': round(tot['claimed'] - tot['policy'], 2) if tot['policy'] else 0,
        'over_sanctioned': round(tot['claimed'] - tot['sanctioned'], 2) if tot['sanctioned'] else 0,
    }


def _status_label(r):
    """A tour programme is finally approved by P&C (HR), so the stock label
    "P&C (HR) Approved - Pending Finance" would misreport it as still in flight."""
    if r.request_type == 'tour_sanction' and r.status == 'hr_approved':
        return 'Approved - Finance notified'
    return r.get_status_display()


def serialize_request(r, detail=False):
    _legs = list(r.legs.all())
    d = {
        'id': r.id, 'type': r.request_type, 'type_label': r.get_request_type_display(),
        'status': r.status, 'status_label': _status_label(r),
        'employee_id': r.user.employee_id, 'employee_name': r.user.name,
        'level': r.user.level, 'department': r.user.department,
        'hq_city': r.user.hq_city,
        'purpose': r.purpose, 'from_date': str(r.from_date) if r.from_date else None,
        'to_date': str(r.to_date) if r.to_date else None, 'number_of_days': r.number_of_days,
        'travel_address': r.travel_address, 'from_city': r.from_city, 'destination_city': r.destination_city,
        'city_grade': r.city_grade, 'contact_number': r.contact_number,
        'sanction_number': r.sanction_number, 'estimate_amount': float(r.estimate_amount),
        'travel_mode': r.travel_mode, 'local_travel_type': r.local_travel_type,
        'travel_mode_date': str(r.travel_mode_date) if r.travel_mode_date else None,
        'travel_mode_time_pref': r.travel_mode_time_pref,
        'travel_mode_time_pref_label': r.get_travel_mode_time_pref_display() if r.travel_mode_time_pref else None,
        'trip_type': r.trip_type, 'trip_type_label': r.get_trip_type_display(),
        'traveller_name': r.traveller_name, 'traveller_age': r.traveller_age,
        'return_travel_mode': r.return_travel_mode,
        'return_booking_mode': r.return_booking_mode,
        'return_booking_status': r.return_booking_status,
        'return_booking_reference': r.return_booking_reference,
        'return_booking_carrier': r.return_booking_carrier,
        'return_booking_fare': float(r.return_booking_fare),
        'return_booking_remarks': r.return_booking_remarks,
        'return_booked_by': r.return_booked_by,
        'return_ticket_url': f'/api/tada/ticket/{r.id}/return/' if r.return_booking_ticket else None,
        'return_mode_date': str(r.return_mode_date) if r.return_mode_date else None,
        'return_mode_time_pref': r.return_mode_time_pref,
        'return_mode_time_pref_label': r.get_return_mode_time_pref_display() if r.return_mode_time_pref else None,
        'est_ticket_amount': float(r.est_ticket_amount), 'est_lodging_amount': float(r.est_lodging_amount),
        'est_food_amount': float(r.est_food_amount), 'est_local_amount': float(r.est_local_amount),
        'est_misc_amount': float(r.est_misc_amount), 'advance_amount': float(r.advance_amount),
        'mode_exception_reason': r.mode_exception_reason,
        'policy_flags': [f for f in (r.policy_flags or '').split('\n') if f],
        'is_claimable': r.is_claimable,
        'needs_booking': r.needs_booking,
        'company_borne_fare': r.company_borne_fare,
            'booking_mode': r.booking_mode, 'booking_mode_label': r.get_booking_mode_display(),
            'booking_status': r.booking_status, 'booking_status_label': r.get_booking_status_display(),
            'booking_reference': r.booking_reference, 'booking_carrier': r.booking_carrier,
            'booking_fare': float(r.booking_fare), 'booking_remarks': r.booking_remarks,
            'booked_by': r.booked_by,
            'booked_at': local_str(r.booked_at, '%Y-%m-%d %H:%M'),
            'ticket_url': f'/api/tada/ticket/{r.id}/trip/' if r.booking_ticket else None,
        # Per-head estimate and the policy ceiling behind it, so a claim form can
        # show both without recomputing policy in the browser.
        'heads': r.estimate_heads if r.request_type == 'tour_sanction' else None,
        'policy_heads': (leg_policy_heads(r.user.level, r.destination_city,
                                          r.number_of_days or 0, max(0, (r.number_of_days or 1) - 1))
                         if r.request_type == 'tour_sanction' and not _legs else None),
        'claim_id': (r.open_claim.id if r.request_type == 'tour_sanction' and r.open_claim else None),
        # For a claim: what the trip was sanctioned to cost, so approvers see
        # estimate against actual rather than a bare total.
        'sanction': ({
            'id': r.sanction.id,
            'sanction_number': r.sanction.sanction_number,
            'from_date': str(r.sanction.from_date) if r.sanction.from_date else None,
            'to_date': str(r.sanction.to_date) if r.sanction.to_date else None,
            'destination_city': r.sanction.destination_city,
            'estimate_amount': float(r.sanction.estimate_amount),
            'advance_amount': float(r.sanction.advance_amount),
            'heads': r.sanction.estimate_heads,
        } if r.sanction else None),
        'advance_adjusted': r.advance_adjusted,
        'net_settlement': r.net_settlement,
        'legs': [{
            'seq': l.seq, 'from_date': str(l.from_date) if l.from_date else None,
            'to_date': str(l.to_date) if l.to_date else None, 'days': l.days,
            'from_city': l.from_city, 'destination_city': l.destination_city, 'travel_address': l.travel_address,
            'city_grade': l.city_grade, 'purpose': l.purpose, 'travel_mode': l.travel_mode,
            'ticket_date': str(l.ticket_date) if l.ticket_date else None,
            'ticket_time_pref_label': l.get_ticket_time_pref_display() if l.ticket_time_pref else None,
            'mode_exception_reason': l.mode_exception_reason,
            'est_ticket_amount': float(l.est_ticket_amount), 'est_lodging_amount': float(l.est_lodging_amount),
            'est_food_amount': float(l.est_food_amount), 'est_local_amount': float(l.est_local_amount),
            'heads': l.estimate_heads,
            'booking_mode': l.booking_mode, 'booking_mode_label': l.get_booking_mode_display(),
            'booking_status': l.booking_status, 'booking_status_label': l.get_booking_status_display(),
            'booking_reference': l.booking_reference, 'booking_carrier': l.booking_carrier,
            'booking_fare': float(l.booking_fare), 'booking_remarks': l.booking_remarks,
            'booked_by': l.booked_by,
            'booked_at': local_str(l.booked_at, '%Y-%m-%d %H:%M'),
            'ticket_url': f'/api/tada/ticket/{r.id}/{l.seq}/' if l.booking_ticket else None,
            'policy_heads': leg_policy_heads(r.user.level, l.destination_city, l.days or 0,
                                             _leg_nights(l, i == len(_legs) - 1)),
        } for i, l in enumerate(_legs)],
        # Every option the desk has offered on this trip, across all its
        # journeys, keyed so each screen filters to the one it is showing.
        'booking_options': [{
            'id': o.id, 'journey_key': o.journey_key, 'mode': o.mode, 'carrier': o.carrier,
            'detail': o.detail, 'date': str(o.date) if o.date else None, 'time': o.time,
            'amount': float(o.amount), 'remarks': o.remarks, 'is_selected': o.is_selected,
        } for o in r.booking_options.all()],
        # The planned stay, so an approver sees what the lodging figure is for.
        'stays': [{
            'seq': sp.seq, 'leg_seq': sp.leg_seq, 'location': sp.location,
            'check_in': str(sp.check_in) if sp.check_in else None,
            'check_out': str(sp.check_out) if sp.check_out else None,
            'nights': sp.nights,
        } for sp in r.stays.all()],
        'total_claimed': float(r.total_claimed), 'total_approved': float(r.total_approved),
        'manager_remarks': r.manager_remarks, 'hr_remarks': r.hr_remarks, 'finance_remarks': r.finance_remarks,
        'created_at': local_str(r.created_at, '%Y-%m-%d %H:%M'),
        'submitted_at': local_str(r.submitted_at, '%Y-%m-%d %H:%M'),
    }
    if detail:
        # Policy vs sanctioned vs claimed — the whole picture an approver needs.
        d['settlement'] = build_settlement(r)
        d['expense_items'] = [{
            'id': i.id, 'category': i.category, 'category_label': i.get_category_display(),
            'leg_id': i.leg_id, 'leg_seq': i.leg.seq if i.leg else None,
            'leg_city': i.leg.destination_city if i.leg else None,
            'date': str(i.date) if i.date else None,
            'to_date': str(i.to_date) if i.to_date else None,
            'description': i.description, 'vendor': i.vendor, 'reference_no': i.reference_no,
            'check_in': local_str(i.check_in, '%Y-%m-%d %H:%M'),
            'check_out': local_str(i.check_out, '%Y-%m-%d %H:%M'),
            'nights': i.nights, 'per_night': i.per_night, 'days_covered': i.days_covered,
            'cap_units': i.cap_units, 'cap_basis': i.cap_basis, 'cap_explained': i.cap_explained,
            'from_location': i.from_location, 'to_location': i.to_location, 'mode': i.mode,
            'km': float(i.km), 'claimed_amount': float(i.claimed_amount),
            'approved_amount': float(i.approved_amount) if i.approved_amount is not None else None,
            'has_bill': bool(i.bill), 'bill_url': f'/api/tada/bill/{i.id}/' if i.bill else None,
            'gst_verified': i.gst_verified,
            'policy_cap': float(i.policy_cap) if i.policy_cap is not None else None,
            'policy_flag': i.policy_flag,
        } for i in r.expense_items.all()]
        d['local_items'] = [{
            'id': i.id, 'date': str(i.date) if i.date else None, 'purpose': i.purpose,
            'from_location': i.from_location, 'to_location': i.to_location, 'mode': i.mode,
            'km': float(i.km), 'amount': float(i.amount), 'policy_flag': i.policy_flag,
        } for i in r.local_items.all()]
        d['logs'] = [{
            'stage': l.stage, 'action': l.action, 'by_name': l.by_name,
            'briefing': l.briefing, 'tour_justification': l.tour_justification,
            'advance_remarks': l.advance_remarks,
            'deviation_justification': l.deviation_justification,
            'remarks': l.remarks, 'timestamp': local_str(l.timestamp, '%Y-%m-%d %H:%M'),
        } for l in r.logs.all()]
    return d


def _get_user(request):
    """Identify the acting user by employee_id passed from the SPA (internal tool)."""
    eid = (request.data.get('employee_id') or request.query_params.get('employee_id') or '').strip()
    if not eid:
        return None
    return TadaUser.objects.filter(employee_id=eid, is_active=True).first()


# ── AUTH (OTP) ────────────────────────────────────────────────────────────────
class SendOTPView(APIView):
    def post(self, request):
        eid = (request.data.get('employee_id') or '').strip()
        if not eid:
            return Response({'error': 'Employee ID is required.'}, status=400)
        u = TadaUser.objects.filter(employee_id=eid, is_active=True).first()
        if not u:
            return Response({'error': 'Employee ID not found in TA/DA portal.'}, status=404)
        if not u.email:
            return Response({'error': 'No email on file. Contact P&C (HR) or IT.'}, status=400)
        TadaOTP.objects.filter(user=u, is_used=False).delete()
        code = f"{secrets.randbelow(1_000_000):06d}"
        TadaOTP.objects.create(user=u, code=code, expires_at=timezone.now() + timedelta(minutes=5))
        try:
            send_mail(
                subject='Your APIS TA/DA Portal Login OTP',
                message=(f"Hi {u.name},\n\nYour OTP for the APIS TA/DA Portal is: {code}\n"
                         f"Valid for 5 minutes. Do not share it.\n\n— APIS Team"),
                from_email=settings.DEFAULT_FROM_EMAIL, recipient_list=[u.email], fail_silently=False,
            )
        except Exception as e:
            return Response({'error': f'Could not send OTP email: {e}'}, status=500)
        return Response({'message': 'OTP sent.', 'masked_email': _mask_email(u.email)})


class VerifyOTPView(APIView):
    def post(self, request):
        eid = (request.data.get('employee_id') or '').strip()
        code = (request.data.get('otp') or '').strip()
        u = TadaUser.objects.filter(employee_id=eid, is_active=True).first()
        if not u:
            return Response({'error': 'Invalid employee.'}, status=404)
        otp = TadaOTP.objects.filter(user=u, code=code, is_used=False).order_by('-created_at').first()
        if not otp:
            return Response({'error': 'Invalid OTP.'}, status=400)
        if otp.expires_at < timezone.now():
            return Response({'error': 'OTP expired. Please request a new one.'}, status=400)
        otp.is_used = True
        otp.save(update_fields=['is_used'])
        return Response({'message': 'Login successful.', 'user': serialize_user(u)})


class AdminOTPView(APIView):
    """Send an OTP to the hardcoded admin email — works before any users are imported."""
    def post(self, request):
        code = f"{secrets.randbelow(1_000_000):06d}"
        cache.set(_ADMIN_OTP_KEY, code, timeout=_ADMIN_OTP_TTL)
        try:
            send_mail(
                subject='APIS TA/DA Portal — Admin Access OTP',
                message=f"Admin login OTP for the APIS TA/DA Portal:\n\n  {code}\n\nValid for 5 minutes.\n\n— APIS System",
                from_email=settings.DEFAULT_FROM_EMAIL, recipient_list=[ADMIN_BOOTSTRAP_EMAIL], fail_silently=False,
            )
        except Exception as e:
            cache.delete(_ADMIN_OTP_KEY)
            return Response({'error': f'Failed to send email: {e}'}, status=500)
        return Response({'message': 'OTP sent to admin email.', 'masked_email': _mask_email(ADMIN_BOOTSTRAP_EMAIL)})


class AdminVerifyView(APIView):
    """Verify the admin OTP and return a synthetic admin session."""
    def post(self, request):
        code = (request.data.get('otp') or '').strip()
        stored = cache.get(_ADMIN_OTP_KEY)
        if stored is None:
            return Response({'error': 'OTP expired or not requested.'}, status=400)
        if stored != code:
            return Response({'error': 'Invalid OTP.'}, status=400)
        cache.delete(_ADMIN_OTP_KEY)
        return Response({'message': 'Admin login successful.', 'user': {
            'id': 0, 'employee_id': 'ADMIN', 'name': 'System Admin', 'role': 'admin',
            'level': '', 'designation': 'Administrator', 'department': 'IT', 'hq_city': '', 'caps': {},
        }})


# ── USER DIRECTORY (import / template / list) ─────────────────────────────────
class UserTemplateView(APIView):
    def get(self, request):
        from django.http import HttpResponse
        from openpyxl.styles import PatternFill, Font, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'TADA Users'
        headers = ['Employee ID *', 'Name *', 'Email *', 'Designation', 'Department',
                   'Level (M1-M7/E1-E4) *', 'HQ City', 'Reporting Manager ID',
                   'Role (employee/manager/hr/finance/travel_desk)', 'Vehicle RC No']
        hf = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hf; c.font = Font(color='FFFFFF', bold=True)
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        samples = [
            ['E1001', 'Rahul Verma', 'rahul@apisindia.com', 'Area Sales Manager', 'Sales', 'M1', 'Delhi', 'M2001', 'employee', 'DL01AB1234'],
            ['M2001', 'Suresh Rao', 'suresh@apisindia.com', 'Regional Sales Manager', 'Sales', 'M4', 'Mumbai', 'M5001', 'manager', ''],
            ['H3001', 'Neha P&C', 'neha@apisindia.com', 'P&C Manager', 'People & Culture', 'M3', 'Delhi', '', 'hr', ''],
            ['F4001', 'Amit Finance', 'amit@apisindia.com', 'Finance Manager', 'Finance', 'M4', 'Delhi', '', 'finance', ''],
            ['T5001', 'Kavita Desk', 'kavita@apisindia.com', 'Travel Desk Executive', 'Administration', 'E3', 'Delhi', '', 'travel_desk', ''],
        ]
        for ri, row in enumerate(samples, 2):
            for ci, v in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=v)
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="TADA_Users_Template.xlsx"'
        return resp


class UserImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {e}'}, status=400)
        ALIASES = {
            'employee id': 'employee_id', 'name': 'name', 'email': 'email',
            'designation': 'designation', 'department': 'department',
            'level': 'level', 'level (m1-m7/e1-e4)': 'level',
            'hq city': 'hq_city', 'reporting manager id': 'reporting_manager_id',
            'role': 'role', 'role (employee/manager/hr/finance)': 'role',
            'role (employee/manager/hr/finance/travel_desk)': 'role',
            'vehicle rc no': 'vehicle_rc_no',
        }
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        cmap = {}
        for ci, cell in enumerate(header):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace('*', '').strip()
            if key in ALIASES:
                cmap[ALIASES[key]] = ci
        created = updated = 0
        errors = []
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            data = {k: (row[ci] if ci < len(row) else None) for k, ci in cmap.items()}
            eid = str(data.get('employee_id') or '').strip()
            name = str(data.get('name') or '').strip()
            if not eid or not name:
                errors.append(f'Row {ri}: missing Employee ID or Name'); continue
            role = str(data.get('role') or 'employee').strip().lower()
            # Spreadsheets carry "travel desk" as often as "travel_desk".
            role = role.replace(' ', '_').replace('-', '_')
            if role in ('traveldesk', 'travel_help_desk', 'desk'):
                role = 'travel_desk'
            if role not in ('employee', 'manager', 'hr', 'finance', 'travel_desk', 'admin'):
                role = 'employee'
            _, was_created = TadaUser.objects.update_or_create(
                employee_id=eid,
                defaults={
                    'name': name, 'email': str(data.get('email') or '').strip(),
                    'designation': str(data.get('designation') or '').strip(),
                    'department': str(data.get('department') or '').strip(),
                    'level': str(data.get('level') or '').strip().upper(),
                    'hq_city': str(data.get('hq_city') or '').strip(),
                    'reporting_manager_id': str(data.get('reporting_manager_id') or '').strip(),
                    'role': role, 'vehicle_rc_no': str(data.get('vehicle_rc_no') or '').strip(),
                    'is_active': True,
                },
            )
            created += was_created; updated += not was_created
        return Response({'message': f'✅ Imported {created} new, {updated} updated.',
                         'created': created, 'updated': updated, 'errors': errors,
                         'total': TadaUser.objects.count()})


class UsersListView(APIView):
    def get(self, request):
        users = TadaUser.objects.all()
        return Response({'users': [serialize_user(u) for u in users], 'total': users.count()})

    def delete(self, request):
        TadaUser.objects.all().delete()
        return Response({'message': 'All TA/DA users cleared.'})


class AdminOverviewView(APIView):
    """Super-admin oversight: all requests + stats across every stage."""
    def get(self, request):
        rs = list(TravelRequest.objects.select_related('user').all())
        by_status, by_type, by_dept = {}, {}, {}
        for r in rs:
            by_status[r.status] = by_status.get(r.status, 0) + 1
            by_type[r.request_type] = by_type.get(r.request_type, 0) + 1
            d = r.user.department or 'Unknown'
            by_dept[d] = by_dept.get(d, 0) + 1
        users = TadaUser.objects.all()
        by_role = {}
        for u in users:
            by_role[u.role] = by_role.get(u.role, 0) + 1
        rejected = sum(1 for r in rs if r.status in ('manager_rejected', 'hr_rejected', 'finance_rejected'))

        # hr_approved means two different things now: a tour programme is
        # finally approved there, while a claim still has Finance ahead of it.
        # Counting them together reported approved trips as pending payment.
        tour_done = sum(1 for r in rs
                        if r.status == 'hr_approved' and r.request_type == 'tour_sanction')
        claims_at_finance = sum(1 for r in rs
                                if r.status == 'hr_approved' and r.request_type != 'tour_sanction')
        return Response({
            'total_requests': len(rs),
            'total_users': users.count(),
            'total_claimed': round(sum(float(r.total_claimed) for r in rs), 2),
            'total_approved': round(sum(float(r.total_approved) for r in rs), 2),
            'pending_manager': by_status.get('submitted', 0),
            'pending_hr': by_status.get('manager_approved', 0),
            'pending_finance': claims_at_finance,
            'approved': by_status.get('finance_approved', 0) + tour_done,
            'paid': by_status.get('paid', 0),
            'rejected': rejected,
            'by_status': by_status,
            'by_type': by_type,
            'by_department': dict(sorted(by_dept.items(), key=lambda x: -x[1])),
            'users_by_role': by_role,
            'requests': [serialize_request(r) for r in rs[:1000]],
        })


class AdminResetView(APIView):
    """Clear TA/DA data. what='requests' (default) or 'all' (also users)."""
    def post(self, request):
        what = (request.data.get('what') or 'requests').strip()
        n = TravelRequest.objects.count()
        TravelRequest.objects.all().delete()   # cascades expense/local items + logs
        msg = f'Cleared {n} requests.'
        if what == 'all':
            un = TadaUser.objects.count()
            TadaUser.objects.all().delete()
            msg += f' Cleared {un} users.'
        return Response({'message': msg})


class CapsView(APIView):
    def get(self, request):
        level = request.query_params.get('level', '')
        city = request.query_params.get('city', '')
        out = policy.get_caps(level)
        if city:
            out['city_grade'] = policy.city_grade(city)
        return Response(out)


class EstimateView(APIView):
    """Live policy estimate for the tour-sanction form.

    The form calls this as the employee picks destination/dates/mode so the
    lodging, food and local-conveyance figures shown are the same ones the
    server will recompute on submit — one source of truth for the numbers.
    """
    def get(self, request):
        q = request.query_params
        u = _get_user(request)
        level = u.level if u else q.get('level', '')
        city = q.get('city', '')
        days = policy.trip_days(_parse_date(q.get('from_date')), _parse_date(q.get('to_date'))) or 0
        out = policy.estimate_breakdown(level, city, days,
                                        misc=_sf(q.get('misc')), ticket=_sf(q.get('ticket')))
        within, entitled, flags = policy.mode_entitlement(level, q.get('mode', ''))
        out['mode_within_entitlement'] = within
        out['entitled_mode'] = entitled
        out['mode_flags'] = flags
        return Response(out)

    def post(self, request):
        """Cost a multi-stop itinerary — each leg at its own city grade."""
        d = request.data
        u = _get_user(request)
        level = u.level if u else d.get('level', '')
        legs = []
        for i, lg in enumerate(d.get('legs') or []):
            lf, lt = _parse_date(lg.get('from_date')), _parse_date(lg.get('to_date'))
            if not lf or not lt:
                continue
            legs.append({'seq': int(lg.get('seq', i) or 0), 'from_date': lf, 'to_date': lt,
                         'destination_city': lg.get('destination_city', ''),
                         'travel_mode': lg.get('travel_mode', ''),
                         'est_ticket_amount': _sf(lg.get('est_ticket_amount'))})
        out = policy.itinerary_estimate(level, legs, misc=_sf(d.get('misc')))
        out['itinerary_flags'] = policy.validate_itinerary(
            legs, _parse_date(d.get('from_date')), _parse_date(d.get('to_date')))
        for leg_out, leg_in in zip(out['legs'], sorted(legs, key=lambda l: l['from_date'])):
            within, entitled, flags = policy.mode_entitlement(level, leg_in['travel_mode'])
            leg_out['mode_within_entitlement'] = within
            leg_out['mode_flags'] = flags
        out['entitled_mode'] = policy.mode_options(level)['entitled_mode']
        return Response(out)


# ── EMPLOYEE: create requests ─────────────────────────────────────────────────
class MyRequestsView(APIView):
    def get(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        rs = TravelRequest.objects.filter(user=u)
        return Response({'requests': [serialize_request(r) for r in rs]})


class ClaimableSanctionsView(APIView):
    """Approved trips this employee still has to file expenses for.

    The gap this closes: a sanction would be approved all the way through,
    the trip happened, and nothing connected the eventual expense claim back
    to what was sanctioned — so nobody could see estimate against actual, or
    net the advance off what was owed.
    """
    def get(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        # Must match TravelRequest.is_claimable — a tour programme is finally
        # approved by P&C (HR), so hr_approved belongs here. The older finance
        # statuses stay for trips approved before Finance stopped signing off.
        sanctions = (TravelRequest.objects
                     .filter(user=u, request_type='tour_sanction',
                             status__in=['hr_approved', 'finance_approved', 'paid'])
                     .order_by('-to_date', '-created_at'))
        return Response({'sanctions': [serialize_request(s, detail=True)
                                       for s in sanctions if s.is_claimable]})


class RequestDetailView(APIView):
    def get(self, request, req_id):
        r = TravelRequest.objects.filter(id=req_id).first()
        if not r:
            return Response({'error': 'Not found'}, status=404)
        out = serialize_request(r, detail=True)
        # Tell the UI what this viewer may actually do, so it never offers a
        # button the server will refuse.
        out['permission'] = action_permission(r, _get_user(request))
        return Response(out)


class CreateTourSanctionView(APIView):
    def post(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        d = request.data
        city = d.get('destination_city', '')
        from_date, to_date = _parse_date(d.get('from_date')), _parse_date(d.get('to_date'))
        days = policy.trip_days(from_date, to_date) or 0

        # Who is raising the tickets. On a multi-stop trip this is per leg.
        trip_booking_mode = 'company' if d.get('booking_mode') == 'company' else 'self'

        def _journey_gaps(mode, travel_mode, ticket_date, from_city, where=''):
            """What is missing before this journey can be approved or booked.

            The travel mode is always needed: it is measured against the band's
            entitled class, and an absent mode passes that check by default.
            from_city is always needed too - an approver or the Travel Help
            Desk seeing only the destination has no route, only a place. The
            date is only needed when the desk has to act on it.
            """
            at = f' for {where}' if where else ''
            gaps = []
            if not (from_city or '').strip():
                gaps.append(f'where the journey starts{at}')
            if not (travel_mode or '').strip():
                gaps.append(f'the travel mode{at}')
            if mode == 'company' and not ticket_date:
                gaps.append(f'the date you want to travel{at}')
            return gaps

        # Multi-stop itinerary, if the employee broke the trip down by city.
        raw_legs = d.get('legs') or []
        legs = []
        for i, lg in enumerate(raw_legs):
            lf, lt = _parse_date(lg.get('from_date')), _parse_date(lg.get('to_date'))
            if not lf or not lt:
                continue
            legs.append({
                'seq': int(lg.get('seq', i) or 0), 'from_date': lf, 'to_date': lt,
                'from_city': (lg.get('from_city') or '').strip(),
                'destination_city': lg.get('destination_city', ''),
                'travel_address': lg.get('travel_address', ''),
                'purpose': lg.get('purpose', ''),
                'travel_mode': lg.get('travel_mode', ''),
                'ticket_date': _parse_date(lg.get('ticket_date')),
                'ticket_time_pref': lg.get('ticket_time_pref', ''),
                'mode_exception_reason': (lg.get('mode_exception_reason') or '').strip(),
                'booking_mode': ('company' if lg.get('booking_mode') == 'company' else 'self'),
                # The company pays this fare directly and it never enters the
                # employee's claim, so an estimate here would only inflate the
                # advance they are asking for against a cost that isn't theirs.
                'est_ticket_amount': (0 if lg.get('booking_mode') == 'company'
                                      else _sf(lg.get('est_ticket_amount'))),
            })

        # Where the employee plans to stay. An approver releasing lodging
        # money was previously shown only the rupee figure, never the stay
        # it was for - so the two could not be judged against each other.
        stays = []
        for i, sp in enumerate(d.get('stays') or []):
            loc = (sp.get('location') or '').strip()
            ci, co = _parse_date(sp.get('check_in')), _parse_date(sp.get('check_out'))
            if not (loc or ci or co):
                continue          # an untouched blank row is not an error
            if not loc:
                return Response({'error': f'Stay {i + 1}: add where you are staying.'},
                                status=400)
            if not ci or not co:
                return Response({'error': f'Stay {i + 1} ({loc}): add both the check-in and'
                                          f' check-out dates.'}, status=400)
            if co < ci:
                return Response({'error': f'Stay {i + 1} ({loc}): check-out cannot be before'
                                          f' check-in.'}, status=400)
            stays.append({
                'seq': i, 'location': loc, 'check_in': ci, 'check_out': co,
                # `or ''` would swallow stop 0, which is falsy - the first
                # stop of every multi-city trip.
                'leg_seq': (int(sp['leg_seq'])
                            if str(sp.get('leg_seq', '')).strip().isdigit() else None),
            })

        # Recompute the estimate server-side rather than trusting the posted
        # total — the browser can send anything, and this number drives an
        # approval and a cash advance.
        misc = _sf(d.get('est_misc_amount'))
        advance = _sf(d.get('advance_amount'))
        flags = []

        if legs:
            # The form bounds the stop pickers to the trip window, but a typed
            # date can still get past that, and a stop outside the trip would be
            # costed for days the trip never covers — refuse it outright rather
            # than flag it.
            if from_date and to_date:
                stray = [lg['destination_city'] or f"Stop {lg['seq'] + 1}"
                         for lg in legs if lg['from_date'] < from_date or lg['to_date'] > to_date]
                if stray:
                    return Response({'error': f"{', '.join(stray)} falls outside your travel dates "
                                              f"({from_date} to {to_date}). Stops must be within the trip."}, status=400)

            # Each stop is costed at its own city grade, then summed.
            it = policy.itinerary_estimate(u.level, legs, misc=misc)
            ticket = it['lines']['ticket']
            lodging = _sf(d.get('est_lodging_amount'), it['lines']['lodging'])
            food = _sf(d.get('est_food_amount'), it['lines']['food'])
            local = _sf(d.get('est_local_amount'), it['lines']['local'])
            total = round(ticket + lodging + food + local + misc, 2)
            flags += policy.validate_itinerary(legs, from_date, to_date)
            # Same ceiling check the single-destination path gets, against the
            # summed per-stop ceilings.
            for head, label, val in (('lodging', 'Lodging', lodging), ('food', 'Food / DA', food),
                                     ('local', 'Conveyance', local)):
                cap = it['caps'].get(head)
                if cap is not None and val > cap:
                    flags.append(f'{label} estimate ₹{val:,.0f} exceeds policy ceiling ₹{cap:,.0f}')
            days = it['total_days']
            if not city:
                city = legs[0]['destination_city']
        else:
            # Same rule as a leg: a company-booked ticket is the desk's cost to
            # quote, not the employee's to estimate or draw an advance against.
            ticket = 0 if trip_booking_mode == 'company' else _sf(d.get('est_ticket_amount'))
            base = policy.estimate_breakdown(u.level, city, days, misc=misc, ticket=ticket)
            lodging = _sf(d.get('est_lodging_amount'), base['lines']['lodging'])
            food = _sf(d.get('est_food_amount'), base['lines']['food'])
            local = _sf(d.get('est_local_amount'), base['lines']['local'])
            total = round(ticket + lodging + food + local + misc, 2)
            # advance is checked once for both paths, just below
            flags += policy.validate_estimate(u.level, city, days, lodging=lodging, food=food,
                                              local=local, advance=0, total=total)

        # A stay outside the travel dates is refused the same way a stray stop
        # is: it would be nights the trip never covers.
        if from_date and to_date:
            for sp in stays:
                if sp['check_in'] < from_date or sp['check_out'] > to_date:
                    return Response({
                        'error': f"{sp['location']} ({sp['check_in']} to {sp['check_out']}) "
                                 f"falls outside your travel dates ({from_date} to {to_date})."
                    }, status=400)

        # Asking for lodging money without saying where you are staying leaves
        # the approver releasing an advance against nothing they can check.
        # Checked against the computed figure, not the posted one: lodging is
        # seeded from policy when the form omits it, and a seeded figure is
        # just as much money as a typed one.
        if lodging > 0 and not stays:
            return Response({
                'error': 'You have estimated lodging, so please add where you plan to stay - '
                         'location, check-in and check-out.'
            }, status=400)

        # The advance is cash released before the trip, so an over-limit request
        # is refused outright rather than flagged for someone to catch later.
        adv_err = policy.advance_error(advance, total)
        if adv_err:
            return Response({'error': adv_err}, status=400)

        # Every mode in play must be within entitlement, or carry a reason —
        # per leg for a multi-stop trip, else the single request-level mode.
        reason = (d.get('mode_exception_reason') or '').strip()
        checks = ([(lg['travel_mode'], lg['mode_exception_reason'], lg['destination_city']) for lg in legs]
                  if legs else [(d.get('travel_mode', ''), reason, city)])
        for mode, why, where in checks:
            if not mode:
                continue
            within_mode, _, mode_flags = policy.mode_entitlement(u.level, mode)
            flags += [f'{where}: {f}' if where and legs else f for f in mode_flags]
            if not within_mode and not why:
                label = f' for {where}' if where and legs else ''
                return Response({'error': f'{mode}{label} is outside your travel entitlement — please give a reason for the exception.'}, status=400)

        # The class entitlement applies to the way home too.
        if (d.get('return_travel_mode') or '').strip() and d.get('trip_type') != 'one_way':
            _ok_ret, _, _ret_flags = policy.mode_entitlement(u.level, d['return_travel_mode'])
            flags += [f'Return journey: {f}' for f in _ret_flags]
            if not _ok_ret and not reason:
                return Response({'error': f"{d['return_travel_mode']} for the journey home is outside your "
                                          "travel entitlement — please give a reason for the exception."}, status=400)

        gaps, needs_desk = [], False
        if legs:
            for lg in legs:
                gaps += _journey_gaps(lg['booking_mode'], lg['travel_mode'], lg['ticket_date'],
                                      lg['from_city'], lg['destination_city'] or f"stop {lg['seq'] + 1}")
                needs_desk = needs_desk or lg['booking_mode'] == 'company'
        else:
            gaps += _journey_gaps(trip_booking_mode, d.get('travel_mode'),
                                  _parse_date(d.get('travel_mode_date')), d.get('from_city'))
            needs_desk = trip_booking_mode == 'company'

        # The way home is half the trip. It was captured as a bare date, so it
        # was never checked against the class entitlement and the desk was never
        # asked to raise it - which is how company-booked trips arrived with an
        # outbound ticket and no way back.
        trip_type = 'one_way' if d.get('trip_type') == 'one_way' else 'round_trip'
        return_booking_mode = 'company' if d.get('return_booking_mode') == 'company' else 'self'
        if trip_type == 'round_trip':
            gaps += _journey_gaps(return_booking_mode, d.get('return_travel_mode'),
                                  _parse_date(d.get('return_mode_date')), 'ok', 'the journey home')
            needs_desk = needs_desk or return_booking_mode == 'company'
        else:
            return_booking_mode = 'self'
        if gaps:
            tail = ('. The Travel Help Desk cannot book without it.' if needs_desk
                    else '. Your travel class is approved against it.')
            return Response({'error': 'Please add ' + ', '.join(gaps) + tail}, status=400)

        # A ticket is raised against a person, not an employee code. Only the
        # employee knows the spelling on their Aadhaar card, the number they
        # will answer while away and their age - and a ticket booked on a guess
        # is refused at the counter, so these are asked for the moment the
        # company is doing the booking.
        traveller_name = (d.get('traveller_name') or '').strip()
        traveller_age = d.get('traveller_age')
        contact_number = (d.get('contact_number') or '').strip()
        if needs_desk:
            who = []
            if not traveller_name:
                who.append('your name exactly as it appears on your Aadhaar card')
            if not contact_number:
                who.append('a contact number you can be reached on while travelling')
            try:
                traveller_age = int(traveller_age)
                if not (14 <= traveller_age <= 100):
                    raise ValueError
            except (TypeError, ValueError):
                traveller_age = None
                who.append('your age')
            if who:
                return Response({'error': 'The Travel Help Desk raises tickets in your name, so please add '
                                          + ', '.join(who) + '.'}, status=400)
        else:
            try:
                traveller_age = int(traveller_age)
            except (TypeError, ValueError):
                traveller_age = None

        r = TravelRequest.objects.create(
            user=u, request_type='tour_sanction', status='submitted',
            purpose=d.get('purpose', ''), travel_address=d.get('travel_address', ''),
            from_city=(d.get('from_city') or '').strip() if not legs else legs[0]['from_city'],
            destination_city=city, city_grade=policy.city_grade(city),
            from_date=from_date, to_date=to_date,
            contact_number=contact_number, sanction_number=d.get('sanction_number', ''),
            traveller_name=traveller_name, traveller_age=traveller_age,
            travel_mode=d.get('travel_mode', ''),
            travel_mode_date=_parse_date(d.get('travel_mode_date')),
            travel_mode_time_pref=d.get('travel_mode_time_pref', ''),
            trip_type=trip_type,
            return_travel_mode=(d.get('return_travel_mode') or '') if trip_type == 'round_trip' else '',
            return_mode_date=_parse_date(d.get('return_mode_date')) if trip_type == 'round_trip' else None,
            return_mode_time_pref=d.get('return_mode_time_pref', '') if trip_type == 'round_trip' else '',
            return_booking_mode=return_booking_mode,
            return_booking_status=('pending' if return_booking_mode == 'company' else 'not_required'),
            booking_mode=trip_booking_mode,
            booking_status=('pending' if trip_booking_mode == 'company' and not legs else 'not_required'),
            est_ticket_amount=ticket, est_lodging_amount=lodging, est_food_amount=food,
            est_local_amount=local, est_misc_amount=misc,
            estimate_amount=total, advance_amount=advance,
            mode_exception_reason=reason, policy_flags='\n'.join(flags),
            submitted_at=timezone.now(),
        )
        for sp in stays:
            StayPlan.objects.create(
                request=r, seq=sp['seq'], leg_seq=sp['leg_seq'],
                location=sp['location'], check_in=sp['check_in'], check_out=sp['check_out'])

        if legs:
            per_leg = {l['seq']: l for l in policy.itinerary_estimate(u.level, legs, misc=misc)['legs']}
            for lg in legs:
                e = per_leg.get(lg['seq'], {}).get('lines', {})
                TravelLeg.objects.create(
                    request=r, seq=lg['seq'], from_date=lg['from_date'], to_date=lg['to_date'],
                    from_city=lg['from_city'],
                    destination_city=lg['destination_city'], travel_address=lg['travel_address'],
                    city_grade=policy.city_grade(lg['destination_city']),
                    purpose=lg['purpose'], travel_mode=lg['travel_mode'],
                    ticket_date=lg['ticket_date'], ticket_time_pref=lg['ticket_time_pref'],
                    mode_exception_reason=lg['mode_exception_reason'],
                    est_ticket_amount=lg['est_ticket_amount'],
                    est_lodging_amount=e.get('lodging', 0), est_food_amount=e.get('food', 0),
                    est_local_amount=e.get('local', 0),
                    booking_mode=lg['booking_mode'],
                    booking_status=('pending' if lg['booking_mode'] == 'company' else 'not_required'),
                )

        ApprovalLog.objects.create(request=r, stage='employee', action='submitted', by_name=u.name)
        notify_submitted(r)
        return Response({'message': 'Tour Programme submitted for Manager approval.',
                         'request': serialize_request(r, detail=True)})


class CreateTravelExpenseView(APIView):
    """Multipart: 'payload' = JSON (request + items[]), files bill_<idx> per item."""
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    @transaction.atomic
    def post(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        try:
            payload = json.loads(request.data.get('payload') or '{}')
        except Exception:
            payload = request.data
        items = payload.get('items', [])

        # A claim may settle a previously approved sanction. Validate it belongs
        # to this employee and is still open — otherwise one trip could be
        # claimed for twice, or against someone else's approval.
        sanction = None
        sid = payload.get('sanction_id')
        if sid:
            sanction = TravelRequest.objects.filter(id=sid, user=u,
                                                    request_type='tour_sanction').first()
            if not sanction:
                return Response({'error': 'That sanction was not found against your account.'}, status=404)
            # A tour programme is finally approved by P&C (HR), so hr_approved counts as
            # settled-and-claimable; the older finance statuses stay valid for
            # trips approved before Finance stopped signing these off.
            if sanction.status not in ('hr_approved', 'finance_approved', 'paid'):
                return Response({'error': f'That trip is not fully approved yet '
                                          f'({_status_label(sanction)}).'}, status=400)
            if sanction.open_claim:
                return Response({'error': 'Expenses have already been filed for that trip.'}, status=400)

        # Fall back to the sanction's own trip details for anything not restated.
        city = payload.get('destination_city') or (sanction.destination_city if sanction else '')
        cgrade = policy.city_grade(city)
        fd = _parse_date(payload.get('from_date')) or (sanction.from_date if sanction else None)
        td = _parse_date(payload.get('to_date')) or (sanction.to_date if sanction else None)

        # Hard-stop: 60-day deadline (checked against the earliest item / from_date)
        if fd and not policy.is_within_deadline(fd):
            return Response({'error': f'Blocked: travel date is beyond the {policy.SUBMISSION_DEADLINE_DAYS}-day '
                                      f'submission deadline.'}, status=400)

        # Check the items before writing anything. A fare the company already
        # paid cannot be claimed back, and finding that out half-way through
        # used to leave a half-built claim behind.
        pre_legs = {l.seq: l for l in sanction.legs.all()} if sanction else {}
        for it in items:
            if it.get('category') != 'travel':
                continue
            src = pre_legs.get(it.get('leg_seq')) or (sanction if not pre_legs else None)
            if src and src.booking_mode == 'company' and src.booking_status == 'booked':
                where = getattr(src, 'destination_city', '') or 'this trip'
                return Response({'error': f'The ticket for {where} was booked and paid for by the '
                                          f'company, so its fare cannot be claimed.'}, status=400)

        r = TravelRequest.objects.create(
            user=u, request_type='travel_expense', status='submitted', sanction=sanction,
            purpose=payload.get('purpose') or (sanction.purpose if sanction else ''),
            destination_city=city, city_grade=cgrade, from_date=fd, to_date=td,
            sanction_number=payload.get('sanction_number') or (sanction.sanction_number if sanction else ''),
            travel_mode=payload.get('travel_mode') or (sanction.travel_mode if sanction else ''),
            submitted_at=timezone.now(),
        )
        # Bills can be filed per stop. Only the sanction's own legs are accepted,
        # so a claim can't attribute spend to a leg of someone else's trip.
        legs_by_seq = {l.seq: l for l in sanction.legs.all()} if sanction else {}


        total = 0.0
        for idx, it in enumerate(items):
            claimed = _sf(it.get('claimed_amount'))
            bill = request.FILES.get(f'bill_{idx}')
            leg = legs_by_seq.get(it.get('leg_seq'))
            # A stop's own city grade decides its caps, not the trip's headline city.
            item_grade = policy.city_grade(leg.destination_city) if leg else cgrade
            # Scale the per-night/per-day ceilings by what this stay actually
            # covers — one invoice for a four-night hotel is not a one-night bill.
            if leg:
                span_days = leg.days or 1
                span_nights = _leg_nights(leg, leg.seq == max(legs_by_seq))
            else:
                span_days = (sanction.number_of_days if sanction else r.number_of_days) or 1
                span_nights = max(1, span_days - 1)
            check_in = _parse_dt(it.get('check_in'))
            check_out = _parse_dt(it.get('check_out'))
            item_from = _parse_date(it.get('date'))
            item_to = _parse_date(it.get('to_date'))

            # Judge the bill on what it actually covers rather than on the
            # stop's length: a three-night folio measured against a one-night
            # ceiling flags a compliant claim, and a two-night stay on a
            # four-night leg would otherwise be allowed twice what it should.
            if check_in and check_out:
                span_nights = max(1, (check_out.date() - check_in.date()).days)
            if item_from and item_to:
                span_days = max(1, (item_to - item_from).days + 1)

            cap, flags = policy.validate_expense_item(
                u.level, it.get('category', 'misc'), item_grade, claimed, bool(bill),
                mode=it.get('mode', ''), km=_sf(it.get('km')), date_val=item_from,
                nights=span_nights, days=span_days)
            item = ExpenseItem.objects.create(
                request=r, leg=leg, category=it.get('category', 'misc'), date=item_from,
                to_date=item_to, description=it.get('description', ''),
                vendor=it.get('vendor', ''), reference_no=it.get('reference_no', ''),
                check_in=check_in, check_out=check_out,
                from_location=it.get('from_location', ''),
                to_location=it.get('to_location', ''), mode=it.get('mode', ''), km=_sf(it.get('km')),
                claimed_amount=claimed, gst_verified=bool(it.get('gst_verified')),
                policy_cap=cap, policy_flag=' | '.join(flags),
                cap_units=(span_nights if it.get('category') == 'lodging' else span_days),
                cap_basis=('night' if it.get('category') == 'lodging' else 'day'),
            )
            if bill:
                item.bill.save(f'bill_{r.id}_{idx}_{bill.name}', bill, save=True)
            total += claimed
        r.total_claimed = total
        r.save(update_fields=['total_claimed'])
        ApprovalLog.objects.create(request=r, stage='employee', action='submitted', by_name=u.name)
        notify_submitted(r)
        return Response({'message': 'Travelling Expenses submitted for Manager approval.',
                         'request': serialize_request(r, detail=True)})


class CreateLocalTravelView(APIView):
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def post(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        try:
            payload = json.loads(request.data.get('payload') or '{}')
        except Exception:
            payload = request.data
        rows = payload.get('items', [])
        r = TravelRequest.objects.create(
            user=u, request_type='local_travel', status='submitted',
            purpose=payload.get('purpose', ''), local_travel_type=payload.get('local_travel_type', ''),
            from_date=_parse_date(payload.get('from_date')), to_date=_parse_date(payload.get('to_date')),
            submitted_at=timezone.now(),
        )
        band = policy.band_for_level(u.level)
        daily_cap = policy.LOCAL_CONVEYANCE.get(band, {}).get('daily')
        total = 0.0
        for it in rows:
            amt = _sf(it.get('amount'))
            flags = []
            if daily_cap is not None and amt > daily_cap:
                flags.append(f'Exceeds daily local conveyance cap ₹{daily_cap}')
            LocalTravelItem.objects.create(
                request=r, date=_parse_date(it.get('date')), purpose=it.get('purpose', ''),
                from_location=it.get('from_location', ''), to_location=it.get('to_location', ''),
                mode=it.get('mode', ''), km=_sf(it.get('km')), amount=amt, policy_flag=' | '.join(flags))
            total += amt
        r.total_claimed = total
        r.save(update_fields=['total_claimed'])
        ApprovalLog.objects.create(request=r, stage='employee', action='submitted', by_name=u.name)
        notify_submitted(r)
        return Response({'message': 'Local Travel submitted for Manager approval.',
                         'request': serialize_request(r, detail=True)})


class BillDownloadView(APIView):
    def get(self, request, item_id):
        from django.http import FileResponse
        it = ExpenseItem.objects.filter(id=item_id).first()
        if not it or not it.bill:
            return Response({'error': 'No bill'}, status=404)
        return FileResponse(it.bill.open('rb'), filename=f'bill_{item_id}.pdf')


class TicketDownloadView(APIView):
    """The uploaded ticket, as the record of what was actually booked."""
    def get(self, request, req_id, journey_key):
        from django.http import FileResponse
        r = TravelRequest.objects.filter(id=req_id).first()
        if not r:
            return Response({'error': 'Not found'}, status=404)
        target, err = _resolve_journey(r, journey_key)
        if err:
            return err
        if not target.booking_ticket:
            return Response({'error': 'No ticket on file for this journey.'}, status=404)
        name = target.booking_ticket.name.rsplit('/', 1)[-1]
        return FileResponse(target.booking_ticket.open('rb'), filename=name)


# ── APPROVALS (Manager / P&C (HR) / Finance) ────────────────────────────────────────
NL = chr(10)


# -- Notifications ------------------------------------------------------------
# Who hears about a request, and when. Mail is a side effect of the workflow and
# never a gate on it: a bounced address or an SMTP outage must not stop an
# approval that has already happened, so every send is best-effort.

def _people(role):
    """Everyone holding a role who can actually receive mail."""
    return list(TadaUser.objects.filter(role=role, is_active=True).exclude(email=""))


def _manager_of(u):
    if not u.reporting_manager_id:
        return None
    return TadaUser.objects.filter(employee_id=u.reporting_manager_id, is_active=True).first()


def _trip_summary(r):
    """Short subject-line description of the request."""
    bits = [r.get_request_type_display()]
    if r.destination_city:
        bits.append(r.destination_city)
    if r.from_date and r.to_date:
        bits.append("%s to %s" % (_d(r.from_date), _d(r.to_date)))
    return " | ".join(bits)


def _d(v):
    """Dates read as 05 Aug 2026 in mail, not 2026-08-05."""
    try:
        return v.strftime("%d %b %Y")
    except Exception:
        return str(v)


def _money(v):
    return "Rs. %s" % format(float(v or 0), ",.0f")


def _detail_block(r):
    """The particulars, laid out as a block so the mail can be read on its own
    without opening the portal."""
    rows = [("Employee", "%s (%s)" % (r.user.name, r.user.employee_id))]
    if r.user.department:
        rows.append(("Department", r.user.department))
    rows.append(("Request Type", r.get_request_type_display()))
    if r.destination_city:
        rows.append(("Destination", r.destination_city))
    if r.from_date and r.to_date:
        rows.append(("Travel Dates", "%s to %s" % (_d(r.from_date), _d(r.to_date))))
    if r.purpose:
        rows.append(("Purpose", r.purpose))
    if r.travel_mode:
        rows.append(("Mode of Travel", r.travel_mode))
    if r.estimate_amount:
        rows.append(("Estimated Cost", _money(r.estimate_amount)))
    if r.advance_amount:
        rows.append(("Advance Requested", _money(r.advance_amount)))
    if r.total_claimed:
        rows.append(("Amount Claimed", _money(r.total_claimed)))

    width = max(len(k) for k, _ in rows)
    out = [k.ljust(width) + " : " + str(v) for k, v in rows]

    # The stay the lodging figure above is actually for - an approver
    # releasing an advance should not have to open the portal to see it.
    stays = list(r.stays.all())
    if stays:
        out.append("")
        out.append("Planned stay:")
        for sp in stays:
            n = sp.nights
            out.append("  - %s : %s to %s%s" % (
                sp.location, _d(sp.check_in), _d(sp.check_out),
                ("" if n is None else " (%d night%s)" % (n, "" if n == 1 else "s"))))

    flags = [f for f in (r.policy_flags or "").split(NL) if f]
    if flags:
        out.append("")
        out.append("Policy observations:")
        out += ["  - " + f for f in flags]
    return out


def _sign_off():
    return ["", "This is an automated message from the APIS TA/DA Portal.",
            "Please do not reply to this email.", "", "Regards,",
            "APIS India Limited"]


def notify(subject, greeting, lines, recipients, attachment=None):
    """Compose and send. Never raises - mail is a side effect of the workflow,
    not a gate on it, so a bad address must not fail an approval.

    attachment, when given, is the ticket itself: the mail confirming a
    booking is also the record of it, so whoever reads it later does not have
    to log into the portal to find what was actually bought.
    """
    to = sorted({x.email for x in recipients if x and x.email})
    if not to:
        return
    body = NL.join([greeting, ""] + lines + _sign_off())
    try:
        if attachment:
            from django.core.mail import EmailMessage
            msg = EmailMessage(subject=subject, body=body, from_email=settings.DEFAULT_FROM_EMAIL,
                               to=to)
            attachment.open('rb')
            msg.attach(attachment.name.rsplit('/', 1)[-1], attachment.read(), None)
            attachment.close()
            msg.send(fail_silently=True)
        else:
            send_mail(subject=subject, message=body,
                      from_email=settings.DEFAULT_FROM_EMAIL, recipient_list=to,
                      fail_silently=True)
    except Exception:
        pass


def notify_submitted(r):
    """Raised - inform the approver that it awaits their action."""
    mgr = _manager_of(r.user)
    lines = (["A new travel request has been submitted and is awaiting your approval.", ""]
             + _detail_block(r)
             + ["", "Kindly review and record your decision in the APIS TA/DA Portal."])
    notify("Action Required | %s | %s" % (_trip_summary(r), r.user.name),
           "Dear %s," % (mgr.name if mgr else "Sir/Madam"),
           lines, [mgr] if mgr else _people("hr"))


def notify_actioned(r, stage, by_name, action, remarks=""):
    """Approved or rejected.

    An approval informs the employee and moves to the next desk. A rejection is
    sent only to the employee - the request stops there, so no one downstream
    has anything to act on.
    """
    stage_name = {"manager": "Reporting Manager", "hr": "P&C (HR)", "finance": "Finance"}.get(stage, stage.title())
    extra = ["", "Remarks: %s" % remarks] if remarks else []

    if action == "rejected":
        lines = (["We regret to inform you that your travel request has not been approved.", "",
                  "Rejected by : %s (%s)" % (by_name, stage_name), ""]
                 + _detail_block(r) + extra
                 + ["", "For any clarification, please contact your %s." % stage_name])
        notify("Travel Request Not Approved | %s" % _trip_summary(r),
               "Dear %s," % r.user.name, lines, [r.user])
        return

    # A tour programme is finally approved by P&C (HR); Finance is kept informed.
    fyi_only = stage == "hr" and r.request_type == "tour_sanction"
    if stage == "manager":
        nxt = _people("hr")
        note = "It has been forwarded to P&C (HR) for further approval."
    elif fyi_only:
        nxt = _people("finance")
        note = "Your tour programme now stands approved. Finance has been informed for their records."
    elif stage == "hr":
        nxt = _people("finance")
        note = "It has been forwarded to Finance for further approval."
    else:
        nxt, note = [], ""

    lines = ["Your travel request has been approved.", "",
             "Approved by : %s (%s)" % (by_name, stage_name), ""]
    lines += _detail_block(r) + extra
    if note:
        lines += ["", note]
    notify("Travel Request Approved | %s | %s" % (_trip_summary(r), r.user.name),
           "Dear %s," % r.user.name, lines, [r.user])

    if not nxt:
        return

    if fyi_only:
        subject = "For Information | Tour Programme Approved | %s" % r.user.name
        opening = ("The following tour programme has been approved by the Reporting Manager "
                   "and P&C (HR), and is shared with Finance for information and records.")
        closing = "Kindly take note of the advance and estimated cost indicated above."
    else:
        subject = "Action Required | %s | %s" % (_trip_summary(r), r.user.name)
        opening = ("The following travel request has been approved by %s (%s) "
                   "and is now pending your approval." % (by_name, stage_name))
        closing = "Kindly review and record your decision in the APIS TA/DA Portal."

    notify(subject, "Dear Sir/Madam,",
           [opening, ""] + _detail_block(r) + extra + ["", closing], nxt)


def _journey_label(x, r):
    """A leg reads as its stop; a single-destination trip as the trip.

    The way home reads as the way home — "Delhi" on both the outbound and the
    return would give the desk two identical lines to book. Outbound legs show
    the route (from -> to), not just the destination, so the desk knows where
    to actually raise the ticket from.
    """
    if getattr(x, 'is_return', False):
        # The return starts wherever the trip actually ends: the last stop on
        # a multi-city itinerary, or the single destination otherwise -
        # r.destination_city alone is only the FIRST stop on a multi-city trip.
        last_leg = r.legs.order_by('-seq').first()
        frm = (last_leg.destination_city if last_leg else r.destination_city) or 'the last stop'
        return '%s -> %s (return)' % (frm, r.user.hq_city or 'base')
    city = getattr(x, 'destination_city', '') or r.destination_city
    frm = getattr(x, 'from_city', '') or getattr(r, 'from_city', '')
    if frm and city:
        return '%s -> %s' % (frm, city)
    return city or 'the journey'


def notify_booking_needed(r):
    """Fully approved, and the desk has tickets to raise."""
    legs = [x for x in r.company_booked_legs if x.booking_status == 'pending']
    if not legs:
        return
    lines = ["The following tour programme is approved and requires ticketing "
             "by the Travel Help Desk.", ""] + _detail_block(r) + ["", "Journeys to be booked:"]
    for x in legs:
        parts = ["  - %s" % _journey_label(x, r)]
        mode = getattr(x, 'travel_mode', '') or ('' if getattr(x, 'is_return', False) else r.travel_mode)
        if mode:
            parts.append("by %s" % mode)
        dt = getattr(x, 'ticket_date', None) or (None if getattr(x, 'is_return', False) else r.travel_mode_date)
        if dt:
            parts.append("on %s" % _d(dt))
        pref = (getattr(x, 'get_ticket_time_pref_display', None) or
                getattr(r, 'get_travel_mode_time_pref_display', None))
        try:
            tp = pref() if pref else ''
        except Exception:
            tp = ''
        if tp:
            parts.append("(%s)" % tp)
        lines.append(" ".join(parts))
    # A carrier asks for a person, not an employee code, so the details the
    # employee gave for exactly this purpose travel with the request.
    lines += ["", "Traveller details, as given by the employee:",
              "  Name (as per Aadhaar) : %s" % (r.traveller_name or r.user.name),
              "  Age                   : %s" % (r.traveller_age or 'not stated'),
              "  Contact while touring : %s" % (r.contact_number or 'not stated'),
              "", "Kindly book and record the ticket details in the APIS TA/DA Portal."]
    notify("Ticketing Required | %s | %s" % (_trip_summary(r), r.user.name),
           "Dear Sir/Madam,", lines, _people("travel_desk"))


def notify_booked(r, target, action, by_name):
    """Tell the employee what the desk did."""
    where = _journey_label(target, r)
    if action == "cancelled":
        lines = ["The Travel Help Desk could not complete the booking for %s." % where, ""]
        if target.booking_remarks:
            lines += ["Reason: %s" % target.booking_remarks, ""]
        lines += _detail_block(r) + ["", "Please contact the Travel Help Desk for the next steps."]
        notify("Booking Not Completed | %s" % _trip_summary(r),
               "Dear %s," % r.user.name, lines, [r.user])
        return

    lines = ["Your ticket has been booked by the Travel Help Desk.", "",
             "Journey   : %s" % where]
    if target.booking_carrier:
        lines.append("Operator  : %s" % target.booking_carrier)
    lines.append("Ticket / PNR : %s" % target.booking_reference)
    lines.append("Fare      : %s (borne by the company)" % _money(target.booking_fare))
    lines.append("Booked by : %s" % by_name)
    if target.booking_remarks:
        lines += ["", "Remarks: %s" % target.booking_remarks]
    lines += ["", "As this ticket is paid for by the company, please do not include "
                  "this fare in your expense claim."]
    if target.booking_ticket:
        lines += ["", "The ticket is attached to this mail, and is also available in the portal."]
    notify("Ticket Booked | %s" % _trip_summary(r), "Dear %s," % r.user.name, lines, [r.user],
           attachment=target.booking_ticket or None)


def notify_options_sent(r, target):
    """The desk found more than one way to travel — the employee has to pick."""
    where = _journey_label(target, r)
    opts = list(BookingOption.objects.filter(request=r, journey_key=target.journey_key))
    lines = ["The Travel Help Desk has found the following options for %s, and needs "
             "you to confirm which one to book." % where, ""]
    for i, o in enumerate(opts, 1):
        lines.append("  %d. %s%s %s — %s" % (
            i, o.mode + ' ' if o.mode else '', o.carrier or '', o.get_time_label(), _money(o.amount)))
        if o.detail:
            lines.append("     %s" % o.detail)
    lines += ["", "Please open the request in the APIS TA/DA Portal and confirm your choice — "
                  "the desk cannot book anything until you do."]
    notify("Please Confirm Your Ticket | %s | %s" % (_trip_summary(r), where),
           "Dear %s," % r.user.name, lines, [r.user])


def notify_option_selected(r, target, chosen, by_name):
    """The employee has chosen — the desk can now actually buy the ticket."""
    where = _journey_label(target, r)
    lines = ["%s has confirmed which option to book for %s:" % (by_name, where), "",
             "  %s%s %s — %s" % (chosen.mode + ' ' if chosen.mode else '', chosen.carrier or '',
                                 chosen.get_time_label(), _money(chosen.amount))]
    if chosen.detail:
        lines.append("  %s" % chosen.detail)
    lines += ["", "Please complete the ticketing and upload the ticket in the portal."]
    notify("Employee Confirmed | %s | %s" % (_trip_summary(r), where),
           "Dear Sir/Madam,", lines, _people("travel_desk"))


def notify_paid(r, by_name):
    lines = (["The payment against your travel claim has been settled.", "",
              "Processed by : %s (Finance)" % by_name, ""] + _detail_block(r))
    notify("Payment Settled | %s" % _trip_summary(r), "Dear %s," % r.user.name, lines, [r.user])


_STAGE_FLOW = {
    'manager': {'from': 'submitted',        'approve': 'manager_approved', 'reject': 'manager_rejected'},
    'hr':      {'from': 'manager_approved', 'approve': 'hr_approved',      'reject': 'hr_rejected'},
    'finance': {'from': 'hr_approved',      'approve': 'finance_approved', 'reject': 'finance_rejected'},
}


def action_permission(r, u):
    """May this user approve/reject/pay this request, and if not, why not?

    One place, used both by the action endpoint and by the detail response the
    UI renders from — otherwise the screen offers an Approve button that the
    server then refuses, which reads as the app being broken.
    """
    if not u:
        return {'can_approve': False, 'can_pay': False, 'reason': 'Login required.'}
    if r.user_id == u.id:
        return {'can_approve': False, 'can_pay': False, 'reason': 'You cannot action your own request.'}
    flow = _STAGE_FLOW.get(u.role)
    if not flow:
        return {'can_approve': False, 'can_pay': False, 'reason': 'Your role cannot action requests.'}
    if u.role == 'manager' and (r.user.reporting_manager_id or '') != u.employee_id:
        return {'can_approve': False, 'can_pay': False,
                'reason': 'This request is not from one of your reports.'}
    # A tour programme is finally approved by P&C (HR). Finance is notified by mail
    # and has nothing to action, so it must not sit in their queue either.
    if u.role == 'finance' and r.request_type == 'tour_sanction':
        return {'can_approve': False, 'can_pay': False,
                'reason': 'Tour programmes are approved by P&C (HR); Finance is notified for information.'}
    can_pay = u.role == 'finance' and r.status == 'finance_approved'
    if r.status != flow['from']:
        return {'can_approve': False, 'can_pay': can_pay,
                'reason': None if can_pay else f'Awaiting a different stage ({r.get_status_display()}).'}
    return {'can_approve': True, 'can_pay': can_pay, 'reason': None}


class PendingQueueView(APIView):
    """Role-based queue of requests awaiting the acting user's action."""
    def get(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        role = u.role
        if role == 'manager':
            # requests submitted by this manager's direct reports
            report_ids = list(TadaUser.objects.filter(reporting_manager_id=u.employee_id)
                              .values_list('employee_id', flat=True))
            rs = TravelRequest.objects.filter(user__employee_id__in=report_ids, status='submitted')
            others = TravelRequest.objects.filter(user__employee_id__in=report_ids).exclude(status='submitted')
        elif role == 'hr':
            rs = TravelRequest.objects.filter(status='manager_approved')
            others = TravelRequest.objects.filter(status__in=['hr_approved', 'hr_rejected', 'finance_approved', 'finance_rejected', 'paid'])
        elif role == 'finance':
            # Tour programmes end at P&C (HR); Finance only ever sees claims to action.
            rs = TravelRequest.objects.filter(status='hr_approved').exclude(request_type='tour_sanction')
            others = TravelRequest.objects.filter(status__in=['finance_approved', 'finance_rejected', 'paid'])
        else:
            return Response({'pending': [], 'processed': []})
        return Response({'pending': [serialize_request(r) for r in rs],
                         'processed': [serialize_request(r) for r in others[:100]]})


class BookingQueueView(APIView):
    """What the Travel Help Desk has to book, and what it has already booked.

    Only fully approved trips appear: booking a journey that P&C (HR) then rejects
    wastes a fare and a cancellation fee.
    """
    def get(self, request):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        if u.role not in ('travel_desk', 'admin'):
            return Response({'error': 'This queue is for the Travel Help Desk.'}, status=403)

        approved = (TravelRequest.objects
                    .filter(request_type='tour_sanction',
                            status__in=['hr_approved', 'finance_approved', 'paid'])
                    .select_related('user').order_by('from_date'))
        pending, booked = [], []
        for r in approved:
            legs = r.company_booked_legs
            if not legs:
                continue
            # Anything short of booked or cancelled is still the desk's to
            # watch — quoting options, waiting on the employee's choice, or
            # ticketing what they picked.
            still_open = any(x.booking_status not in ('booked', 'cancelled') for x in legs)
            (pending if still_open else booked).append(serialize_request(r, detail=True))
        return Response({'pending': pending, 'booked': booked[:100]})


def _resolve_journey(r, journey_key):
    """journey_key -> the TravelRequest, TravelLeg or ReturnJourney it names.

    One place for this, because BookingActionView, BookingOptionsView and
    BookingSelectView all need to turn the same three-shaped key back into the
    object whose booking_* fields actually get written.
    """
    key = str(journey_key)
    if key == 'return':
        if r.trip_type != 'round_trip':
            return None, Response({'error': 'This is a one-way trip — there is no return to book.'}, status=400)
        return ReturnJourney(r), None
    if key in ('', 'trip', 'None'):
        return r, None
    leg = r.legs.filter(seq=int(key)).first() if key.isdigit() else None
    if not leg:
        return None, Response({'error': 'That stop is not part of this trip.'}, status=404)
    return leg, None


class BookingOptionsView(APIView):
    """The desk lists what it found for a journey — several flights or trains,
    not the one it has already decided on — and the employee is asked to pick.
    """
    def post(self, request, req_id):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        if u.role not in ('travel_desk', 'admin'):
            return Response({'error': 'Only the Travel Help Desk can offer ticket options.'}, status=403)

        r = TravelRequest.objects.filter(id=req_id).first()
        if not r:
            return Response({'error': 'Not found'}, status=404)
        if r.status not in ('hr_approved', 'finance_approved', 'paid'):
            return Response({'error': f'That trip is not approved yet ({_status_label(r)}).'}, status=400)

        journey_key = str(request.data.get('journey_key') or 'trip')
        target, err = _resolve_journey(r, journey_key)
        if err:
            return err
        if target.booking_mode != 'company':
            return Response({'error': 'That journey is being booked by the employee.'}, status=400)

        raw = request.data.get('options') or []
        if isinstance(raw, str):
            raw = json.loads(raw)
        options = [o for o in raw if (o.get('carrier') or o.get('detail') or o.get('amount'))]
        if len(options) < 2:
            return Response({'error': 'Add at least two options for the employee to choose between.'}, status=400)
        if len(options) > 6:
            return Response({'error': 'That is a lot of options — six is plenty for anyone to choose from.'}, status=400)
        for i, o in enumerate(options):
            if not (o.get('carrier') or '').strip():
                return Response({'error': f'Option {i + 1} needs the airline or train name.'}, status=400)
            if _sf(o.get('amount')) <= 0:
                return Response({'error': f'Option {i + 1} needs the fare.'}, status=400)

        with transaction.atomic():
            # A fresh set replaces the last — the desk is re-quoting, not adding
            # to a stale list the employee already looked at once.
            BookingOption.objects.filter(request=r, journey_key=journey_key).delete()
            for i, o in enumerate(options):
                BookingOption.objects.create(
                    request=r, journey_key=journey_key, seq=i,
                    mode=(o.get('mode') or '').strip(), carrier=(o.get('carrier') or '').strip(),
                    detail=(o.get('detail') or '').strip(), date=_parse_date(o.get('date')),
                    time=(o.get('time') or '').strip(), amount=_sf(o.get('amount')),
                    remarks=(o.get('remarks') or '').strip(), added_by=u.name)
            target.booking_status = 'options_sent'
            target.save()

        ApprovalLog.objects.create(request=r, stage='travel_desk', action='options_sent', by_name=u.name,
                                   remarks=f'{len(options)} option(s) offered.')
        notify_options_sent(r, target)
        return Response({'message': 'Options sent to the employee.', 'request': serialize_request(r, detail=True)})


class BookingSelectView(APIView):
    """The employee's turn: pick one of the desk's options."""
    def post(self, request, req_id):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)

        r = TravelRequest.objects.filter(id=req_id).first()
        if not r:
            return Response({'error': 'Not found'}, status=404)
        if r.user_id != u.id:
            return Response({'error': 'This is not your request.'}, status=403)

        journey_key = str(request.data.get('journey_key') or 'trip')
        target, err = _resolve_journey(r, journey_key)
        if err:
            return err
        if target.booking_status != 'options_sent':
            return Response({'error': 'There is nothing awaiting your confirmation for this journey.'}, status=400)

        option_id = request.data.get('option_id')
        chosen = BookingOption.objects.filter(id=option_id, request=r, journey_key=journey_key).first()
        if not chosen:
            return Response({'error': 'Choose one of the options offered.'}, status=400)

        with transaction.atomic():
            BookingOption.objects.filter(request=r, journey_key=journey_key).update(is_selected=False)
            chosen.is_selected = True
            chosen.save(update_fields=['is_selected'])
            target.booking_status = 'confirmed'
            target.save()

        ApprovalLog.objects.create(request=r, stage='employee', action='confirmed', by_name=u.name,
                                   remarks=f'Chose: {chosen.carrier} {chosen.get_time_label()}')
        notify_option_selected(r, target, chosen, u.name)
        return Response({'message': 'Choice confirmed — the desk has been notified.',
                         'request': serialize_request(r, detail=True)})


class BookingActionView(APIView):
    """Record what the desk booked — or that it could not be booked.

    When the employee has already confirmed one of the offered options, that
    option's fare and carrier are what gets booked unless the desk overrides
    them — the whole point of asking was to book what was actually chosen.
    """
    def post(self, request, req_id):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        if u.role not in ('travel_desk', 'admin'):
            return Response({'error': 'Only the Travel Help Desk can record a booking.'}, status=403)

        r = TravelRequest.objects.filter(id=req_id).first()
        if not r:
            return Response({'error': 'Not found'}, status=404)
        if r.status not in ('hr_approved', 'finance_approved', 'paid'):
            return Response({'error': f'That trip is not approved yet ({_status_label(r)}).'}, status=400)

        d = request.data
        journey_key = str(d.get('leg_seq') if d.get('leg_seq') is not None else d.get('journey_key') or 'trip')
        target, err = _resolve_journey(r, journey_key)
        if err:
            return err
        if target.booking_mode != 'company':
            return Response({'error': 'That journey is being booked by the employee.'}, status=400)

        action = (d.get('action') or 'booked').strip().lower()
        if action not in ('booked', 'cancelled'):
            return Response({'error': 'action must be booked or cancelled.'}, status=400)

        chosen = BookingOption.objects.filter(request=r, journey_key=journey_key, is_selected=True).first()

        if action == 'booked':
            ref = (d.get('booking_reference') or '').strip()
            carrier = (d.get('booking_carrier') or '').strip() or (chosen.carrier if chosen else '')
            fare_in = d.get('booking_fare')
            fare = _sf(fare_in) if fare_in not in (None, '') else (float(chosen.amount) if chosen else 0)
            if not ref:
                return Response({'error': 'A PNR or ticket number is needed to record a booking.'}, status=400)
            if fare <= 0:
                return Response({'error': 'Enter the fare actually paid.'}, status=400)
            ticket = request.FILES.get('ticket')
            if not ticket:
                return Response({'error': 'Upload the ticket — it is the record of what was actually booked.'}, status=400)
            target.booking_reference = ref
            target.booking_carrier = carrier
            target.booking_fare = fare
            target.booking_ticket.save(f'ticket_{r.id}_{journey_key}_{ticket.name}', ticket, save=False)
        target.booking_status = action
        target.booking_remarks = (d.get('booking_remarks') or '').strip()
        target.booked_by = u.name
        target.booked_at = timezone.now()
        target.save()

        ApprovalLog.objects.create(request=r, stage='travel_desk', action=action, by_name=u.name,
                                   remarks=target.booking_remarks)
        notify_booked(r, target, action, u.name)
        return Response({'message': f'Booking {action}.', 'request': serialize_request(r, detail=True)})


class ActionView(APIView):
    """Approve / reject a request at the acting user's stage."""
    def post(self, request, req_id):
        u = _get_user(request)
        if not u:
            return Response({'error': 'Login required.'}, status=401)
        r = TravelRequest.objects.filter(id=req_id).first()
        if not r:
            return Response({'error': 'Not found'}, status=404)
        action = (request.data.get('action') or '').strip().lower()   # approve / reject / paid
        remarks = request.data.get('remarks', '')

        # Same authority check the UI renders its buttons from.
        perm = action_permission(r, u)
        if not (perm['can_approve'] or perm['can_pay']):
            # Own request, wrong team and wrong role are refusals of authority;
            # being at another stage is simply not this user's turn yet.
            wrong_stage = perm['reason'] and perm['reason'].startswith('Awaiting')
            return Response({'error': perm['reason']}, status=400 if wrong_stage else 403)
        flow = _STAGE_FLOW[u.role]

        if action == 'paid' and u.role == 'finance' and r.status == 'finance_approved':
            r.status = 'paid'
            r.finance_action_at = timezone.now()
            r.save()
            ApprovalLog.objects.create(request=r, stage='finance', action='paid', by_name=u.name, remarks=remarks)
            notify_paid(r, u.name)
            return Response({'message': 'Marked as Paid.', 'request': serialize_request(r, detail=True)})

        if r.status != flow['from']:
            return Response({'error': f'Request is not awaiting your action (status: {r.get_status_display()}).'}, status=400)
        if action not in ('approve', 'reject'):
            return Response({'error': 'action must be approve or reject.'}, status=400)

        # Approving a tour programme is a judgement, so it is recorded in the
        # approver's own words. The two stages answer different questions: the
        # manager briefed the employee and owns any deviation from policy,
        # while P&C (HR) is endorsing the tour itself.
        briefing = (request.data.get('briefing') or '').strip()
        tour_justification = (request.data.get('tour_justification') or '').strip()
        advance_remarks = (request.data.get('advance_remarks') or '').strip()
        deviation = (request.data.get('deviation_justification') or '').strip()
        flags = [f for f in (r.policy_flags or '').split(NL) if f]

        missing = []
        if action == 'approve' and r.request_type == 'tour_sanction':
            if u.role == 'manager':
                if not briefing:
                    missing.append('what you briefed the employee about this programme')
                if flags and not deviation:
                    missing.append('a justification for the policy deviation on this request')
            elif u.role == 'hr':
                if not tour_justification:
                    missing.append('your justification for this tour')
            if u.role in ('manager', 'hr') and not advance_remarks:
                missing.append('your remarks on the advance')

        # A decision with no words behind it is not an audit trail. Marking a
        # settled claim as paid is bookkeeping, not a judgement, so it is exempt.
        if action in ('approve', 'reject') and not str(remarks).strip():
            missing.append('your remarks on this decision')

        if missing:
            verb = 'approving' if action == 'approve' else 'rejecting'
            return Response({'error': 'Before %s, please add: ' % verb + '; '.join(missing) + '.',
                             'missing': missing}, status=400)

        new_status = flow['approve'] if action == 'approve' else flow['reject']
        r.status = new_status
        now = timezone.now()
        if u.role == 'manager':
            r.manager_remarks = remarks; r.manager_action_at = now
        elif u.role == 'hr':
            r.hr_remarks = remarks; r.hr_action_at = now
            # P&C (HR) is the last approver of a tour programme, so the amount
            # sanctioned is settled here. A sanction has nothing *claimed* yet —
            # it is pre-travel — so what is actually authorised is the advance;
            # recording total_claimed would file every trip as approved for zero.
            if action == 'approve' and r.request_type == 'tour_sanction':
                r.total_approved = r.advance_amount
        elif u.role == 'finance':
            r.finance_remarks = remarks; r.finance_action_at = now
            if action == 'approve':
                r.total_approved = r.total_claimed
        r.save()
        # 'approve' + 'd' gives "approved", but 'reject' + 'd' gave "rejectd" —
        # a typo that has been in the audit trail all along, and which also made
        # the rejection branch below never match.
        past = 'approved' if action == 'approve' else 'rejected'
        ApprovalLog.objects.create(request=r, stage=u.role, action=past, by_name=u.name,
                                   remarks=remarks, briefing=briefing,
                                   tour_justification=tour_justification,
                                   advance_remarks=advance_remarks, deviation_justification=deviation)
        notify_actioned(r, u.role, u.name, past, remarks)
        # Only a sanctioned trip goes to the desk — booking one that is later
        # rejected wastes a fare and a cancellation charge.
        if past == 'approved' and r.needs_booking:
            notify_booking_needed(r)
        return Response({'message': f'Request {action}d.', 'request': serialize_request(r, detail=True)})
