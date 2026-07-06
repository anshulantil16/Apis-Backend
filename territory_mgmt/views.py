import uuid
import openpyxl
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from .models import OrganizationData


class TerritoryTemplateView(APIView):
    """Download Excel template for territory management upload."""

    def get(self, request):
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import FileResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Territory Data'

        # Header row
        headers = ['S.NO', 'Code', 'Name', 'Designation', 'HQ', 'STATE', 'ZONE', 'RM']
        ws.append(headers)

        # Format header
        header_fill = PatternFill(start_color='1f4788', end_color='1f4788', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Sample data
        sample_data = [
            [1, 'EMP001', 'Rajesh Kumar', 'Sales Manager', 'Delhi', 'Delhi', 'North', 'Vikram Singh'],
            [2, 'EMP002', 'Priya Sharma', 'Sales Executive', 'Mumbai', 'Maharashtra', 'West', 'Amit Patel'],
            [3, 'EMP003', 'Suresh Reddy', 'Area Manager', 'Bangalore', 'Karnataka', 'South', 'Sanjay Reddy'],
        ]

        for row in sample_data:
            ws.append(row)

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18

        # Save
        file_path = '/tmp/territory_template.xlsx'
        wb.save(file_path)

        response = FileResponse(open(file_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="territory_template.xlsx"'
        return response


class TerritoryUploadView(APIView):
    """Upload territory management data from Excel."""
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        if 'file' not in request.FILES:
            return Response({'error': 'No file provided'}, status=400)

        file = request.FILES['file']
        batch_id = str(uuid.uuid4())[:8]

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            created = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    sno, code, name, designation, hq, state, zone, rm = row[:8]

                    if not code or not name:
                        continue

                    OrganizationData.objects.update_or_create(
                        code=str(code).strip(),
                        defaults={
                            'sno': int(sno) if sno else row_idx,
                            'name': str(name).strip(),
                            'designation': str(designation).strip() if designation else '',
                            'hq': str(hq).strip() if hq else '',
                            'state': str(state).strip() if state else '',
                            'zone': str(zone).strip() if zone else '',
                            'rm': str(rm).strip() if rm else '',
                            'batch_id': batch_id,
                            'updated_at': timezone.now(),
                        }
                    )
                    created += 1

                except Exception as e:
                    errors.append(f'Row {row_idx}: {str(e)}')

            return Response({
                'message': f'Successfully uploaded {created} records',
                'created': created,
                'batch_id': batch_id,
                'errors': errors,
            })

        except Exception as e:
            return Response({
                'error': 'Upload failed',
                'detail': str(e),
            }, status=400)


class TerritoryDashboardView(APIView):
    """Analytics dashboard with multi-filter support."""

    def get(self, request):
        # Get filter params
        designations = request.query_params.getlist('designation')
        states = request.query_params.getlist('state')
        zones = request.query_params.getlist('zone')
        rms = request.query_params.getlist('rm')

        # Base queryset
        records = OrganizationData.objects.all()

        # Apply filters
        if designations:
            records = records.filter(designation__in=designations)
        if states:
            records = records.filter(state__in=states)
        if zones:
            records = records.filter(zone__in=zones)
        if rms:
            records = records.filter(rm__in=rms)

        # Prepare data
        data = [{
            'id': r.id,
            'sno': r.sno,
            'code': r.code,
            'name': r.name,
            'designation': r.designation,
            'hq': r.hq,
            'state': r.state,
            'zone': r.zone,
            'rm': r.rm,
        } for r in records]

        # Get all unique values for filters
        all_records = OrganizationData.objects.all()
        filter_options = {
            'designations': sorted(set(all_records.values_list('designation', flat=True))),
            'states': sorted(set(all_records.values_list('state', flat=True))),
            'zones': sorted(set(all_records.values_list('zone', flat=True))),
            'rms': sorted(set(all_records.values_list('rm', flat=True))),
        }

        # Analytics
        total_employees = len(data)
        unique_designations = records.values('designation').distinct().count()
        unique_zones = records.values('zone').distinct().count()
        unique_states = records.values('state').distinct().count()

        # Count by RM
        rm_count = {}
        for r in records:
            rm_count[r.rm] = rm_count.get(r.rm, 0) + 1

        # Count by Designation
        designation_count = {}
        for r in records:
            designation_count[r.designation] = designation_count.get(r.designation, 0) + 1

        # Count by Zone
        zone_count = {}
        for r in records:
            zone_count[r.zone] = zone_count.get(r.zone, 0) + 1

        # Count by State
        state_count = {}
        for r in records:
            state_count[r.state] = state_count.get(r.state, 0) + 1

        return Response({
            'summary': {
                'total_employees': total_employees,
                'unique_designations': unique_designations,
                'unique_zones': unique_zones,
                'unique_states': unique_states,
            },
            'breakdown': {
                'by_rm': rm_count,
                'by_designation': designation_count,
                'by_zone': zone_count,
                'by_state': state_count,
            },
            'filter_options': filter_options,
            'data': data,
            'record_count': len(data),
        })
