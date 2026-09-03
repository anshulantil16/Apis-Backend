"""
EOM Excel Export — full database dump for management reporting.
GET /api/eom/export/?cycle_id=<id>   (cycle_id optional — omit for all cycles)
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response

from eom.models import EOMNomination, EOMCycle
from config.tz import local_str


def _border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_fill():
    return PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')


def _subheader_fill():
    return PatternFill(start_color='2d6a9f', end_color='2d6a9f', fill_type='solid')


def _winner_fill():
    return PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')


class EOMExportView(APIView):
    """GET /api/eom/export/ — full Excel export for management."""

    def get(self, request):
        cycle_id = request.query_params.get('cycle_id')

        qs = EOMNomination.objects.select_related('employee', 'cycle').order_by(
            'cycle__year', 'cycle__month', '-is_winner', 'employee__name'
        )
        if cycle_id:
            qs = qs.filter(cycle_id=cycle_id)

        wb = Workbook()

        # ── Sheet 1: Full Nominations Data ────────────────────────────────
        ws = wb.active
        ws.title = 'All Nominations'

        COLUMNS = [
            # Basic info
            ('Cycle',                   18),
            ('Month',                   10),
            ('Year',                    8),
            ('Employee ID',             14),
            ('Employee Name',           22),
            ('Designation',             20),
            ('Department',              18),
            ('Zone',                    16),
            ('Track',                   14),
            ('Status',                  18),
            ('Winner',                  10),
            ('Submitted At',            18),

            # Employee form
            ('Achievement (Part A)',    40),
            ('SMART: Specific',         30),
            ('SMART: Measurable',       30),
            ('SMART: Achievable',       30),
            ('SMART: Relevant',         30),
            ('SMART: Timebound',        30),
            ('Evidence 1 Description',  30),
            ('Evidence 1 Source',       20),
            ('Evidence 2 Description',  30),
            ('Evidence 2 Source',       20),
            ('Support Document',        25),

            # HOD Scorecard
            ('HOD: Dim 1 Score (/50)',  18),
            ('HOD: Dim 1 Comments',     35),
            ('HOD: Recommendation',     18),
            ('HOD: Panel Name',         20),
            ('HOD: Remarks',            30),
            ('HOD: Reviewed At',        18),

            # Panel Scorecard
            ('Panel: Dim 2 – APIS Values (/10)',              20),
            ('Panel: Dim 2 Comments',                         30),
            ('Panel: Dim 3 – Survey/Compliance/Fun (/10)',    22),
            ('Panel: Dim 3 Comments',                         30),
            ('Panel: Dim 4 Score (/10)',                      18),
            ('Panel: Dim 4 Comments',                         30),
            ('Panel: Dim 5 Score (/10)',                      18),
            ('Panel: Dim 5 Comments',                         30),
            ('Panel: Dim 6 Score (/10)',                      18),
            ('Panel: Dim 6 Comments',                         30),
            ('Panel: Sustainability Bonus (/5)',               20),
            ('Panel: Sustainability Description',              35),
            ('Panel: Sustainability Justification',            35),
            ('Panel: Recommendation',                          18),
            ('Panel: Panel Name',                              20),
            ('Panel: Remarks',                                 30),
            ('Panel: Reviewed At',                             18),

            # Scores
            ('HOD Total (/50)',         14),
            ('Panel Total (/50)',       14),
            ('Sustainability Bonus',    14),
            ('GRAND TOTAL (/105)',      16),

            # HR
            ('HR: Remarks',            30),
            ('HR: Finalized At',       18),
        ]

        # Title row
        ws.merge_cells(f'A1:{get_column_letter(len(COLUMNS))}1')
        title_cell = ws['A1']
        title_cell.value = 'APIS INDIA LIMITED — Employee of the Month: Complete Nominations Data'
        title_cell.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        title_cell.fill = _header_fill()
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        # Header row
        for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            cell.fill = _subheader_fill()
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = _border()
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width
        ws.row_dimensions[2].height = 36

        # Data rows
        for row_idx, nom in enumerate(qs, 3):
            emp = nom.employee
            cyc = nom.cycle

            hod_total = nom.hod_dim1_score or 0
            panel_total = (
                (nom.panel_dim2_score or 0) +
                (nom.panel_dim3_score or 0) +
                (nom.panel_dim4_score or 0) +
                (nom.panel_dim5_score or 0) +
                (nom.panel_dim6_score or 0)
            )
            bonus = nom.panel_sustainability_bonus or 0
            grand_total = hod_total + panel_total + bonus

            row_data = [
                cyc.name,
                cyc.get_month_display() if hasattr(cyc, 'get_month_display') else cyc.month,
                cyc.year,
                emp.employee_id,
                emp.name,
                emp.designation,
                emp.department,
                emp.zone,
                nom.track,
                nom.get_status_display(),
                'YES ★' if nom.is_winner else '',
                local_str(nom.submitted_at, '%d-%m-%Y %H:%M') or '',

                nom.part_a_achievement,
                nom.smart_specific,
                nom.smart_measurable,
                nom.smart_achievable,
                nom.smart_relevant,
                nom.smart_timebound,
                nom.evidence_1_description,
                nom.evidence_1_source,
                nom.evidence_2_description,
                nom.evidence_2_source,
                nom.support_document_name or '',

                nom.hod_dim1_score,
                nom.hod_dim1_comments,
                nom.hod_recommendation,
                nom.hod_panel_name,
                nom.hod_remarks,
                local_str(nom.hod_reviewed_at, '%d-%m-%Y %H:%M') or '',

                nom.panel_dim2_score,
                nom.panel_dim2_comments,
                nom.panel_dim3_score,
                nom.panel_dim3_comments,
                nom.panel_dim4_score,
                nom.panel_dim4_comments,
                nom.panel_dim5_score,
                nom.panel_dim5_comments,
                nom.panel_dim6_score,
                nom.panel_dim6_comments,
                bonus,
                nom.panel_sustainability_desc,
                nom.panel_sustainability_just,
                nom.panel_recommendation,
                nom.panel_panel_name,
                nom.panel_remarks,
                local_str(nom.panel_reviewed_at, '%d-%m-%Y %H:%M') or '',

                hod_total,
                panel_total,
                bonus,
                grand_total,

                nom.hr_remarks,
                local_str(nom.hr_finalized_at, '%d-%m-%Y %H:%M') or '',
            ]

            fill = _winner_fill() if nom.is_winner else None
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Calibri', size=9)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = _border()
                if fill:
                    cell.fill = fill
            ws.row_dimensions[row_idx].height = 40

        # Freeze header
        ws.freeze_panes = 'A3'

        # ── Sheet 2: Winners Summary ──────────────────────────────────────
        ws2 = wb.create_sheet('Winners Summary')
        winners = [n for n in qs if n.is_winner]

        ws2.merge_cells('A1:J1')
        ws2['A1'].value = 'APIS INDIA LIMITED — Employee of the Month: Winners'
        ws2['A1'].font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        ws2['A1'].fill = PatternFill(start_color='856404', end_color='856404', fill_type='solid')
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 28

        winner_headers = ['Cycle', 'Employee ID', 'Name', 'Designation', 'Department',
                          'Zone', 'Track', 'HOD Score', 'Panel Score', 'Grand Total']
        gold = PatternFill(start_color='B8860B', end_color='B8860B', fill_type='solid')
        for ci, h in enumerate(winner_headers, 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = gold
            c.alignment = Alignment(horizontal='center')
            c.border = _border()
            ws2.column_dimensions[get_column_letter(ci)].width = 18

        for ri, nom in enumerate(winners, 3):
            hod_t = nom.hod_dim1_score or 0
            panel_t = sum(v or 0 for v in [nom.panel_dim2_score, nom.panel_dim3_score,
                                             nom.panel_dim4_score, nom.panel_dim5_score,
                                             nom.panel_dim6_score])
            bonus = nom.panel_sustainability_bonus or 0
            row = [nom.cycle.name, nom.employee.employee_id, nom.employee.name,
                   nom.employee.designation, nom.employee.department, nom.employee.zone,
                   nom.track, hod_t, panel_t + bonus, hod_t + panel_t + bonus]
            for ci, v in enumerate(row, 1):
                c = ws2.cell(row=ri, column=ci, value=v)
                c.font = Font(name='Calibri', size=9, bold=True)
                c.fill = _winner_fill()
                c.border = _border()

        # ── Sheet 3: Score Summary per Cycle ─────────────────────────────
        ws3 = wb.create_sheet('Scores by Cycle')
        ws3.merge_cells('A1:H1')
        ws3['A1'].value = 'APIS INDIA LIMITED — EOM Score Summary by Cycle'
        ws3['A1'].font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        ws3['A1'].fill = _header_fill()
        ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[1].height = 28

        score_headers = ['Cycle', 'Employee', 'Department', 'Zone',
                         'HOD Score', 'Panel Score', 'Bonus', 'Grand Total']
        for ci, h in enumerate(score_headers, 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _subheader_fill()
            c.alignment = Alignment(horizontal='center')
            c.border = _border()
            ws3.column_dimensions[get_column_letter(ci)].width = 20

        for ri, nom in enumerate(qs, 3):
            hod_t = nom.hod_dim1_score or 0
            panel_t = sum(v or 0 for v in [nom.panel_dim2_score, nom.panel_dim3_score,
                                             nom.panel_dim4_score, nom.panel_dim5_score,
                                             nom.panel_dim6_score])
            bonus = nom.panel_sustainability_bonus or 0
            row = [nom.cycle.name, nom.employee.name, nom.employee.department,
                   nom.employee.zone, hod_t, panel_t, bonus, hod_t + panel_t + bonus]
            for ci, v in enumerate(row, 1):
                c = ws3.cell(row=ri, column=ci, value=v)
                c.font = Font(name='Calibri', size=9)
                c.border = _border()
                if nom.is_winner:
                    c.fill = _winner_fill()

        # Return as Excel file download
        filename = f'EOM_Data{"_Cycle_" + cycle_id if cycle_id else "_All"}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
