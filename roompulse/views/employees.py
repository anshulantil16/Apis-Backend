"""Super Admin: employee directory — template, bulk upload, list, clear."""
import io
import openpyxl
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from ..models import Employee, AdminUser
from ..ingest import map_headers, build_template, is_admin_value
from .perms import require_role, actor_role
from .auth import SUPER_ADMIN_EMAIL


class EmployeeTemplateView(APIView):
    def get(self, request):
        buf = build_template()
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="RoomPulse_Employee_Template.xlsx"'
        return resp


class EmployeeUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        if (err := require_role(request, 'super_admin')):
            return err
        _, acting_email = actor_role(request)
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {e}'}, status=400)

        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return Response({'error': 'The sheet is empty.'}, status=400)

        col_map, unknown = map_headers(header_row)
        raw_headers = [str(c).strip() for c in header_row if c is not None and str(c).strip()]
        if 'name' not in col_map or 'email' not in col_map:
            missing = [lbl for field, lbl in (('name', 'Name'), ('email', 'Email')) if field not in col_map]
            shown = ', '.join(raw_headers[:15]) + (' …' if len(raw_headers) > 15 else '')
            return Response({
                'error': f'Missing required column(s): {", ".join(missing)}. '
                         f'Your file has: {shown or "(no headers found)"}.',
                'detected_columns': raw_headers,
            }, status=400)

        def cell(row, field):
            ci = col_map.get(field)
            if ci is None or ci >= len(row):
                return None
            v = row[ci]
            return str(v).strip() if v is not None else ''

        created = updated = skipped = admins_granted = 0
        skipped_rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(v is not None and str(v).strip() != '' for v in row):
                continue
            email = (cell(row, 'email') or '').lower()
            name = cell(row, 'name') or ''
            if not email or '@' not in email or not name:
                skipped += 1
                skipped_rows.append(row_idx)
                continue

            # Role column GRANTS admin access, never revokes it — a blank or
            # "Employee" cell on someone already admin leaves them untouched.
            # Revoking admin access stays a deliberate action in the Team tab.
            wants_admin = is_admin_value(cell(row, 'role')) and email != SUPER_ADMIN_EMAIL
            role = 'admin' if wants_admin else None  # None = "don't change existing role"

            obj, was_created = Employee.objects.update_or_create(
                email=email,
                defaults={
                    'name': name[:200],
                    'employee_code': (cell(row, 'employee_code') or '')[:50],
                    'department': (cell(row, 'department') or '')[:150],
                    'designation': (cell(row, 'designation') or '')[:150],
                    'location': (cell(row, 'location') or '')[:150],
                    'reporting_manager': (cell(row, 'reporting_manager') or '')[:200],
                    **({'role': 'admin'} if wants_admin else {}),
                },
            )
            created += was_created
            updated += not was_created

            if wants_admin:
                _, was_new_admin = AdminUser.objects.get_or_create(
                    email=email, defaults={'name': name[:200],
                                           'added_by': f'employee upload by {acting_email}'})
                admins_granted += was_new_admin

        warnings = []
        if skipped_rows:
            head = ', '.join(str(n) for n in skipped_rows[:15])
            warnings.append(f'{skipped} row(s) skipped — missing Name or a valid Email '
                            f'(sheet row {head}{"…" if len(skipped_rows) > 15 else ""}).')
        if unknown:
            warnings.append(f'{len(unknown)} column(s) not recognised: {", ".join(unknown[:10])}.')
        if admins_granted:
            warnings.append(f'{admins_granted} employee(s) granted Admin access via the Role column.')

        return Response({
            'message': f'{created} added, {updated} updated.'
                       + (f' {admins_granted} granted Admin access.' if admins_granted else ''),
            'created': created, 'updated': updated, 'skipped': skipped,
            'admins_granted': admins_granted,
            'detected_columns': raw_headers, 'warnings': warnings,
        })


class EmployeeListView(APIView):
    def get(self, request):
        from django.db.models import Q
        qs = Employee.objects.all()
        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(email__icontains=search) |
                           Q(department__icontains=search) | Q(employee_code__icontains=search))
        total = qs.count()
        try:
            limit = max(1, min(1000, int(request.query_params.get('limit', 200))))
        except (TypeError, ValueError):
            limit = 200
        results = [{
            'id': e.id, 'employee_code': e.employee_code, 'name': e.name, 'email': e.email,
            'department': e.department, 'designation': e.designation, 'location': e.location,
            'reporting_manager': e.reporting_manager, 'role': e.role,
        } for e in qs[:limit]]
        return Response({'results': results, 'count': total})

    def delete(self, request):
        if (err := require_role(request, 'super_admin')):
            return err
        n = Employee.objects.count()
        Employee.objects.all().delete()
        return Response({'message': f'Cleared {n} employee record(s).', 'deleted': n})
