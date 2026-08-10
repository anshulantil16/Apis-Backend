"""Employee-directory Excel template + tolerant upload parsing.

Same alias-matching approach as sales/ingest.py: headers are matched against
known spellings rather than required verbatim, and unrecognised columns are
reported back instead of silently dropped.
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUMN_ALIASES = {
    'employee_code':     ['employee id', 'employee code', 'emp id', 'emp code', 'code'],
    'name':               ['name', 'employee name', 'full name'],
    'email':              ['email', 'email address', 'official email'],
    'department':         ['department', 'dept'],
    'designation':        ['designation', 'title', 'job title'],
    'location':           ['location', 'work location', 'office', 'city'],
    'reporting_manager':  ['reporting manager', 'manager', 'rm'],
    'role':               ['role', 'access', 'access level', 'user role', 'admin', 'is admin'],
}

# Values in the Role column that grant Admin access on upload. Anything else
# (blank, "Employee", "No", ...) leaves the row as a plain Employee — a
# stray/blank cell can never accidentally REVOKE existing admin access via
# upload, only grant it (see EmployeeUploadView, which never demotes).
ADMIN_ROLE_VALUES = {'admin', 'administrator', 'yes', 'y', 'true', '1'}


def _norm(h):
    s = str(h or '').strip().lower().replace('*', '')
    for ch in ('_', '-', '.', '/', '(', ')', ':'):
        s = s.replace(ch, ' ')
    return ' '.join(s.split())


_LOOKUP = {}
for field, aliases in COLUMN_ALIASES.items():
    for a in aliases:
        _LOOKUP[_norm(a)] = field


def map_headers(header_row):
    col_map, unknown = {}, []
    for ci, cell in enumerate(header_row):
        if cell is None or str(cell).strip() == '':
            continue
        key = _norm(cell)
        field = _LOOKUP.get(key)
        if field is None:
            unknown.append(str(cell).strip())
        elif field not in col_map:
            col_map[field] = ci
    return col_map, unknown


def is_admin_value(v):
    """True if a Role-column cell should grant Admin access."""
    return str(v or '').strip().lower() in ADMIN_ROLE_VALUES


def build_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    headers = ['Employee ID', 'Name *', 'Email *', 'Department', 'Designation',
              'Location', 'Reporting Manager', 'Role']
    border = Border(*(Side(style='thin'),) * 4)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        required = '*' in h
        c.fill = PatternFill(start_color='6366F1' if required else 'A5B4FC',
                             end_color='6366F1' if required else 'A5B4FC', fill_type='solid')
        c.font = Font(color='FFFFFF', bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    samples = [
        ['EMP001', 'Rahul Sharma', 'rahul.sharma@apisindia.com', 'Sales', 'Sales Manager',
         'Delhi HO', 'Vikas Gupta', 'Admin'],
        ['EMP002', 'Priya Singh', 'priya.singh@apisindia.com', 'Operations', 'Executive',
         'Mumbai', 'Anita Desai', 'Employee'],
    ]
    for ri, row in enumerate(samples, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=v).border = border

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
    ws.freeze_panes = 'A2'

    # Reference sheet — same pattern as sales/ingest.py's "How To Use" tab.
    ref = wb.create_sheet('Role Column')
    ref.column_dimensions['A'].width = 26
    ref.column_dimensions['B'].width = 90
    rows = [
        ('Role column — grants Admin access', ''),
        ('', ''),
        ('Admin', 'Write "Admin" to give this person Admin access — they can approve/reject '
                  'booking requests and book rooms directly. Also accepts: Administrator, Yes, Y, '
                  'True, 1 (case-insensitive).'),
        ('Employee (default)', 'Leave blank, or write "Employee" / "No" — no change to their access.'),
        ('Important', 'This upload only GRANTS admin access, it never removes it. A blank or '
                      '"Employee" cell on someone who is already an Admin does NOT revoke them — '
                      'remove an admin from the Team tab in RoomPulse instead.'),
        ('Super Admin', 'The Super Admin account is fixed in the system and cannot be changed '
                        'via this column.'),
    ]
    for ri, (a, b) in enumerate(rows, 1):
        ref.cell(row=ri, column=1, value=a).font = Font(bold=True, size=12 if ri == 1 else 10)
        c = ref.cell(row=ri, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical='top')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
