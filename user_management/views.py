import pandas as pd
import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
import io
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def _format_cell(val):
    """Convert any pandas/Excel cell value to a clean string."""
    if val is None:
        return ''
    try:
        if pd.isna(val):
            return ''
    except (TypeError, ValueError):
        pass
    if isinstance(val, datetime.time):
        return val.strftime('%H:%M')
    if isinstance(val, datetime.datetime):
        return val.strftime('%d/%m/%Y')
    if isinstance(val, datetime.date):
        return val.strftime('%d/%m/%Y')
    s = str(val).strip()
    return '' if s in ('nan', 'NaT', 'None', 'NaN') else s


def _read_raw(file, file_name):
    """Read file into a headerless DataFrame, resetting the file cursor first."""
    file.seek(0)
    if file_name.endswith('.csv'):
        try:
            return pd.read_csv(file, header=None, dtype=str)
        except UnicodeDecodeError:
            file.seek(0)
            return pd.read_csv(file, header=None, dtype=str, encoding='latin1')
    else:
        return pd.read_excel(file, header=None)


def _parse_pocket_hrms(file, file_name):
    """
    Pocket HRMS exports have two structural quirks this function handles:

    1. Metadata rows at the top (company name, address, 'AttendanceSummary',
       'Run Date') before the real column-header row.

    2. Hierarchical / merged-cell layout for multi-date exports:
         Row N   : Code | Name | Category | Department | Designation | Date | ...
         Row N+1 :  ←blank for A-E→                                  | Date | ...
       Each employee occupies one row per date; identity columns (Code, Name,
       Category, Department, Designation) are only filled in the first row and
       left blank for subsequent dates.  We forward-fill those columns so every
       row becomes a self-contained record.

    Returns (DataFrame, error_message_or_None).
    """
    df_raw = _read_raw(file, file_name)

    # ── Step 1: find the real header row ────────────────────────────────────
    required = {'code', 'name', 'type'}
    header_idx = None
    for idx, row in df_raw.iterrows():
        vals = {str(v).strip().lower() for v in row.values if str(v).strip().lower() not in ('', 'nan')}
        if required.issubset(vals):
            header_idx = idx
            break

    if header_idx is None:
        return None, "Could not locate column headers (Code / Name / Type) in the uploaded file."

    # ── Step 2: re-read using the correct header row ─────────────────────────
    file.seek(0)
    if file_name.endswith('.csv'):
        try:
            df = pd.read_csv(file, header=header_idx)
        except UnicodeDecodeError:
            file.seek(0)
            df = pd.read_csv(file, header=header_idx, encoding='latin1')
    else:
        df = pd.read_excel(file, header=header_idx)

    df.columns = df.columns.astype(str).str.replace('\n', ' ', regex=False).str.strip()

    # ── Step 3: forward-fill identity columns ────────────────────────────────
    # In multi-date exports only the first row per employee has Code/Name/etc.;
    # the remaining date rows leave those cells blank.
    identity_col_names = {'code', 'name', 'category', 'department', 'designation'}
    identity_cols = [c for c in df.columns if c.strip().lower() in identity_col_names]
    for col in identity_cols:
        # treat empty strings and 'nan' as missing so ffill can propagate
        df[col] = df[col].replace('', pd.NA).replace('nan', pd.NA)
        df[col] = df[col].ffill()

    # ── Step 4: drop separator / empty rows (no Date value) ──────────────────
    date_col = next((c for c in df.columns if c.strip().lower() == 'date'), None)
    if date_col:
        df = df[df[date_col].apply(
            lambda x: str(x).strip().lower() not in ('', 'nan', 'nat', 'date', 'none')
        )]

    # ── Step 5: stringify all cells (handles datetime.time punch-time objects)
    for col in df.columns:
        df[col] = df[col].apply(_format_cell)

    # ── Step 6: final guard — drop any rows still without a Code ─────────────
    code_col = next((c for c in df.columns if c.strip().lower() == 'code'), None)
    if code_col:
        df = df[df[code_col].apply(lambda x: str(x).strip().lower() not in ('', 'nan'))]

    df = df.reset_index(drop=True)
    return df, None


class ExcelUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')

        tool = request.data.get('tool', 'joining')

        if not file:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            file_name = file.name.lower()

            # ── Delhi / Pocket HRMS attendance ──────────────────────────────
            if tool == 'delhi':
                df, err = _parse_pocket_hrms(file, file_name)
                if err:
                    return Response({"error": err}, status=status.HTTP_400_BAD_REQUEST)
                headers = list(df.columns)
                data = df.to_dict(orient='records')
                return Response({"message": "File processed successfully", "headers": headers, "data": data})

            # ── Generic read for all other non-medical tools ─────────────────
            if file_name.endswith('.csv'):
                try:
                    df = pd.read_csv(file)
                except UnicodeDecodeError:
                    file.seek(0)
                    df = pd.read_csv(file, encoding='latin1')
            else:
                df = pd.read_excel(file)

            df.columns = df.columns.astype(str).str.replace('\n', ' ', regex=False).str.replace('\r', '', regex=False).str.strip()

            # NOTE: this branch's column-mapping/family-extraction logic below produces
            # exactly the Medical Report columns (Sr. No / Employee Number / Name Of
            # Member / Relation / Gender / DOB / Age / Sum Insured / GPA / DOJ /
            # Designation Name / Location / Contact number / Email details), so it must
            # run for the 'medical' tool — NOT 'joining' (a separate, unrelated tool:
            # "Joining Form Processor"). Previously this was gated on 'joining' by
            # mistake, so Medical Reports silently fell through to the generic
            # unprocessed passthrough below.
            if tool != 'medical':
                df = df.fillna('')
                headers = list(df.columns)
                data = df.to_dict(orient='records')
                return Response({
                    "message": "File processed successfully",
                    "headers": headers,
                    "data": data
                })
            
            # Drop rows that are entirely blank (trailing rows are common in
            # exported sheets) BEFORE counting, so they never become phantom
            # employees in either report.
            df = df.fillna('')
            df = df[df.apply(lambda r: any(str(v).strip() != '' for v in r), axis=1)]
            rows_in_file = len(df)

            # NOTE: there is deliberately NO HR-Remarks filter here. Both
            # reports must cover EVERY employee in the uploaded sheet.
            # A previous version kept only rows whose "HR Remarks" contained
            # "Done", which silently dropped anyone with a blank/"Pending"
            # remark from both reports (and from insurance enrolment). Do not
            # re-introduce row filtering above these projections — the only
            # rows removed are fully-blank sheet rows, handled above.
            excluded_not_done = 0

            df = df.fillna('')
            
            final_rows = []
            sr_no = 1
            
            import re

            def _clean_value(v):
                """Excel numeric cells arrive as floats, so identity numbers come
                out as '9876500001.0' / '111122223333.0'. Render whole numbers as
                plain integers so Aadhaar, UAN, mobile, pincode and bank account
                numbers are exact — a trailing '.0' in a bank A/C or Aadhaar makes
                the row unusable for payroll and insurance enrolment.
                Genuine decimals (e.g. a salary) are left untouched."""
                if v is None:
                    return ''
                if isinstance(v, float):
                    if pd.isna(v):
                        return ''
                    if v.is_integer():
                        return str(int(v))
                return v

            # Columns describing somebody OTHER than the employee. get_val must
            # never read these: it returns the first NON-EMPTY match, so without
            # this guard a blank employee field silently falls through to a
            # relative's value — e.g. an employee with no Gender would inherit
            # "Gender of Child 1", or no DOB would inherit "Father Date of Birth".
            _OTHER_PERSON_WORDS = (
                'father', 'mother', 'spouse', 'wife', 'husband', 'child', 'children',
                'son', 'daughter', 'dependant', 'dependent', 'emergency', 'nominee',
                'reporting', 'manager',
            )

            def _is_other_person_col(col_lower):
                return any(re.search(r'\b' + w + r'\b', col_lower) for w in _OTHER_PERSON_WORDS)

            def get_val(row, *substrings):
                """The EMPLOYEE'S OWN value for the first matching column."""
                for col in df.columns:
                    col_lower = str(col).lower()
                    if _is_other_person_col(col_lower):
                        continue
                    for sub in substrings:
                        if re.search(r'\b' + re.escape(sub.lower()) + r'\b', col_lower):
                            val = row[col]
                            if val != '':
                                return _clean_value(val)
                return ''

            def format_date(val):
                if pd.isna(val) or val == '':
                    return ''
                try:
                    dt = pd.to_datetime(val)
                    return dt.strftime('%d-%m-%Y')
                except:
                    s = str(val).strip()
                    if ' 00:00:00' in s:
                        return s.replace(' 00:00:00', '')
                    return s

            def _first_non_empty(row, cols):
                """First non-blank value across candidate columns, in order.
                The raw form asks several fields twice (conditional sections),
                so a value may sit in any one of them."""
                for c in cols or ():
                    if c is None:
                        continue
                    v = row[c]
                    if v is None:
                        continue
                    if isinstance(v, float) and pd.isna(v):
                        continue
                    if str(v).strip() == '':
                        continue
                    return _clean_value(v)
                return ''

            def compute_age(dob_val):
                """Fallback when the sheet has no (or a blank) Age column: derive
                age in whole years from Date Of Birth, as of today."""
                if dob_val is None or (isinstance(dob_val, float) and pd.isna(dob_val)) or str(dob_val).strip() == '':
                    return ''
                try:
                    # dayfirst=True for DD-MM-YYYY sheets; warnings suppressed so a
                    # 500-row upload doesn't spam the log once per parsed cell.
                    import warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter('ignore')
                        dob = pd.to_datetime(dob_val, dayfirst=True)
                    if pd.isna(dob):
                        return ''
                    today = pd.Timestamp.now()
                    years = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
                    return years if years >= 0 else ''
                except Exception:
                    return ''

            def infer_gender(relation, explicit_gender):
                if explicit_gender and str(explicit_gender).strip() != '':
                    return explicit_gender
                rel = str(relation).strip().upper()
                if rel in ['WIFE', 'DAUGHTER', 'MOTHER']:
                    return 'FEMALE'
                if rel in ['HUSBAND', 'SON', 'FATHER']:
                    return 'MALE'
                return ''

            for idx, row in df.iterrows():
                # Extract Primary Employee Names
                first_name = get_val(row, 'first name')
                middle_name = get_val(row, 'middle name')
                last_name = get_val(row, 'last name')
                
                names = [str(n).strip() for n in [first_name, middle_name, last_name] if str(n).strip()]
                full_name = ' '.join(names) if names else get_val(row, 'name', 'employee name')
                
                emp_no = get_val(row, 'emp id', 'employee code', 'employee number', 'emp code')
                
                primary_row = {
                    "Sr. No": sr_no,
                    "Employee Number": "",
                    "Name Of Member": full_name,
                    "Relation": "SELF",
                    "Gender": get_val(row, 'gender', 'sex'),
                    "Date Of Birth": format_date(get_val(row, 'dob', 'date of birth')),
                    "Age": get_val(row, 'age') or compute_age(get_val(row, 'dob', 'date of birth')),
                    "Sum Insured": "",
                    "GPA": "",
                    "DOJ": format_date(get_val(row, 'doj', 'date of joining')),
                    "Designation Name": get_val(row, 'designation'),
                    "Location": get_val(row, 'location', 'headquarter', 'district'),
                    "Contact number": get_val(row, 'mobile', 'contact', 'phone'),
                    "Email details": get_val(row, 'email')
                }
                final_rows.append(primary_row)
                
                # Dynamically extract family members
                # Common patterns in HR forms for family.
                # A generic "Spouse" column doesn't say wife-or-husband, so derive
                # it from the EMPLOYEE'S gender — hardcoding WIFE mislabelled every
                # female employee's husband (and flipped his gender to FEMALE).
                _emp_gender = str(primary_row["Gender"]).strip().upper()
                _spouse_rel = 'HUSBAND' if _emp_gender.startswith('F') else 'WIFE'
                family_patterns = [
                    ('Spouse', _spouse_rel), ('Wife', 'WIFE'), ('Husband', 'HUSBAND'),
                    ('Child 1', 'CHILD'), ('Child 2', 'CHILD'), ('Child 3', 'CHILD'),
                    ('Son', 'SON'), ('Daughter', 'DAUGHTER'),
                    ('Father', 'FATHER'), ('Mother', 'MOTHER'),
                    ('Dependant 1', 'DEPENDANT'), ('Dependant 2', 'DEPENDANT')
                ]
                
                for pat, rel_name in family_patterns:
                    # The raw form asks some relations TWICE ("Father Name" and
                    # "Father Name 2", same for Mother) because of conditional
                    # form sections — different employees fill different ones.
                    # Collect EVERY matching column and take the first non-empty
                    # value per row. Previously each match overwrote the last, so
                    # only the final column ("Father Name 2") was ever read and
                    # anyone who filled the first one silently lost their father.
                    name_cols_f = []
                    dob_cols_f = []
                    gender_cols_f = []
                    age_cols_f = []
                    fallback_cols_f = []

                    for col in df.columns:
                        col_lower = str(col).lower()
                        pat_lower = pat.lower()

                        if re.search(r'\b' + re.escape(pat_lower) + r'\b', col_lower):
                            if 'email' in col_lower or 'mail' in col_lower:
                                continue # Skip email columns so they aren't incorrectly mapped as names
                            if 'name' in col_lower:
                                name_cols_f.append(col)
                            elif 'dob' in col_lower or 'birth' in col_lower:
                                dob_cols_f.append(col)
                            elif 'gender' in col_lower or 'sex' in col_lower:
                                gender_cols_f.append(col)
                            elif 'age' in col_lower:
                                age_cols_f.append(col)
                            else:
                                fallback_cols_f.append(col) # e.g. a bare "Spouse" column

                    if not name_cols_f:
                        name_cols_f = fallback_cols_f

                    if name_cols_f:
                        member_name = _first_non_empty(row, name_cols_f)
                        if member_name != '' and pd.notna(member_name):
                            computed_gender = infer_gender(rel_name, _first_non_empty(row, gender_cols_f))
                            computed_relation = rel_name
                            if rel_name in ['CHILD', 'DEPENDANT']:
                                if str(computed_gender).strip().upper() == 'MALE':
                                    computed_relation = 'SON'
                                elif str(computed_gender).strip().upper() == 'FEMALE':
                                    computed_relation = 'DAUGHTER'
                                    
                            family_row = {
                                "Sr. No": "",
                                "Employee Number": "",
                                "Name Of Member": str(member_name).strip(),
                                "Relation": computed_relation,
                                "Gender": computed_gender,
                                "Date Of Birth": format_date(_first_non_empty(row, dob_cols_f)),
                                "Age": _first_non_empty(row, age_cols_f) or compute_age(_first_non_empty(row, dob_cols_f)),
                                "Sum Insured": "",
                                "GPA": "",
                                "DOJ": "",
                                "Designation Name": "",
                                "Location": "",
                                "Contact number": "",
                                "Email details": ""
                            }
                            # Check if name is mistakenly an email address (e.g. they put email in the 'Son' column)
                            if '@' not in str(member_name):
                                final_rows.append(family_row)
                
                # Insert blank row to separate families
                blank_row = {k: "" for k in primary_row.keys()}
                final_rows.append(blank_row)
                
                sr_no += 1
            
            if not final_rows:
                return Response({"headers": [], "data": []})

            headers = list(final_rows[0].keys())

            # ── Sales Report: same raw upload, different column projection ──
            # Reuses the same parsed sheet so one upload can produce BOTH the
            # Medical Report (above) and this Sales Report without re-uploading.
            # The raw sheet is a Google-Forms-style joining form whose headers
            # ("Reporting Manager Name", "Bank Account Number", "Aadhar No", ...)
            # do NOT match the target column names verbatim, so matching is by
            # alias substring, not exact name. Fields the raw form has no
            # equivalent for (Employee Code, Department, City/Pincode, CTC
            # breakdown, Medical/Conveyance/Special Allowance) are left blank
            # for manual entry, per requirement.
            SALES_COLUMNS = [
                'S No', 'Employee Code', 'Name', 'Father Name', 'Designation',
                'Date of Joining', 'Date of Birth', 'UAN Number', 'Reporting Emp Name',
                'HQ.', 'STATE', 'Gender', 'Department', 'Mobile', 'PAN No', 'Aadhaar',
                'Address 1', 'City', 'Pincode', 'State', 'Official E Mail', 'Personal E Mail',
                'Employee  Name IN BANK', 'BANK NAME', 'Employee Bank Account No', 'IFSC CODE',
                'CURRENT MONTHLY CTC', 'CURRENT YEARLY CTC', 'IN WORDS', 'GROSS/CTC',
                'Medical', 'Conveyance', 'Special Allowance',
            ]
            DATE_SALES_COLS = {'Date of Joining', 'Date of Birth'}
            # target column -> ordered alias phrases to try (first match wins).
            # None = composed/generated field, handled separately below.
            SALES_ALIASES = {
                'S No': None,
                'Employee Code': ('employee code', 'emp code', 'employee id'),
                'Name': None,
                # The form asks for the father TWICE (conditional sections), so
                # list both — the first non-blank one wins per employee.
                'Father Name': ('father name', 'father name 2'),
                'Designation': ('designation',),
                'Date of Joining': ('date of joining', 'doj'),
                'Date of Birth': ('date of birth', 'dob'),
                'UAN Number': ('uan number', 'uan no'),
                'Reporting Emp Name': ('reporting manager name', 'reporting emp name'),
                'HQ.': ('headquarter', 'hq'),
                'STATE': ('state',),
                'Gender': ('gender',),
                'Department': ('department',),
                'Mobile': ('mobile number', 'mobile', 'contact number'),
                'PAN No': ('pan no', 'pan number'),
                'Aadhaar': ('aadhar no', 'aadhaar no', 'aadhar', 'aadhaar'),
                'Address 1': ('present address', 'permanent address', 'address'),
                'City': ('city',),
                'Pincode': ('pincode', 'pin code'),
                'State': ('state',),
                'Official E Mail': ('email address', 'official email'),
                'Personal E Mail': ('personal email id', 'personal email'),
                'Employee  Name IN BANK': ('your name in bank', 'name in bank'),
                'BANK NAME': ('name of bank', 'bank name'),
                'Employee Bank Account No': ('bank account number', 'bank account no', 'account number'),
                'IFSC CODE': ('ifsc code', 'ifsc'),
                'CURRENT MONTHLY CTC': None,
                'CURRENT YEARLY CTC': None,
                'IN WORDS': None,
                'GROSS/CTC': None,
                'Medical': None,
                'Conveyance': None,
                'Special Allowance': None,
            }

            def _norm(s):
                return ' '.join(str(s).replace('.', '').split()).lower()

            # Consume-from-pool matching: once a raw column is used for one
            # target, it's removed from the pool so a later target sharing the
            # same alias (e.g. two "state"-like targets, but only one raw State
            # column) can't silently reuse it — leaving the 2nd blank instead,
            # rather than the New-Department/New-Function-style clobbering bug
            # this app hit before.
            # Each target collects ALL columns matching its aliases (in alias
            # order) so duplicated form fields — "Father Name" / "Father Name 2"
            # — coalesce to the first non-blank value per employee, instead of
            # one arbitrarily winning and the other's data disappearing.
            pool = list(df.columns)  # order preserved; items removed as matched
            sales_cols_for_target = []
            for t in SALES_COLUMNS:
                aliases = SALES_ALIASES.get(t)
                if not aliases:
                    sales_cols_for_target.append([])
                    continue
                # Pass 1: exact normalised equality, collecting every alias hit.
                # Exact-only here is deliberate: a loose substring pass would let
                # "Date of Birth" also swallow "Father/Mother/Spouse Date of
                # Birth" and silently fall back to a relative's DOB.
                matches = []
                for alias in aliases:
                    for col in pool:
                        if _norm(col) == alias and col not in matches:
                            matches.append(col)
                # Pass 2: word-boundary substring — only when nothing matched
                # exactly (e.g. "Headquarter- District", "Name Of Bank (EXCEPT
                # CANARA BANK )"). Takes the single best hit, not all of them.
                if not matches:
                    for alias in aliases:
                        pat = r'\b' + re.escape(alias) + r'\b'
                        hit = next((c for c in pool if re.search(pat, _norm(c))), None)
                        if hit:
                            matches.append(hit)
                            break
                for m in matches:
                    pool.remove(m)
                sales_cols_for_target.append(matches)

            name_cols = {}
            for key, aliases in (('first', ('first name',)), ('middle', ('middle name',)), ('last', ('last name',))):
                for alias in aliases:
                    for col in df.columns:
                        if re.search(r'\b' + re.escape(alias) + r'\b', _norm(col)):
                            name_cols[key] = col
                            break
                    if key in name_cols:
                        break

            sales_rows = []
            for s_no, (idx, row) in enumerate(df.iterrows(), 1):
                out = {}
                for t, rcs in zip(SALES_COLUMNS, sales_cols_for_target):
                    if t == 'S No':
                        out[t] = s_no
                    elif t == 'Name':
                        parts = [str(row[name_cols[k]]).strip() for k in ('first', 'middle', 'last') if k in name_cols and str(row[name_cols[k]]).strip()]
                        out[t] = ' '.join(parts)
                    elif not rcs:
                        out[t] = ''
                    elif t in DATE_SALES_COLS:
                        out[t] = format_date(_first_non_empty(row, rcs))
                    else:
                        out[t] = _first_non_empty(row, rcs)
                sales_rows.append(out)

            employee_count = sum(1 for r in final_rows if str(r.get("Relation", "")).strip().upper() == "SELF")
            return Response({
                "message": "File processed successfully",
                "headers": headers,
                "data": final_rows,
                "sales_headers": SALES_COLUMNS,
                "sales_data": sales_rows,
                # Both reports are built from the same filtered rows, so
                # medical_employees and sales_employees must always be equal.
                # Surfaced so any future divergence is caught immediately
                # instead of being discovered by eyeballing two Excel files.
                "stats": {
                    "rows_in_file": rows_in_file,
                    "excluded_not_done": excluded_not_done,
                    "medical_employees": employee_count,
                    "medical_total_rows": len(final_rows),
                    "sales_employees": len(sales_rows),
                    "counts_match": employee_count == len(sales_rows),
                },
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ExcelExportView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            data = request.data.get('data', [])
            if not data:
                return Response({"error": "No data provided to export"}, status=status.HTTP_400_BAD_REQUEST)
                
            df = pd.DataFrame(data)
            
            # Create an in-memory buffer
            buffer = io.BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Extracted Data')
                
                workbook = writer.book
                worksheet = writer.sheets['Extracted Data']
                
                # Styles
                header_fill = PatternFill(start_color='333F50', end_color='333F50', fill_type='solid') # Dark Blue
                header_font = Font(color='FFFFFF', bold=True)
                center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                green_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Format Headers
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                # Find column indices dynamically
                emp_col_idx = None
                rel_col_idx = None
                for idx, col in enumerate(df.columns, start=1):
                    if str(col).strip().upper() == 'EMPLOYEE NUMBER':
                        emp_col_idx = idx
                    elif str(col).strip().upper() == 'RELATION':
                        rel_col_idx = idx
                
                # Format Data Rows
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row), start=2):
                    relation_val = ""
                    if rel_col_idx:
                        relation_val = str(worksheet.cell(row=row_idx, column=rel_col_idx).value).upper()
                    
                    for col_idx, cell in enumerate(row, start=1):
                        cell.border = thin_border
                        cell.alignment = center_alignment
                        
                        # If it's the "Employee Number" column and Relation is SELF, color it green
                        if emp_col_idx and col_idx == emp_col_idx and relation_val == 'SELF':
                            cell.fill = green_fill
                            cell.font = Font(color='000000', bold=False)
                            
                # Adjust column widths
                for col in worksheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    # Set max width to avoid extremely wide columns
                    worksheet.column_dimensions[column].width = min(adjusted_width, 40)
            
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=filtered_data.xlsx'
            
            return response
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
