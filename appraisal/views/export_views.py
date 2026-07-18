"""
Appraisal Excel Export — full database dump for management reporting.
GET /api/appraisal/export/?cycle_id=<id>
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.views import APIView

from appraisal.models import GoalCard, PerformanceCycle


def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _hfill(color='1a3a5c'):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


# ── Synopsis score helpers (mirror the exact formulas used in the role UIs) ──────
def _kpi_achievement(kpi):
    """Self-achievement % for one KPI — same logic as calcScoreAchievement() in the
    employee/manager React views (actual vs target, by parameter direction)."""
    try:
        plan = float(kpi.target_value)
        actual = float(kpi.actual_achievement)
    except (TypeError, ValueError):
        return None
    if plan == 0:
        return None
    d = (kpi.parameter_type or '').lower()
    if 'higher' in d:
        return (actual / plan) * 100
    if 'lower' in d:
        return None if actual == 0 else (plan / actual) * 100
    if 'target' in d:
        return max(0.0, (actual / plan) * 100)
    return None


def _self_achievement_total(gc):
    """Weighted self-achievement out of 100 = Σ (achievement% / 100 × weightage)."""
    total, seen = 0.0, False
    for kra in gc.goals.all():
        for kpi in kra.kpis.all():
            sc = _kpi_achievement(kpi)
            wt = kpi.weightage or 0
            if sc is not None and wt:
                total += (sc / 100.0) * wt
                seen = True
    return round(total, 2) if seen else None


def _sum_kpi_field(gc, field):
    """Σ of a per-KPI score field (manager_score / hod_score) → out of 100.
    Matches totalManagerScore / totalHODScore in the manager & HOD views."""
    total, seen = 0.0, False
    for kra in gc.goals.all():
        for kpi in kra.kpis.all():
            v = getattr(kpi, field, None)
            if v is not None and v != '':
                try:
                    total += float(v)
                    seen = True
                except (TypeError, ValueError):
                    pass
    return round(total, 2) if seen else None


_BAND_FILL = {
    'outstanding': 'C6EFCE', 'exceeds': 'DDEBF7', 'meets': 'FFF2CC',
    'below': 'FCE4D6', 'poor': 'F8CBAD',
}


class AppraisalExportView(APIView):
    """GET /api/appraisal/export/ — full Excel export for management."""

    def get(self, request):
        cycle_id = request.query_params.get('cycle_id')

        qs = GoalCard.objects.select_related(
            'employee', 'cycle'
        ).prefetch_related(
            'goals__kpis', 'competency_ratings', 'review', 'approval_logs'
        ).order_by('cycle__fiscal_year', 'cycle__quarter', 'employee__name')

        if cycle_id:
            qs = qs.filter(cycle_id=cycle_id)

        wb = Workbook()

        # ── Sheet 1: Employee Summary ─────────────────────────────────────
        ws = wb.active
        ws.title = 'Employee Summary'

        SUMMARY_COLS = [
            ('Cycle',                    18),
            ('Employee ID',              14),
            ('Name',                     22),
            ('Designation',              20),
            ('Department',               18),
            ('Zone',                     16),
            ('Sub Zone',                 16),
            ('Manager ID',               14),
            ('HOD ID',                   14),
            ('Status',                   18),
            ('Submitted At',             18),
            ('Manager Reviewed At',      20),
            ('HOD Reviewed At',          20),
            # Self review
            ('Training Programs',        30),
            ('Feedback to Manager',      30),
            ('Feedback Rating (Manager)',20),
            ('Feedback to Org',          30),
            ('Feedback Rating (Org)',    18),
            # Manager remarks
            ('Manager: Special Achievements', 35),
            ('Manager: Promoted?',       14),
            ('Manager: Promotion Justification', 35),
            ('Manager: Salary Correction', 30),
            ('Manager: Salary Justification', 30),
            # HOD remarks
            ('HOD: Remarks',             30),
            ('HOD: Special Achievements',35),
            ('HOD: Promoted?',           14),
            ('HOD: Promotion Justification', 35),
            ('HOD: Salary Correction',   30),
            ('HOD: Salary Justification',30),
            # Review
            ('Final Score',              14),
            ('Performance Band',         18),
            ('HR Remarks',               30),
        ]

        # Title
        ws.merge_cells(f'A1:{get_column_letter(len(SUMMARY_COLS))}1')
        t = ws['A1']
        t.value = 'APIS INDIA LIMITED — Appraisal Hub: Employee Summary'
        t.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        t.fill = _hfill()
        t.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(SUMMARY_COLS, 1):
            c = ws.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _hfill('2d6a9f')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = _thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 36

        for ri, gc in enumerate(qs, 3):
            emp = gc.employee
            review = getattr(gc, 'review', None)
            row = [
                gc.cycle.name, emp.employee_id, emp.name, emp.designation,
                emp.department, emp.zone, emp.subzone,
                emp.reporting_manager_id, emp.hod_id,
                gc.get_status_display(),
                gc.submitted_at.strftime('%d-%m-%Y %H:%M') if gc.submitted_at else '',
                gc.manager_reviewed_at.strftime('%d-%m-%Y %H:%M') if gc.manager_reviewed_at else '',
                gc.hod_reviewed_at.strftime('%d-%m-%Y %H:%M') if gc.hod_reviewed_at else '',
                gc.training_programs,
                gc.feedback_manager,
                gc.feedback_manager_rating,
                gc.feedback_organization,
                gc.feedback_organization_rating,
                gc.manager_special_achievements,
                gc.manager_promoted,
                gc.manager_promoted_justification,
                gc.manager_salary_correction,
                gc.manager_salary_justification,
                gc.hod_remarks,
                gc.hod_special_achievements,
                gc.hod_promoted,
                gc.hod_promoted_justification,
                gc.hod_salary_correction,
                gc.hod_salary_justification,
                float(review.final_weighted_score) if review and review.final_weighted_score else '',
                review.get_performance_band_display() if review and review.performance_band else '',
                review.hr_comments if review else '',
            ]
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = Font(name='Calibri', size=9)
                c.alignment = Alignment(vertical='top', wrap_text=True)
                c.border = _thin_border()
            ws.row_dimensions[ri].height = 30

        ws.freeze_panes = 'A3'

        # ── Sheet 2: KRAs and KPIs ────────────────────────────────────────
        ws2 = wb.create_sheet('KRAs & KPIs')
        KPI_COLS = [
            ('Cycle',20), ('Employee ID',14), ('Name',22), ('Department',18),
            ('KRA Category',18), ('KRA Title',30), ('KPI Metric',30),
            ('Target Value',18), ('Weightage',12), ('Frequency',14),
            ('Unit of Measurement',18), ('Data Source',20),
            ('Actual Achievement',20), ('Self Completion %',16),
            ('Self Rating',12), ('Self Comments',30),
            ('Manager Rating',14), ('Manager Comments',30),
            ('Manager Score',14), ('HOD Score',12),
            ('HR Rating',12), ('HR Comments',30),
            ('Final Score',12),
        ]

        ws2.merge_cells(f'A1:{get_column_letter(len(KPI_COLS))}1')
        t2 = ws2['A1']
        t2.value = 'APIS INDIA LIMITED — Appraisal Hub: KRAs & KPIs Detail'
        t2.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        t2.fill = _hfill()
        t2.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(KPI_COLS, 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _hfill('2d6a9f')
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = _thin_border()
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[2].height = 36

        ri = 3
        for gc in qs:
            emp = gc.employee
            for kra in gc.goals.all():
                for kpi in kra.kpis.all():
                    row = [
                        gc.cycle.name, emp.employee_id, emp.name, emp.department,
                        kra.category, kra.title, kpi.metric,
                        kpi.target_value, kpi.weightage, kpi.frequency,
                        kpi.unit_of_measurement, kpi.data_source,
                        kpi.actual_achievement,
                        kpi.self_completion_pct, kpi.self_rating, kpi.self_comments,
                        kpi.manager_rating, kpi.manager_comments, kpi.manager_score,
                        kpi.hod_score, kpi.hr_rating, kpi.hr_comments,
                        float(kpi.final_score) if kpi.final_score else '',
                    ]
                    for ci, v in enumerate(row, 1):
                        c = ws2.cell(row=ri, column=ci, value=v)
                        c.font = Font(name='Calibri', size=9)
                        c.alignment = Alignment(vertical='top', wrap_text=True)
                        c.border = _thin_border()
                    ws2.row_dimensions[ri].height = 25
                    ri += 1

        ws2.freeze_panes = 'A3'

        # ── Sheet 3: Competency Ratings ───────────────────────────────────
        ws3 = wb.create_sheet('Competency Ratings')
        COMP_COLS = [
            ('Cycle',20), ('Employee ID',14), ('Name',22), ('Department',18),
            ('Competency',25), ('Marks',10), ('Manager Remarks',35),
        ]

        ws3.merge_cells(f'A1:{get_column_letter(len(COMP_COLS))}1')
        t3 = ws3['A1']
        t3.value = 'APIS INDIA LIMITED — Appraisal Hub: Competency Ratings'
        t3.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        t3.fill = _hfill()
        t3.alignment = Alignment(horizontal='center', vertical='center')
        ws3.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(COMP_COLS, 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _hfill('2d6a9f')
            c.alignment = Alignment(horizontal='center')
            c.border = _thin_border()
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[2].height = 36

        ri = 3
        for gc in qs:
            emp = gc.employee
            for cr in gc.competency_ratings.all():
                row = [gc.cycle.name, emp.employee_id, emp.name, emp.department,
                       cr.get_competency_display(), cr.marks, cr.manager_remarks]
                for ci, v in enumerate(row, 1):
                    c = ws3.cell(row=ri, column=ci, value=v)
                    c.font = Font(name='Calibri', size=9)
                    c.border = _thin_border()
                ws3.row_dimensions[ri].height = 22
                ri += 1

        ws3.freeze_panes = 'A3'

        # ── Sheet 4: Approval Audit Trail ─────────────────────────────────
        ws4 = wb.create_sheet('Audit Trail')
        AUDIT_COLS = [
            ('Cycle',20), ('Employee ID',14), ('Name',22),
            ('Role',14), ('Actor Name',22), ('Action',18),
            ('Comment',40), ('Timestamp',20),
        ]

        ws4.merge_cells(f'A1:{get_column_letter(len(AUDIT_COLS))}1')
        t4 = ws4['A1']
        t4.value = 'APIS INDIA LIMITED — Appraisal Hub: Approval Audit Trail'
        t4.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        t4.fill = _hfill()
        t4.alignment = Alignment(horizontal='center', vertical='center')
        ws4.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(AUDIT_COLS, 1):
            c = ws4.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _hfill('2d6a9f')
            c.alignment = Alignment(horizontal='center')
            c.border = _thin_border()
            ws4.column_dimensions[get_column_letter(ci)].width = w
        ws4.row_dimensions[2].height = 36

        ri = 3
        for gc in qs:
            emp = gc.employee
            for log in gc.approval_logs.all():
                row = [gc.cycle.name, emp.employee_id, emp.name,
                       log.actor_role, log.actor_name, log.action,
                       log.comment, log.timestamp.strftime('%d-%m-%Y %H:%M')]
                for ci, v in enumerate(row, 1):
                    c = ws4.cell(row=ri, column=ci, value=v)
                    c.font = Font(name='Calibri', size=9)
                    c.border = _thin_border()
                ws4.row_dimensions[ri].height = 22
                ri += 1

        ws4.freeze_panes = 'A3'

        # ── Sheet 5: Leaderboard ──────────────────────────────────────────
        ws5 = wb.create_sheet('Leaderboard')
        LB_COLS = [
            ('Rank',8), ('Cycle',18), ('Employee ID',14), ('Name',22),
            ('Designation',20), ('Department',18), ('Zone',16),
            ('Final Score',14), ('Performance Band',20),
        ]

        ws5.merge_cells(f'A1:{get_column_letter(len(LB_COLS))}1')
        t5 = ws5['A1']
        t5.value = 'APIS INDIA LIMITED — Appraisal Hub: Leaderboard'
        t5.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        t5.fill = _hfill()
        t5.alignment = Alignment(horizontal='center', vertical='center')
        ws5.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(LB_COLS, 1):
            c = ws5.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _hfill('2d6a9f')
            c.alignment = Alignment(horizontal='center')
            c.border = _thin_border()
            ws5.column_dimensions[get_column_letter(ci)].width = w
        ws5.row_dimensions[2].height = 36

        ranked = sorted(
            [gc for gc in qs if hasattr(gc, 'review') and gc.review and gc.review.final_weighted_score],
            key=lambda g: float(g.review.final_weighted_score), reverse=True
        )
        for ri, gc in enumerate(ranked, 3):
            row = [
                ri - 2, gc.cycle.name, gc.employee.employee_id, gc.employee.name,
                gc.employee.designation, gc.employee.department, gc.employee.zone,
                float(gc.review.final_weighted_score),
                gc.review.get_performance_band_display(),
            ]
            medal_fill = None
            rank = ri - 2
            if rank == 1:
                medal_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
            elif rank == 2:
                medal_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            elif rank == 3:
                medal_fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')

            for ci, v in enumerate(row, 1):
                c = ws5.cell(row=ri, column=ci, value=v)
                c.font = Font(name='Calibri', size=9, bold=(rank <= 3))
                c.border = _thin_border()
                if medal_fill:
                    c.fill = medal_fill

        ws5.freeze_panes = 'A3'

        # ── Sheet 0: Synopsis (Score Extract) — inserted as the FIRST tab ─────
        ws0 = wb.create_sheet('Synopsis (Score Extract)', 0)
        SYN_COLS = [
            ('Cycle',                   18),
            ('Employee ID',             14),
            ('Name',                    22),
            ('Designation',             20),
            ('Department',              18),
            ('Zone',                    16),
            ('Sub Zone',                16),
            ('Manager ID',              13),
            ('HOD ID',                  13),
            ('Status',                  18),
            ('Total KRAs',              10),
            ('Total KPIs',              10),
            ('Total Weightage %',       14),
            ('Self Achievement /100',   16),
            ('Manager Score /100',      15),
            ('HOD Score /100',          14),
            ('Mgr vs HOD Δ',       12),
            ('Final Weighted Score',    16),
            ('Performance Band',        18),
            ('Competency Total',        13),
            ('Competency Avg',          12),
            ('Mgr Promotion',           12),
            ('HOD Promotion',           12),
            ('Mgr Salary Correction',   20),
            ('HOD Salary Correction',   20),
            ('Feedback→Mgr Rating', 14),
            ('Feedback→Org Rating', 14),
            ('Submitted At',            18),
            ('HOD Reviewed At',         18),
        ]

        ws0.merge_cells(f'A1:{get_column_letter(len(SYN_COLS))}1')
        t0 = ws0['A1']
        t0.value = 'APIS INDIA LIMITED — Appraisal Hub: Synopsis (Final Score Extract, out of 100)'
        t0.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
        t0.fill = _hfill()
        t0.alignment = Alignment(horizontal='center', vertical='center')
        ws0.row_dimensions[1].height = 28

        for ci, (h, w) in enumerate(SYN_COLS, 1):
            c = ws0.cell(row=2, column=ci, value=h)
            c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
            c.fill = _hfill('2d6a9f')
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = _thin_border()
            ws0.column_dimensions[get_column_letter(ci)].width = w
        ws0.row_dimensions[2].height = 40

        for ri, gc in enumerate(qs, 3):
            emp = gc.employee
            review = getattr(gc, 'review', None)

            kras = gc.goals.all()
            n_kras = len(kras)
            n_kpis = sum(len(kra.kpis.all()) for kra in kras)
            total_wt = round(gc.total_weightage, 2) if gc.total_weightage else 0

            self_score = _self_achievement_total(gc)
            mgr_score = _sum_kpi_field(gc, 'manager_score')
            hod_score = _sum_kpi_field(gc, 'hod_score')
            mgr_hod_delta = (round(mgr_score - hod_score, 2)
                             if mgr_score is not None and hod_score is not None else '')
            final_score = (float(review.final_weighted_score)
                           if review and review.final_weighted_score is not None else '')
            band = review.get_performance_band_display() if review and review.performance_band else ''

            comp_marks = [cr.marks for cr in gc.competency_ratings.all() if cr.marks is not None]
            comp_total = sum(comp_marks) if comp_marks else ''
            comp_avg = round(sum(comp_marks) / len(comp_marks), 2) if comp_marks else ''

            row = [
                gc.cycle.name, emp.employee_id, emp.name, emp.designation,
                emp.department, emp.zone, emp.subzone,
                emp.reporting_manager_id, emp.hod_id, gc.get_status_display(),
                n_kras, n_kpis, total_wt,
                self_score if self_score is not None else '',
                mgr_score if mgr_score is not None else '',
                hod_score if hod_score is not None else '',
                mgr_hod_delta,
                final_score, band,
                comp_total, comp_avg,
                gc.manager_promoted or '', gc.hod_promoted or '',
                gc.manager_salary_correction or '', gc.hod_salary_correction or '',
                gc.feedback_manager_rating if gc.feedback_manager_rating else '',
                gc.feedback_organization_rating if gc.feedback_organization_rating else '',
                gc.submitted_at.strftime('%d-%m-%Y %H:%M') if gc.submitted_at else '',
                gc.hod_reviewed_at.strftime('%d-%m-%Y %H:%M') if gc.hod_reviewed_at else '',
            ]
            for ci, v in enumerate(row, 1):
                c = ws0.cell(row=ri, column=ci, value=v)
                c.font = Font(name='Calibri', size=9,
                              bold=ci in (14, 15, 16, 18))  # emphasise the score columns
                c.alignment = Alignment(vertical='center', wrap_text=True,
                                        horizontal='center' if ci >= 11 else 'left')
                c.border = _thin_border()
            # tint the Performance Band cell by band
            if review and review.performance_band and review.performance_band in _BAND_FILL:
                bc = ws0.cell(row=ri, column=19)
                bc.fill = PatternFill(start_color=_BAND_FILL[review.performance_band],
                                      end_color=_BAND_FILL[review.performance_band], fill_type='solid')
            ws0.row_dimensions[ri].height = 26

        ws0.freeze_panes = 'D3'

        # Return response
        filename = f'Appraisal_Data{"_Cycle_" + cycle_id if cycle_id else "_All"}.xlsx'
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
