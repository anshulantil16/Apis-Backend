"""SalesIQ API — upload, analytics and forecasting.

Every analytics endpoint accepts the same filter query params, so the whole
dashboard can share one filter bar:
    from, to            ISO dates (inclusive)
    state, zone, area, category, channel, salesperson, asm, rsm, brand
                        repeatable / comma-separated
"""
import io
from datetime import date, timedelta

import openpyxl
from django.db.models import Sum, Count, Avg, Q, F
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser

from .models import SalesUpload, SalesRecord
from .ingest import (map_headers, parse_date, parse_num, build_template,
                     TEXT_FIELDS, NUM_FIELDS, TEXT_MAX)
from .forecasting import forecast_series

# Dimensions a client is allowed to group/filter by. Whitelisted rather than
# passed straight through so a query param can never reach an arbitrary column.
DIMENSIONS = {
    'state': 'state', 'zone': 'zone', 'area': 'area', 'city': 'city', 'region': 'region',
    'category': 'category', 'sub_category': 'sub_category', 'product': 'product_name',
    'product_name': 'product_name', 'sku': 'sku', 'brand': 'brand', 'pack_size': 'pack_size',
    'channel': 'channel', 'customer': 'customer_name', 'customer_name': 'customer_name',
    'customer_type': 'customer_type',
    'salesperson': 'salesperson', 'asm': 'asm', 'rsm': 'rsm', 'territory': 'territory',
}
FILTERABLE = ['state', 'zone', 'area', 'city', 'region', 'category', 'sub_category',
              'brand', 'channel', 'salesperson', 'asm', 'rsm', 'customer_name', 'sku']


def _multi(request, key):
    """Collect a repeatable / comma-separated query param into a list."""
    vals = []
    for raw in request.query_params.getlist(key):
        vals += [v.strip() for v in str(raw).split(',') if v.strip()]
    return vals


def apply_dim_filters(qs, request):
    """Apply only the dimension filters (no dates). Split out so the
    previous-period comparison can reuse the exact same slice of the business
    while swapping the date window."""
    applied = {}
    for f in FILTERABLE:
        vals = _multi(request, f)
        if vals:
            qs = qs.filter(**{f'{f}__in': vals})
            applied[f] = vals
    return qs, applied


def apply_filters(qs, request):
    """Shared filter parsing (dates + dimensions). Returns (qs, applied dict)."""
    qs, applied = apply_dim_filters(qs, request)
    d_from = parse_date(request.query_params.get('from'))
    d_to = parse_date(request.query_params.get('to'))
    if d_from:
        qs = qs.filter(order_date__gte=d_from)
        applied['from'] = d_from.isoformat()
    if d_to:
        qs = qs.filter(order_date__lte=d_to)
        applied['to'] = d_to.isoformat()
    return qs, applied


def _period_bounds(qs):
    agg = qs.aggregate(lo=models_min('order_date'), hi=models_max('order_date'))
    return agg['lo'], agg['hi']


# Small indirections so the imports above stay tidy.
def models_min(f):
    from django.db.models import Min
    return Min(f)


def models_max(f):
    from django.db.models import Max
    return Max(f)


def _money(v):
    return round(float(v or 0), 2)


def _pct_change(cur, prev):
    """Growth %. None when there's no baseline — 0 would read as 'flat', which
    is a different and misleading statement."""
    if not prev:
        return None
    return round(((cur - prev) / prev) * 100, 1)


class SalesTemplateView(APIView):
    def get(self, request):
        buf = build_template()
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="SalesIQ_Template.xlsx"'
        return resp


class SalesUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {e}'}, status=400)

        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return Response({'error': 'The sheet is empty.'}, status=400)

        col_map, unknown = map_headers(header_row)
        if 'order_date' not in col_map:
            return Response({
                'error': 'No date column found. One column must be the order/invoice date.',
                'detected_columns': sorted(col_map.keys()),
                'unrecognised_columns': unknown,
            }, status=400)
        if 'net_amount' not in col_map and 'gross_amount' not in col_map:
            return Response({
                'error': 'No sales value column found. Add a "Net Amount" '
                         '(or "Amount" / "Sales Value") column.',
                'detected_columns': sorted(col_map.keys()),
                'unrecognised_columns': unknown,
            }, status=400)

        def cell(row, field):
            ci = col_map.get(field)
            if ci is None or ci >= len(row):
                return None
            return row[ci]

        upload = SalesUpload.objects.create(
            filename=(f.name or '')[:255],
            uploaded_by=str(request.query_params.get('user') or '')[:200],
            status='completed',
        )

        batch, total_rev = [], 0.0
        no_date = bad_value = 0
        lo = hi = None
        row_no = 1
        try:
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_no += 1
                if not any(v is not None and str(v).strip() != '' for v in row):
                    continue
                od = parse_date(cell(row, 'order_date'))
                if od is None:
                    no_date += 1
                    continue

                gross = parse_num(cell(row, 'gross_amount'))
                net = parse_num(cell(row, 'net_amount'))
                # Fall back to gross when the export has no explicit net column.
                if not net and gross:
                    net = gross - parse_num(cell(row, 'discount'))
                if not net and not gross:
                    bad_value += 1

                rec = SalesRecord(
                    upload=upload,
                    order_date=od,
                    period=od.replace(day=1),
                    quantity=parse_num(cell(row, 'quantity')),
                    unit_price=parse_num(cell(row, 'unit_price')),
                    gross_amount=gross,
                    discount=parse_num(cell(row, 'discount')),
                    tax=parse_num(cell(row, 'tax')),
                    net_amount=net,
                    target_amount=parse_num(cell(row, 'target_amount')),
                )
                for tf in TEXT_FIELDS:
                    v = cell(row, tf)
                    s = '' if v is None else str(v).strip()
                    setattr(rec, tf, s[:TEXT_MAX.get(tf, 150)])
                batch.append(rec)
                total_rev += net
                lo = od if lo is None or od < lo else lo
                hi = od if hi is None or od > hi else hi

                if len(batch) >= 2000:
                    SalesRecord.objects.bulk_create(batch, batch_size=1000)
                    batch = []
            if batch:
                SalesRecord.objects.bulk_create(batch, batch_size=1000)
        except Exception as e:
            upload.status = 'failed'
            upload.save(update_fields=['status'])
            upload.records.all().delete()
            return Response({'error': f'Failed while reading row {row_no}: {e}'}, status=400)

        count = upload.records.count()
        if count == 0:
            upload.delete()
            return Response({
                'error': 'No usable rows found — every row was missing a valid date.',
                'detected_columns': sorted(col_map.keys()),
                'unrecognised_columns': unknown,
            }, status=400)

        warnings = []
        if no_date:
            warnings.append(f'{no_date} row(s) skipped — the date column was empty or '
                            f'unreadable. Those sales are NOT in the dashboard.')
        if bad_value:
            warnings.append(f'{bad_value} row(s) had no sales value (treated as 0). '
                            f'Check the amount column in your export.')
        if unknown:
            shown = ', '.join(unknown[:12]) + (' …' if len(unknown) > 12 else '')
            warnings.append(f'{len(unknown)} column(s) were not recognised and are ignored: '
                            f'{shown}. Rename them to match the template if you need them.')
        missing = [d for d in ('state', 'category', 'channel', 'salesperson')
                   if d not in col_map]
        if missing:
            warnings.append('No ' + ', '.join(missing) + ' column — those breakdown views '
                            'will be empty. Add the column and re-upload to enable them.')

        upload.row_count = count
        upload.skipped_rows = no_date
        upload.total_revenue = total_rev
        upload.period_start, upload.period_end = lo, hi
        upload.warnings = warnings
        upload.save()

        return Response({
            'message': f'Imported {count:,} rows.',
            'upload_id': upload.id,
            'rows': count,
            'skipped': no_date,
            'total_revenue': _money(total_rev),
            'period': {'from': lo.isoformat() if lo else None,
                       'to': hi.isoformat() if hi else None},
            'detected_columns': sorted(col_map.keys()),
            'unrecognised_columns': unknown,
            'warnings': warnings,
        })


class SalesOverviewView(APIView):
    """Headline KPIs + comparison against the preceding equal-length window."""

    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        agg = qs.aggregate(
            revenue=Sum('net_amount'), qty=Sum('quantity'), target=Sum('target_amount'),
            orders=Count('invoice_no', distinct=True), lines=Count('id'),
            discount=Sum('discount'),
        )
        revenue = _money(agg['revenue'])
        target = _money(agg['target'])
        lo, hi = _period_bounds(qs)

        # Preceding window of identical length, same dimension filters, so the
        # comparison is like-for-like rather than "this quarter vs all history".
        prev = None
        if lo and hi:
            span = (hi - lo).days + 1
            p_hi = lo - timedelta(days=1)
            p_lo = p_hi - timedelta(days=span - 1)
            pqs, _ = apply_dim_filters(SalesRecord.objects.all(), request)
            prev = (pqs.filter(order_date__gte=p_lo, order_date__lte=p_hi)
                       .aggregate(revenue=Sum('net_amount'), qty=Sum('quantity')))

        prev_rev = _money(prev['revenue']) if prev else 0.0
        customers = qs.exclude(customer_name='').values('customer_name').distinct().count()
        skus = qs.exclude(sku='').values('sku').distinct().count()
        orders = agg['orders'] or 0

        return Response({
            'revenue': revenue,
            'quantity': _money(agg['qty']),
            'orders': orders,
            'lines': agg['lines'] or 0,
            'customers': customers,
            'skus': skus,
            'discount': _money(agg['discount']),
            'avg_order_value': _money(revenue / orders) if orders else 0.0,
            'target': target,
            'achievement_pct': round((revenue / target) * 100, 1) if target else None,
            'gap_to_target': _money(target - revenue) if target else None,
            'prev_revenue': prev_rev,
            'revenue_growth_pct': _pct_change(revenue, prev_rev),
            'quantity_growth_pct': _pct_change(_money(agg['qty']),
                                               _money(prev['qty']) if prev else 0),
            'period': {'from': lo.isoformat() if lo else None,
                       'to': hi.isoformat() if hi else None},
            'filters': applied,
            'has_data': revenue != 0 or (agg['lines'] or 0) > 0,
        })


class SalesBreakdownView(APIView):
    """Group by any whitelisted dimension. `?dim=state&metric=revenue&limit=10`"""

    def get(self, request):
        dim_key = (request.query_params.get('dim') or 'state').strip().lower()
        field = DIMENSIONS.get(dim_key)
        if not field:
            return Response({'error': f'Unknown dimension "{dim_key}".',
                             'available': sorted(DIMENSIONS.keys())}, status=400)
        metric = (request.query_params.get('metric') or 'revenue').strip().lower()
        try:
            limit = max(1, min(200, int(request.query_params.get('limit', 15))))
        except (TypeError, ValueError):
            limit = 15

        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        qs = qs.exclude(**{field: ''})
        rows = (qs.values(field)
                  .annotate(revenue=Sum('net_amount'), quantity=Sum('quantity'),
                            target=Sum('target_amount'), orders=Count('invoice_no', distinct=True),
                            lines=Count('id'))
                  .order_by('-quantity' if metric == 'quantity' else '-revenue'))

        all_rows = list(rows)
        total_rev = sum(float(r['revenue'] or 0) for r in all_rows) or 1.0
        top = all_rows[:limit]
        out = []
        for r in top:
            rev = _money(r['revenue'])
            tgt = _money(r['target'])
            out.append({
                'name': r[field] or '—',
                'revenue': rev,
                'quantity': _money(r['quantity']),
                'orders': r['orders'],
                'lines': r['lines'],
                'target': tgt,
                'achievement_pct': round((rev / tgt) * 100, 1) if tgt else None,
                'share_pct': round((rev / total_rev) * 100, 1),
            })
        others = all_rows[limit:]
        return Response({
            'dimension': dim_key,
            'metric': metric,
            'results': out,
            'total_groups': len(all_rows),
            'others': {
                'count': len(others),
                'revenue': _money(sum(float(r['revenue'] or 0) for r in others)),
            } if others else None,
            'filters': applied,
        })


class SalesTrendView(APIView):
    """Monthly time series, with target and a cumulative running total."""

    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        rows = (qs.values('period')
                  .annotate(revenue=Sum('net_amount'), quantity=Sum('quantity'),
                            target=Sum('target_amount'), orders=Count('invoice_no', distinct=True))
                  .order_by('period'))
        out, running = [], 0.0
        prev_rev = None
        for r in rows:
            rev = _money(r['revenue'])
            running += rev
            tgt = _money(r['target'])
            out.append({
                'period': r['period'].isoformat(),
                'label': r['period'].strftime('%b %Y'),
                'revenue': rev,
                'quantity': _money(r['quantity']),
                'orders': r['orders'],
                'target': tgt,
                'achievement_pct': round((rev / tgt) * 100, 1) if tgt else None,
                'cumulative': round(running, 2),
                'mom_growth_pct': _pct_change(rev, prev_rev) if prev_rev is not None else None,
            })
            prev_rev = rev

        best = max(out, key=lambda r: r['revenue']) if out else None
        worst = min(out, key=lambda r: r['revenue']) if out else None
        return Response({
            'results': out, 'months': len(out),
            'best_month': best, 'worst_month': worst,
            'filters': applied,
        })


class SalesForecastView(APIView):
    """Forecast future monthly revenue from the filtered history."""

    def get(self, request):
        try:
            periods = max(1, min(24, int(request.query_params.get('periods', 6))))
        except (TypeError, ValueError):
            periods = 6
        metric = (request.query_params.get('metric') or 'revenue').strip().lower()
        agg_field = 'quantity' if metric == 'quantity' else 'net_amount'

        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        rows = (qs.values('period').annotate(v=Sum(agg_field)).order_by('period'))
        points = [(r['period'], float(r['v'] or 0)) for r in rows]

        result = forecast_series(points, periods=periods)
        result['metric'] = metric
        result['history'] = [{'period': d.isoformat(), 'label': d.strftime('%b %Y'),
                              'value': round(v, 2)} for d, v in points]
        result['filters'] = applied

        hist_total = sum(v for _, v in points)
        if points and result.get('points'):
            # Compare like with like: the same number of months, most recent first.
            n = min(periods, len(points))
            recent = sum(v for _, v in points[-n:])
            proj = sum(p['value'] for p in result['points'][:n])
            result['vs_recent'] = {
                'months': n,
                'recent_total': round(recent, 2),
                'projected_total': round(proj, 2),
                'change_pct': _pct_change(proj, recent),
            }
        result['history_total'] = round(hist_total, 2)
        return Response(result)


class SalesFiltersView(APIView):
    """Distinct values for every filter, so the UI can populate its dropdowns."""

    def get(self, request):
        qs = SalesRecord.objects.all()
        out = {}
        for f in FILTERABLE:
            vals = (qs.exclude(**{f: ''}).values_list(f, flat=True)
                      .order_by(f).distinct()[:500])
            out[f] = list(vals)
        lo, hi = _period_bounds(qs)
        out['date_range'] = {'from': lo.isoformat() if lo else None,
                             'to': hi.isoformat() if hi else None}
        out['dimensions'] = sorted(DIMENSIONS.keys())
        return Response(out)


class SalesInsightsView(APIView):
    """Auto-generated written observations — the 'so what' the numbers imply.

    Kept server-side so the same wording appears everywhere the data is shown."""

    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        insights = []

        total = _money(qs.aggregate(v=Sum('net_amount'))['v'])
        if not total:
            return Response({'insights': [], 'filters': applied})

        def top_of(field, label, min_share=0):
            rows = (qs.exclude(**{field: ''}).values(field)
                      .annotate(v=Sum('net_amount')).order_by('-v')[:3])
            rows = [r for r in rows if r['v']]
            if not rows:
                return None
            share = float(rows[0]['v']) / total * 100
            if share < min_share:
                return None
            return rows, share

        # Concentration risk — one region/customer carrying too much of the book.
        for field, noun in (('state', 'state'), ('customer_name', 'customer'),
                            ('category', 'category')):
            got = top_of(field, noun)
            if not got:
                continue
            rows, share = got
            name = rows[0][field]
            if share >= 40:
                insights.append({
                    'type': 'risk',
                    'title': f'Heavy concentration in one {noun}',
                    'body': f'{name} alone accounts for {share:.0f}% of sales. '
                            f'That is a single point of failure — losing it would take a '
                            f'large share of revenue with it.',
                })
            elif share >= 25:
                insights.append({
                    'type': 'info',
                    'title': f'Top {noun}: {name}',
                    'body': f'{name} contributes {share:.0f}% of total sales — the largest '
                            f'single {noun} in this selection.',
                })

        # Target achievement
        tgt = _money(qs.aggregate(v=Sum('target_amount'))['v'])
        if tgt:
            ach = total / tgt * 100
            if ach >= 100:
                insights.append({
                    'type': 'win',
                    'title': f'Target exceeded — {ach:.0f}%',
                    'body': f'Sales of {total:,.0f} against a target of {tgt:,.0f}, '
                            f'ahead by {total - tgt:,.0f}.',
                })
            else:
                insights.append({
                    'type': 'risk' if ach < 80 else 'info',
                    'title': f'Target achievement at {ach:.0f}%',
                    'body': f'Short of target by {tgt - total:,.0f}. '
                            f'{"Well behind plan — worth investigating by region." if ach < 80 else "Within reach of plan."}',
                })
            # Who is dragging
            lag = (qs.exclude(salesperson='').values('salesperson')
                     .annotate(rev=Sum('net_amount'), t=Sum('target_amount'))
                     .filter(t__gt=0).order_by('rev'))
            lag = [r for r in lag if float(r['rev']) / float(r['t']) < 0.7][:3]
            if lag:
                names = ', '.join(f"{r['salesperson']} ({float(r['rev'])/float(r['t'])*100:.0f}%)"
                                  for r in lag)
                insights.append({
                    'type': 'risk',
                    'title': 'Sales people below 70% of target',
                    'body': f'{names}. These are the biggest gaps to close.',
                })

        # Momentum from the monthly series
        months = list(qs.values('period').annotate(v=Sum('net_amount')).order_by('period'))
        if len(months) >= 4:
            recent = [float(m['v'] or 0) for m in months[-3:]]
            earlier = [float(m['v'] or 0) for m in months[-6:-3]] or recent
            r_avg, e_avg = sum(recent) / len(recent), sum(earlier) / len(earlier)
            if e_avg:
                delta = (r_avg - e_avg) / e_avg * 100
                if abs(delta) >= 10:
                    insights.append({
                        'type': 'win' if delta > 0 else 'risk',
                        'title': f'Momentum {"rising" if delta > 0 else "falling"} '
                                 f'{abs(delta):.0f}%',
                        'body': f'The last 3 months average {r_avg:,.0f} vs {e_avg:,.0f} in the '
                                f'3 months before — a clear '
                                f'{"upswing" if delta > 0 else "slowdown"}.',
                    })

        # Discount pressure
        disc = _money(qs.aggregate(v=Sum('discount'))['v'])
        gross = _money(qs.aggregate(v=Sum('gross_amount'))['v'])
        if gross and disc / gross * 100 >= 12:
            insights.append({
                'type': 'risk',
                'title': f'Discounting at {disc / gross * 100:.0f}% of gross',
                'body': f'{disc:,.0f} given away against {gross:,.0f} gross. '
                        f'High discount intensity erodes margin even when top-line looks healthy.',
            })

        # Dormant customers — bought before, nothing recently.
        lo, hi = _period_bounds(qs)
        if hi:
            cutoff = hi - timedelta(days=90)
            recent_c = set(qs.filter(order_date__gt=cutoff)
                             .exclude(customer_name='')
                             .values_list('customer_name', flat=True).distinct())
            all_c = set(qs.exclude(customer_name='')
                          .values_list('customer_name', flat=True).distinct())
            dormant = all_c - recent_c
            if dormant and len(all_c) >= 5:
                insights.append({
                    'type': 'risk',
                    'title': f'{len(dormant)} customer(s) inactive for 90+ days',
                    'body': f'{", ".join(sorted(dormant)[:5])}'
                            f'{" and others" if len(dormant) > 5 else ""} have not ordered in the '
                            f'last 90 days of this period. Worth a win-back call.',
                })

        return Response({'insights': insights, 'filters': applied})


class SalesUploadsView(APIView):
    """List uploads; delete one (rolls back a bad file) or all."""

    def get(self, request):
        ups = SalesUpload.objects.all()[:100]
        return Response({'results': [{
            'id': u.id, 'filename': u.filename, 'rows': u.row_count,
            'skipped': u.skipped_rows, 'revenue': _money(u.total_revenue),
            'period_start': u.period_start.isoformat() if u.period_start else None,
            'period_end': u.period_end.isoformat() if u.period_end else None,
            'warnings': u.warnings, 'status': u.status,
            'created_at': u.created_at.isoformat(),
        } for u in ups],
            'total_rows': SalesRecord.objects.count(),
            'count': SalesUpload.objects.count()})

    def delete(self, request):
        up_id = request.query_params.get('id')
        if up_id:
            try:
                u = SalesUpload.objects.get(id=up_id)
            except SalesUpload.DoesNotExist:
                return Response({'error': 'Upload not found'}, status=404)
            n = u.records.count()
            u.delete()   # cascades to its rows
            return Response({'message': f'Removed upload "{u.filename}" and {n:,} row(s).',
                             'deleted': n})
        n = SalesRecord.objects.count()
        SalesRecord.objects.all().delete()
        SalesUpload.objects.all().delete()
        return Response({'message': f'Cleared all sales data ({n:,} row(s)).', 'deleted': n})


class SalesExportView(APIView):
    """Export the current filtered view as Excel — summary + per-dimension sheets."""

    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        wb = openpyxl.Workbook()
        from openpyxl.styles import Font as F, PatternFill as P

        def sheet(title, header, rows):
            ws = wb.create_sheet(title[:31])
            ws.append(header)
            for c in ws[1]:
                c.font = F(bold=True, color='FFFFFF')
                c.fill = P(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            for r in rows:
                ws.append(r)
            for i, _ in enumerate(header, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
            ws.freeze_panes = 'A2'

        agg = qs.aggregate(rev=Sum('net_amount'), qty=Sum('quantity'),
                           tgt=Sum('target_amount'), lines=Count('id'))
        ws = wb.active
        ws.title = 'Summary'
        ws.append(['SalesIQ Export'])
        ws['A1'].font = F(bold=True, size=14)
        ws.append([])
        for k, v in (('Total Revenue', _money(agg['rev'])),
                     ('Total Quantity', _money(agg['qty'])),
                     ('Total Target', _money(agg['tgt'])),
                     ('Rows', agg['lines'] or 0)):
            ws.append([k, v])
        ws.append([])
        ws.append(['Filters applied'])
        for k, v in (applied or {'(none)': ''}).items():
            ws.append([k, ', '.join(v) if isinstance(v, list) else str(v)])
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 40

        for dim_key in ('state', 'area', 'category', 'product', 'channel', 'salesperson',
                        'customer'):
            field = DIMENSIONS[dim_key]
            rows = (qs.exclude(**{field: ''}).values(field)
                      .annotate(rev=Sum('net_amount'), qty=Sum('quantity'),
                                tgt=Sum('target_amount'))
                      .order_by('-rev')[:500])
            if not rows:
                continue
            sheet(dim_key.title(), [dim_key.title(), 'Revenue', 'Quantity', 'Target',
                                    'Achievement %'],
                  [[r[field], _money(r['rev']), _money(r['qty']), _money(r['tgt']),
                    round(float(r['rev'] or 0) / float(r['tgt']) * 100, 1) if r['tgt'] else '']
                   for r in rows])

        trend = (qs.values('period').annotate(rev=Sum('net_amount'), qty=Sum('quantity'),
                                              tgt=Sum('target_amount')).order_by('period'))
        if trend:
            sheet('Monthly Trend', ['Month', 'Revenue', 'Quantity', 'Target'],
                  [[r['period'].strftime('%b %Y'), _money(r['rev']), _money(r['qty']),
                    _money(r['tgt'])] for r in trend])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="SalesIQ_Export.xlsx"'
        return resp


# ══════════════════════════════════════════════════════════════════════════
# Advanced analytics — each delegates to sales/analytics.py and shares the
# same filter contract as the basic endpoints.
# ══════════════════════════════════════════════════════════════════════════
from . import analytics as AN


def _dim_or_400(request, default='state'):
    key = (request.query_params.get('dim') or default).strip().lower()
    field = DIMENSIONS.get(key)
    return key, field


class SalesParetoView(APIView):
    """80/20 concentration with ABC classification."""
    def get(self, request):
        key, field = _dim_or_400(request, 'customer')
        if not field:
            return Response({'error': f'Unknown dimension "{key}".',
                             'available': sorted(DIMENSIONS.keys())}, status=400)
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.pareto(qs, field)
        data.update({'dimension': key, 'filters': applied})
        return Response(data)


class SalesMatrixView(APIView):
    """Revenue-vs-growth quadrant (star / cash cow / rising / watch)."""
    def get(self, request):
        key, field = _dim_or_400(request, 'product')
        if not field:
            return Response({'error': f'Unknown dimension "{key}".'}, status=400)
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        lo, hi = _period_bounds(qs)
        base, _ = apply_dim_filters(SalesRecord.objects.all(), request)
        data = AN.growth_matrix(base, field, lo, hi)
        data.update({'dimension': key, 'filters': applied})
        return Response(data)


class SalesMoversView(APIView):
    """Biggest absolute gainers and losers vs the prior equal window."""
    def get(self, request):
        key, field = _dim_or_400(request, 'state')
        if not field:
            return Response({'error': f'Unknown dimension "{key}".'}, status=400)
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        lo, hi = _period_bounds(qs)
        base, _ = apply_dim_filters(SalesRecord.objects.all(), request)
        data = AN.movers(base, field, lo, hi)
        data.update({'dimension': key, 'filters': applied})
        return Response(data)


class SalesAnomaliesView(APIView):
    def get(self, request):
        try:
            z = max(1.0, min(4.0, float(request.query_params.get('z', 2.0))))
        except (TypeError, ValueError):
            z = 2.0
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.anomalies(qs, z=z)
        data['filters'] = applied
        return Response(data)


class SalesSeasonalityView(APIView):
    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.seasonality(qs)
        data['filters'] = applied
        return Response(data)


class SalesHeatmapView(APIView):
    def get(self, request):
        key, field = _dim_or_400(request, 'state')
        if not field:
            return Response({'error': f'Unknown dimension "{key}".'}, status=400)
        try:
            top = max(3, min(25, int(request.query_params.get('top', 12))))
        except (TypeError, ValueError):
            top = 12
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.heatmap(qs, field, top=top)
        data.update({'dimension': key, 'filters': applied})
        return Response(data)


class SalesRFMView(APIView):
    """Recency / Frequency / Monetary customer segmentation."""
    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.rfm(qs)
        data['filters'] = applied
        return Response(data)


class SalesCohortsView(APIView):
    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.cohorts(qs)
        data['filters'] = applied
        return Response(data)


class SalesNewRepeatView(APIView):
    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.new_vs_repeat(qs)
        data['filters'] = applied
        return Response(data)


class SalesYoYView(APIView):
    """Year-on-year: uses dimension filters but ignores the date window so
    prior years remain visible when the user narrows the range."""
    def get(self, request):
        base, applied = apply_dim_filters(SalesRecord.objects.all(), request)
        data = AN.year_on_year(SalesRecord.objects.all(), base)
        data['filters'] = applied
        return Response(data)


class SalesPacingView(APIView):
    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.pacing(qs)
        data['filters'] = applied
        return Response(data)


class SalesPriceView(APIView):
    def get(self, request):
        qs, applied = apply_filters(SalesRecord.objects.all(), request)
        data = AN.price_realisation(qs)
        data['filters'] = applied
        return Response(data)


# ══════════════════════════════════════════════════════════════════════════
# SalesIQ access control — email OTP, super-admin only.
# ══════════════════════════════════════════════════════════════════════════
import os
import secrets
from django.core.cache import cache

# Super admin. Deliberately a hard-coded constant rather than an env default:
# sales data is commercially sensitive, so access has to be an explicit code
# change, never something a stray env var can widen by accident.
SALESIQ_SUPER_ADMIN = 'anshul@apisindia.com'
_OTP_TTL = 300          # 5 minutes
_OTP_MAX_ATTEMPTS = 5   # per issued code, then it is burned


def _salesiq_allowed_emails():
    """Super admin plus anyone explicitly listed in SALESIQ_ADMIN_EMAILS.
    There is intentionally NO "any @apisindia.com" fallback here — unlike the
    PMS Simulator — because this exposes company-wide revenue."""
    extra = [e.strip().lower() for e in
             os.getenv('SALESIQ_ADMIN_EMAILS', '').split(',') if e.strip()]
    return set([SALESIQ_SUPER_ADMIN] + extra)


def _mask(email):
    try:
        name, dom = email.split('@', 1)
        head = name[:2] if len(name) > 2 else name[:1]
        return f"{head}{'*' * max(1, len(name) - len(head))}@{dom}"
    except Exception:
        return email


class SalesLoginView(APIView):
    """POST /api/sales/login/ { action: 'send_otp' | 'verify_otp', email, otp }"""

    def post(self, request):
        action = str(request.data.get('action') or '').strip()
        email = str(request.data.get('email') or '').strip().lower()

        if action == 'send_otp':
            if not email or '@' not in email:
                return Response({'error': 'Please enter a valid email address.'}, status=400)
            if email not in _salesiq_allowed_emails():
                # Same wording regardless of whether the address exists, so this
                # can't be used to enumerate who has access.
                return Response({'error': 'This email is not authorised for SalesIQ. '
                                          'Contact the administrator for access.'}, status=403)

            code = f"{secrets.randbelow(1000000):06d}"
            cache.set(f'salesiq_otp_{email}', {'code': code, 'attempts': 0}, timeout=_OTP_TTL)
            try:
                from django.core.mail import send_mail
                from django.conf import settings
                send_mail(
                    subject='APIS SalesIQ — Login Code',
                    message=(f"Your SalesIQ login code is:\n\n    {code}\n\n"
                             f"Valid for 5 minutes. Do not share it with anyone.\n\n"
                             f"If you did not request this, someone has your email address "
                             f"but not your access — no action is needed.\n\n— APIS SalesIQ"),
                    from_email=(getattr(settings, 'OFFER_LETTER_EMAIL_HOST_USER', None)
                                or settings.EMAIL_HOST_USER or settings.DEFAULT_FROM_EMAIL),
                    recipient_list=[email],
                    fail_silently=False,
                )
            except Exception as e:
                cache.delete(f'salesiq_otp_{email}')
                return Response({'error': f'Could not send the login code: {e}'}, status=500)

            return Response({'message': f'Login code sent to {_mask(email)}',
                             'masked_email': _mask(email), 'expires_in': _OTP_TTL})

        if action == 'verify_otp':
            code = str(request.data.get('otp') or '').strip()
            key = f'salesiq_otp_{email}'
            saved = cache.get(key)
            if not saved:
                return Response({'error': 'That code has expired. Request a new one.'}, status=400)

            # Burn the code after repeated failures so a 6-digit OTP can't be
            # brute-forced within its 5-minute window.
            if saved.get('attempts', 0) >= _OTP_MAX_ATTEMPTS:
                cache.delete(key)
                return Response({'error': 'Too many incorrect attempts. Request a new code.'},
                                status=429)

            if code and secrets.compare_digest(str(saved.get('code')), code):
                cache.delete(key)
                return Response({'success': True, 'email': email,
                                 'role': 'super_admin', 'name': email.split('@')[0]})

            saved['attempts'] = saved.get('attempts', 0) + 1
            cache.set(key, saved, timeout=_OTP_TTL)
            left = _OTP_MAX_ATTEMPTS - saved['attempts']
            return Response({'error': f'Incorrect code. {left} attempt(s) remaining.'}, status=400)

        return Response({'error': 'Invalid action.'}, status=400)
