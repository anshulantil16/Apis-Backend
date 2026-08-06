"""Excel template + tolerant parsing for SalesIQ uploads.

Real sales exports never match a fixed schema — the same field appears as
"State", "STATE NAME" or "Billing State" depending on who ran the report. So
headers are matched against an alias table rather than required verbatim, and
anything unrecognised is reported back instead of silently dropped.
"""
import io
import re
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Canonical field -> every header spelling seen in the wild. Matching is done
# on a normalised key (lowercased, punctuation stripped).
COLUMN_ALIASES = {
    'order_date':    ['order date', 'date', 'invoice date', 'bill date', 'txn date',
                      'transaction date', 'sale date', 'posting date', 'month'],
    'invoice_no':    ['invoice no', 'invoice number', 'invoice', 'bill no', 'bill number',
                      'document no', 'order no', 'order number'],
    'zone':          ['zone', 'sales zone'],
    'state':         ['state', 'state name', 'billing state', 'customer state'],
    'city':          ['city', 'town', 'city name', 'billing city'],
    'area':          ['area', 'area name', 'district', 'beat', 'sales area'],
    'region':        ['region', 'region name'],
    'sku':           ['sku', 'sku code', 'item code', 'product code', 'material code',
                      'article code', 'item no'],
    'product_name':  ['product', 'product name', 'item name', 'item description',
                      'material description', 'description'],
    'category':      ['category', 'product category', 'item category', 'segment'],
    'sub_category':  ['sub category', 'subcategory', 'sub segment', 'product sub category'],
    'brand':         ['brand', 'brand name'],
    'pack_size':     ['pack size', 'pack', 'size', 'grammage', 'weight'],
    'uom':           ['uom', 'unit', 'unit of measure'],
    'channel':       ['channel', 'sales channel', 'trade channel', 'route to market'],
    'customer_code': ['customer code', 'party code', 'distributor code', 'dealer code',
                      'buyer code', 'account code'],
    'customer_name': ['customer', 'customer name', 'party name', 'distributor',
                      'distributor name', 'dealer', 'dealer name', 'buyer', 'account name'],
    'customer_type': ['customer type', 'party type', 'account type'],
    'salesperson':   ['salesperson', 'sales person', 'sales executive', 'se', 'so',
                      'sales officer', 'executive', 'employee name', 'sales rep'],
    'asm':           ['asm', 'area sales manager', 'area manager'],
    'rsm':           ['rsm', 'regional sales manager', 'regional manager'],
    'territory':     ['territory', 'territory name', 'beat name'],
    'quantity':      ['quantity', 'qty', 'sales qty', 'billed qty', 'volume', 'units'],
    'unit_price':    ['unit price', 'rate', 'price', 'mrp', 'selling price'],
    'gross_amount':  ['gross amount', 'gross', 'gross value', 'gross sales'],
    'discount':      ['discount', 'discount amount', 'scheme', 'scheme amount'],
    'tax':           ['tax', 'gst', 'tax amount', 'gst amount'],
    'net_amount':    ['net amount', 'net sales', 'net value', 'amount', 'sales value',
                      'sales', 'revenue', 'total', 'net revenue', 'value', 'net'],
    'target_amount': ['target', 'target amount', 'budget', 'budget amount', 'plan',
                      'target value', 'goal'],
}

TEXT_FIELDS = ['invoice_no', 'zone', 'state', 'city', 'area', 'region', 'sku',
               'product_name', 'category', 'sub_category', 'brand', 'pack_size', 'uom',
               'channel', 'customer_code', 'customer_name', 'customer_type',
               'salesperson', 'asm', 'rsm', 'territory']
NUM_FIELDS = ['quantity', 'unit_price', 'gross_amount', 'discount', 'tax',
              'net_amount', 'target_amount']

# Max chars per text field, mirroring the model's max_length values so a long
# free-text cell truncates instead of raising a DB "Data too long" error.
TEXT_MAX = {
    'invoice_no': 100, 'zone': 100, 'state': 100, 'city': 100, 'area': 150, 'region': 100,
    'sku': 100, 'product_name': 255, 'category': 150, 'sub_category': 150, 'brand': 150,
    'pack_size': 80, 'uom': 40, 'channel': 100, 'customer_code': 100, 'customer_name': 255,
    'customer_type': 100, 'salesperson': 200, 'asm': 200, 'rsm': 200, 'territory': 150,
}


def _norm(h):
    """Normalise a header cell for alias matching."""
    s = str(h or '').strip().lower().replace('*', '')
    for ch in ('_', '-', '.', '/', '(', ')', ':'):
        s = s.replace(ch, ' ')
    return ' '.join(s.split())


_LOOKUP = {}
for field, aliases in COLUMN_ALIASES.items():
    for a in aliases:
        _LOOKUP[_norm(a)] = field


def map_headers(header_row):
    """-> (field -> column index, list of unrecognised header names)."""
    col_map, unknown = {}, []
    for ci, cell in enumerate(header_row):
        if cell is None or str(cell).strip() == '':
            continue
        key = _norm(cell)
        field = _LOOKUP.get(key)
        if field is None:
            unknown.append(str(cell).strip())
        elif field not in col_map:      # first occurrence wins
            col_map[field] = ci
    return col_map, unknown


_ISO_RE = re.compile(r'^\d{4}-\d{1,2}-\d{1,2}(?:[T ]|$)')


def parse_date(v):
    """Parse a date from Excel cells OR API query params.

    ISO (YYYY-MM-DD) is checked FIRST and parsed strictly. dateutil with
    dayfirst=True reads "2026-02-01" as 2 January, not 1 February — which
    silently corrupted every date-range filter coming from an <input type=
    "date"> whenever the day was <= 12. Indian sheets still need dayfirst for
    "05-04-2026", so both rules coexist: ISO wins when the shape is ISO.
    """
    if v is None or str(v).strip() == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if _ISO_RE.match(s):
        try:
            return date.fromisoformat(s[:10])
        except ValueError:
            pass    # e.g. 2026-13-45 — fall through to the tolerant parser
    try:
        from dateutil import parser as _p
        return _p.parse(s, dayfirst=True).date()
    except Exception:
        return None


def parse_num(v, default=0.0):
    """Tolerant of Excel text cells: '1,234.50', '₹1,234', '(500)' negatives, '12%'."""
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return default
    neg = s.startswith('(') and s.endswith(')')
    for ch in ('₹', 'Rs.', 'Rs', ',', '%', '(', ')', '"'):
        s = s.replace(ch, '')
    s = s.strip()
    if not s or s in ('-', '--'):
        return default
    try:
        val = float(s)
    except ValueError:
        return default
    return -val if neg else val


def build_template():
    """Excel template: one sheet of columns + a reference sheet explaining them."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sales Data'

    groups = [
        ('When',     ['Order Date *', 'Invoice No'],                                  '1F4E79'),
        ('Where',    ['Zone', 'State', 'City', 'Area', 'Region'],                     '2E75B6'),
        ('Product',  ['SKU', 'Product Name', 'Category', 'Sub Category', 'Brand',
                      'Pack Size', 'UOM'],                                            '548235'),
        ('Customer', ['Channel', 'Customer Code', 'Customer Name', 'Customer Type'],  'BF8F00'),
        ('Sales Team', ['Salesperson', 'ASM', 'RSM', 'Territory'],                    '7030A0'),
        ('Money',    ['Quantity', 'Unit Price', 'Gross Amount', 'Discount', 'Tax',
                      'Net Amount *', 'Target Amount'],                               'C00000'),
    ]
    headers, colours = [], []
    for _, cols, colour in groups:
        headers += cols
        colours += [colour] * len(cols)

    border = Border(*(Side(style='thin'),) * 4)
    for ci, (h, colour) in enumerate(zip(headers, colours), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = PatternFill(start_color=colour, end_color=colour, fill_type='solid')
        c.font = Font(color='FFFFFF', bold=True, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border

    # Sample rows using APIS India's actual product range.
    samples = [
        ['2026-04-05', 'INV-1001', 'North', 'Delhi', 'New Delhi', 'Karol Bagh', 'North 1',
         'APS-HNY-500', 'APIS Himalaya Honey 500g', 'Honey', 'Natural Honey', 'APIS',
         '500g', 'PCS', 'Distributor', 'CUST-001', 'Sharma Traders', 'Distributor',
         'Rahul Sharma', 'Vikas Gupta', 'Anil Mehra', 'Delhi North',
         120, 250, 30000, 1500, 1425, 29925, 35000],
        ['2026-04-08', 'INV-1002', 'North', 'Punjab', 'Ludhiana', 'Model Town', 'North 2',
         'APS-DTS-500', 'APIS Premium Dates 500g', 'Dates', 'Seedless Dates', 'APIS',
         '500g', 'PCS', 'Modern Trade', 'CUST-014', 'BigBasket', 'Modern Trade',
         'Priya Singh', 'Vikas Gupta', 'Anil Mehra', 'Punjab Central',
         80, 320, 25600, 2000, 1180, 24780, 22000],
        ['2026-04-12', 'INV-1003', 'West', 'Maharashtra', 'Mumbai', 'Andheri', 'West 1',
         'APS-JAM-450', 'APIS Mixed Fruit Jam 450g', 'Jams & Spreads', 'Fruit Jam', 'APIS',
         '450g', 'PCS', 'E-Commerce', 'CUST-021', 'Amazon Retail', 'E-Commerce',
         'Amit Kumar', 'Rohit Deshmukh', 'Sunil Rao', 'Mumbai West',
         200, 180, 36000, 3600, 1620, 34020, 40000],
        ['2026-04-15', 'INV-1004', 'South', 'Karnataka', 'Bengaluru', 'Koramangala', 'South 1',
         'APS-GHE-1L', 'APIS Pure Cow Ghee 1L', 'Ghee', 'Cow Ghee', 'APIS',
         '1L', 'PCS', 'Retail', 'CUST-033', 'Sri Lakshmi Stores', 'Retailer',
         'Neha Gupta', 'Karthik Iyer', 'Sunil Rao', 'Bengaluru South',
         45, 720, 32400, 800, 1580, 33180, 30000],
        ['2026-05-03', 'INV-1005', 'East', 'West Bengal', 'Kolkata', 'Salt Lake', 'East 1',
         'APS-HNY-1KG', 'APIS Himalaya Honey 1kg', 'Honey', 'Natural Honey', 'APIS',
         '1kg', 'PCS', 'Distributor', 'CUST-045', 'Bose Enterprises', 'Distributor',
         'Sourav Das', 'Debasish Roy', 'Anil Mehra', 'Kolkata East',
         150, 460, 69000, 3450, 3277, 68827, 65000],
    ]
    for ri, row in enumerate(samples, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border

    widths = {1: 14, 2: 14, 9: 28, 18: 22, 19: 16, 20: 16, 21: 16}
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(i, 16)
    ws.freeze_panes = 'A2'

    # ── Reference sheet ──
    ref = wb.create_sheet('How To Use')
    ref.column_dimensions['A'].width = 26
    ref.column_dimensions['B'].width = 96
    rows = [
        ('SalesIQ — Upload Guide', ''),
        ('', ''),
        ('Required', 'Only TWO columns are mandatory: Order Date and Net Amount. '
                     'Everything else is optional — fill what you have.'),
        ('More columns = more insight',
         'Each optional column unlocks its own analysis. No State column means no '
         'state-wise view; no Salesperson means no team leaderboard.'),
        ('Column names are flexible',
         'Headers are auto-detected. "State", "STATE NAME" and "Billing State" all map '
         'to the same field. Unrecognised columns are reported after upload, not dropped '
         'silently.'),
        ('Net Amount', 'The headline sales figure every KPI is built on. If you only have '
                       'a single "Amount" or "Sales Value" column, name it that — it maps here.'),
        ('Target Amount', 'Optional. Fill it and the dashboard adds achievement %, '
                          'gap-to-target and who is behind plan. Leave blank to skip.'),
        ('Quantity', 'Optional but recommended — enables volume analysis alongside value, '
                     'which is how you spot price-led vs volume-led growth.'),
        ('Dates', 'Any common format works: 2026-04-05, 05-04-2026, 05/04/2026, 5 Apr 2026. '
                  'Day-first is assumed for ambiguous dates (05-04-2026 = 5 April).'),
        ('Row granularity', 'Invoice-line level is ideal. Pre-aggregated monthly rows also '
                            'work — just put the first of the month as Order Date.'),
        ('Forecasting', '6+ months of history enables trend forecasting; 24+ months enables '
                        'seasonal forecasting. Upload as much history as you have.'),
        ('Multiple uploads', 'Uploads add to the dataset. Each upload can be removed '
                             'individually from the dashboard if you load the wrong file.'),
    ]
    for ri, (a, b) in enumerate(rows, 1):
        ref.cell(row=ri, column=1, value=a).font = Font(bold=True, size=12 if ri == 1 else 10)
        c = ref.cell(row=ri, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical='top')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
