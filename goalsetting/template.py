"""The employee-master template an admin downloads, fills in and uploads back.

Generated rather than kept as a file in the repo, so it cannot drift from the
importer: the column headings ARE the importer's canonical names and the
user_type values ARE the ones it recognises. A template that has quietly gone
out of date is worse than none, because it fails at upload with a complaint
about a column the person is certain they included.

The data sheet holds NOTHING but the header row. An earlier draft put a
guidance row under the headings and three example people below it, with a note
in red saying to delete them - which is a trap, not a design: forget, and four
imaginary employees are imported. Per-column help is attached to the header
cells as Excel comments instead, and the worked example lives on the second
sheet where it cannot be uploaded by accident.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (heading, width, required, what it is)
COLUMNS = [
    ('employee_id', 18, True,
     'Their staff ID. Must be unique.\n\nEverything else points at a person by '
     'this, so it has to match exactly wherever it appears.'),
    ('name', 26, True, 'Full name, as it should appear on screen.'),
    ('email', 32, False,
     'Where their sign-in code is sent.\n\nWithout it they cannot sign in at all.'),
    ('designation', 24, False, 'Job title. Optional.'),
    ('department', 22, False, 'Used for filtering and grouping. Optional.'),
    ('reporting_manager_id', 24, False,
     'The employee_id of their manager - NOT the manager\'s name.\n\nThis is what '
     'puts someone into a manager\'s team list.'),
    ('hod_id', 18, False,
     'The employee_id of their HOD.\n\nThis is what puts someone into an HOD\'s '
     'team list.'),
    ('user_type', 18, False,
     'One of: employee, manager, hod, admin.\n\nLeave blank for a normal employee. '
     '"admin" hands that person the whole console, including the ability to edit '
     'anyone else\'s goals.'),
    ('zone', 18, False, 'Optional.'),
    ('subzone', 18, False, 'Optional.'),
    ('phone', 18, False, 'Optional.'),
    ('joined_date', 16, False,
     'Date of joining, written day-month-year.\n\nFor example 25-04-2024.'),
]

# Filled-in rows, so it is obvious what each column should look like - and
# harmless if they are left in. Their IDs begin with SAMPLE-, and the importer
# skips those, reporting how many it ignored. That is what makes them safe:
# an example you must remember to delete is a trap, not a help.
SAMPLE_ROWS = [
    # id, name, email, designation, department, manager, hod, type, zone, subzone, phone, joined
    ['SAMPLE-001', 'Rahul Sharma', 'rahul.sharma@apisindia.com', 'Sales Executive',
     'Sales', 'SAMPLE-002', 'SAMPLE-003', 'employee', 'North', 'Delhi',
     '9876543210', '25-04-2024'],
    ['SAMPLE-002', 'Arun Mishra', 'arun.mishra@apisindia.com', 'Area Sales Manager',
     'Sales', 'SAMPLE-003', 'SAMPLE-003', 'manager', 'North', '',
     '9876543211', '01-06-2021'],
    ['SAMPLE-003', 'Narendra Gangwar', 'narendra.gangwar@apisindia.com', 'Head - Sales',
     'Sales', '', '', 'hod', 'North', '', '9876543212', '15-01-2018'],
]

NOTES = [
    ('The three SAMPLE rows are yours to overwrite', [
        'They show what a filled-in row looks like, including how a reporting line '
        'is written.',
        'You can delete them, type over them, or simply leave them - the upload '
        'ignores any row whose employee_id starts with SAMPLE-, and tells you how '
        'many it skipped.',
    ]),
    ('Two columns are compulsory', [
        'employee_id and name. Every other column may be left blank.',
        'A row with neither is skipped, and reported back to you after the upload.',
    ]),
    ('Managers and HODs need their own row too', [
        'A manager is not created by being named in somebody else\'s '
        'reporting_manager_id column.',
        'If Arun manages Rahul, Arun needs his own row, with user_type set to manager.',
    ]),
    ('Reporting lines are IDs, never names', [
        'reporting_manager_id and hod_id hold an employee_id, not a person\'s name.',
        'Getting this wrong is the commonest reason a team list comes up empty.',
    ]),
    ('Uploading again is safe', [
        'Anyone already on the list is updated rather than duplicated - people are '
        'matched on employee_id.',
        'So you can correct a mistake by fixing the sheet and uploading the whole '
        'thing again.',
        'Re-uploading does NOT remove people who have dropped out of the sheet. '
        'Switch a leaver off with the Edit button in People instead.',
    ]),
    ('Things that will not import', [
        'A renamed or merged heading. Row 1 is read as the column names - leave it '
        'exactly as it is.',
        'Dates written month-first. Use day-month-year, as in 25-04-2024.',
        'A second header row. Start your people on row 2.',
    ]),
]

HEAD_FILL = PatternFill('solid', fgColor='1A1410')
REQ_FILL = PatternFill('solid', fgColor='B45309')
GUIDE_FILL = PatternFill('solid', fgColor='FEF3C7')


def build_template() -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = 'Employees'
    ws.freeze_panes = 'A2'          # headings stay visible while typing

    for i, (name, width, required, help_text) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = REQ_FILL if required else HEAD_FILL
        c.alignment = Alignment(horizontal='left', vertical='center')
        # Hover help, rather than a guidance row that would be imported.
        c.comment = Comment(
            ('REQUIRED\n\n' if required else '') + help_text, 'Goal Setting')
        c.comment.width = 300
        c.comment.height = 130
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 26

    # Filled-in rows, tinted so they read as examples rather than as data. Safe
    # to leave in: the importer skips SAMPLE- ids.
    for r, row in enumerate(SAMPLE_ROWS, start=2):
        for i, value in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=value)
            c.fill = GUIDE_FILL
            c.font = Font(size=10, color='78350F')

    # No hint text on this sheet, deliberately. Anything written in a cell below
    # the header is read back as a row of data - a note saying "these are only
    # examples" would itself arrive at the importer as a nameless employee. The
    # SAMPLE- ids say what they are, the headings carry hover notes, and the
    # second sheet holds the rest.

    guide = wb.create_sheet('How to fill this in')
    guide.column_dimensions['A'].width = 3
    for col, w in zip('BCDEF', (20, 26, 24, 16, 14)):
        guide.column_dimensions[col].width = w

    t = guide.cell(row=1, column=2, value='Goal Setting - employee master')
    t.font = Font(bold=True, size=15, color='1A1410')
    s = guide.cell(row=2, column=2,
                   value='Fill in the Employees sheet - headings in row 1, your people from '
                         'row 2 down - then upload it from Admin > People > Upload sheet. '
                         'Hover any heading for what that column means.')
    s.font = Font(size=10, color='64748B')
    s.alignment = Alignment(wrap_text=True, vertical='top')
    guide.merge_cells('B2:F2')
    guide.row_dimensions[2].height = 30

    r = 4
    h = guide.cell(row=r, column=2, value='A worked example')
    h.font = Font(bold=True, size=11, color='B45309')
    r += 1
    note = guide.cell(row=r, column=2,
                      value='Rahul reports to Arun; both sit under Narendra as HOD. '
                            'Note that all three appear as rows of their own.')
    note.font = Font(size=10, color='334155')
    guide.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 2

    hierarchy = [
        ['employee_id', 'name', 'reporting_manager_id', 'hod_id', 'user_type'],
        ['SAMPLE-001', 'Rahul Sharma', 'SAMPLE-002', 'SAMPLE-003', 'employee'],
        ['SAMPLE-002', 'Arun Mishra', 'SAMPLE-003', 'SAMPLE-003', 'manager'],
        ['SAMPLE-003', 'Narendra Gangwar', '', '', 'hod'],
    ]
    for j, row in enumerate(hierarchy):
        for i, value in enumerate(row, start=2):
            c = guide.cell(row=r, column=i, value=value)
            if j == 0:
                c.font = Font(bold=True, size=9, color='78350F')
            else:
                c.font = Font(size=10, color='334155')
                c.fill = GUIDE_FILL
        r += 1
    r += 2

    for heading, lines in NOTES:
        h = guide.cell(row=r, column=2, value=heading)
        h.font = Font(bold=True, size=11, color='B45309')
        r += 1
        for line in lines:
            c = guide.cell(row=r, column=2, value='•  ' + line)
            c.font = Font(size=10, color='334155')
            c.alignment = Alignment(wrap_text=True, vertical='top')
            guide.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            guide.row_dimensions[r].height = 28
            r += 1
        r += 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
