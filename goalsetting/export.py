"""The agreed goals for everyone, as a workbook.

The point of this product is the sheet everybody signed off on, so that is what
the export leads with: one row per KPI, flat, with the person's details repeated
on every row. Flat rather than nested because the file exists to be filtered,
pivoted and pasted into a review deck - a shape that reads nicely in Excel and
cannot be sorted is the wrong shape for a spreadsheet.

Three sheets, because an admin asks three different questions of the same data:
what was agreed, how does each person's sheet total up, and who has not finished.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.tz import local_str

HEAD_FILL = PatternFill('solid', fgColor='1A1410')
GROUP_FILL = PatternFill('solid', fgColor='B45309')
THIN = Side(style='thin', color='E2E8F0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GOAL_COLUMNS = [
    ('Employee ID', 14), ('Name', 22), ('Designation', 20), ('Department', 18),
    ('Reports to', 20), ('HOD', 20), ('Cycle', 18), ('Status', 22),
    ('Agreed on', 14),
    ('Category', 24), ('KRA', 34), ('KPI / Metric', 30), ('Weight %', 10),
    ('Frequency', 13), ('Unit', 14), ('Direction', 18), ('Data source', 18),
    ('Plan / Target', 15),
]

SUMMARY_COLUMNS = [
    ('Employee ID', 14), ('Name', 22), ('Designation', 20), ('Department', 18),
    ('Reports to', 20), ('HOD', 20), ('Status', 24), ('KRAs', 8), ('KPIs', 8),
    ('Total weight %', 14), ('Submitted', 14), ('Agreed on', 14),
    ('Steps', 8), ('Last changed by', 22),
]

PENDING_COLUMNS = [
    ('Employee ID', 14), ('Name', 22), ('Designation', 20), ('Department', 18),
    ('Reports to', 20), ('HOD', 20), ('Where it is stuck', 26), ('Waiting on', 24),
]

WAITING_ON = {
    'draft': 'The employee, to fill it in',
    'submitted': 'The manager, to review',
    'with_hod': 'The HOD, to review',
    'awaiting_employee': 'The employee, to accept',
    'returned': 'The employee, to make changes',
    'accepted': '-',
}


def _d(value):
    """dd-mm-yyyy in IST, matching the rest of the intranet. These are
    DateTimeFields stored in UTC, so a bare strftime would print the UTC
    calendar day, not the day it actually was in India."""
    return local_str(value, '%d-%m-%Y') or ''


def _header(ws, columns, fill=HEAD_FILL):
    for i, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(columns))}1'


def _write(ws, row, values, tint=None):
    for i, value in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=value)
        c.font = Font(size=10)
        c.border = BORDER
        c.alignment = Alignment(vertical='top', wrap_text=i in (10, 11, 12))
        if tint:
            c.fill = tint


def build_export(plans, all_plans, people, cycle_name='All cycles'):
    """plans: the sheets being exported (agreed ones, by default).
    all_plans: EVERY sheet in scope, whatever its status.
    people: everyone, so the third sheet can name those who have not started.

    all_plans is separate from plans deliberately. The third sheet answers "who
    still owes me goals", and building it from the filtered list reported anyone
    mid-review as "Not started" - which is the opposite of the truth, in the one
    column an admin would use to chase people.
    """
    wb = Workbook()

    # ── 1. the goals themselves ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Goals'
    _header(ws, GOAL_COLUMNS)

    row = 2
    for plan in plans:
        emp = plan.employee
        manager = emp.manager
        hod = emp.hod
        base = [
            emp.employee_id, emp.name, emp.designation, emp.department,
            manager.name if manager else emp.reporting_manager_id,
            hod.name if hod else emp.hod_id,
            plan.cycle.name, plan.get_status_display(), _d(plan.accepted_at),
        ]
        wrote = False
        for kra in plan.kras.all():
            for kpi in kra.kpis.all():
                _write(ws, row, base + [
                    kra.category, kra.title, kpi.metric, kpi.weightage,
                    kpi.frequency, kpi.unit_of_measurement, kpi.parameter_type,
                    kpi.data_source, kpi.target_value,
                ])
                row += 1
                wrote = True
            if not kra.kpis.exists():
                # A KRA with no measures is still worth seeing - it is usually
                # the reason a sheet is stuck.
                _write(ws, row, base + [kra.category, kra.title,
                                        '(no KPI set)', None, '', '', '', '', ''])
                row += 1
                wrote = True
        if not wrote:
            _write(ws, row, base + ['(nothing filled in)', '', '', None, '', '', '', '', ''])
            row += 1

    if row == 2:
        c = ws.cell(row=2, column=1, value='No goal sheets matched. Try including every '
                                           'status rather than only the agreed ones.')
        c.font = Font(italic=True, color='94A3B8', size=10)

    # ── 2. one line per person ───────────────────────────────────────────────
    summary = wb.create_sheet('Summary')
    _header(summary, SUMMARY_COLUMNS, GROUP_FILL)

    for r, plan in enumerate(plans, start=2):
        emp = plan.employee
        manager, hod = emp.manager, emp.hod
        versions = list(plan.versions.all())
        last = versions[-1] if versions else None
        _write(summary, r, [
            emp.employee_id, emp.name, emp.designation, emp.department,
            manager.name if manager else emp.reporting_manager_id,
            hod.name if hod else emp.hod_id,
            plan.get_status_display(),
            plan.kras.count(), plan.kpi_count, plan.total_weightage,
            _d(plan.submitted_at), _d(plan.accepted_at),
            len(versions),
            (last.actor_name or last.actor_role) if last else '',
        ])

    # ── 3. who has not finished ──────────────────────────────────────────────
    pending = wb.create_sheet('Not agreed yet')
    _header(pending, PENDING_COLUMNS, GROUP_FILL)

    started = {p.employee_id: p for p in all_plans}
    r = 2
    for emp in people:
        plan = started.get(emp.id)
        if plan and plan.status == 'accepted':
            continue
        manager, hod = emp.manager, emp.hod
        _write(pending, r, [
            emp.employee_id, emp.name, emp.designation, emp.department,
            manager.name if manager else emp.reporting_manager_id,
            hod.name if hod else emp.hod_id,
            plan.get_status_display() if plan else 'Not started',
            WAITING_ON.get(plan.status, '') if plan else 'The employee, to start it',
        ])
        r += 1

    if r == 2:
        c = pending.cell(row=2, column=1, value='Everyone has agreed their goals.')
        c.font = Font(italic=True, color='059669', size=10)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
