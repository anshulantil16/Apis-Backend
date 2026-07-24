"""PMS views — comprehensive HR data management with full Excel support."""
import io
import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from .models import PMSEmployee, PMSAuditLog, PMSSettings, OfferLetter, GRADE_META


def _apply_global_mgmt(employees):
    """Overlay the company-wide Management Score onto each employee instance (in memory
    only, not saved) so final_score / grade / increment all use the single value."""
    mgmt = PMSSettings.get_solo().management_score
    if mgmt is not None:
        for e in employees:
            e.management_score = mgmt
    return mgmt

GRADE_ORDER = ['A+', 'A', 'B+', 'B', 'C', 'D']


def serialize_emp(e):
    cfg = e.grade_config
    return {
        'id': e.id,
        'employee_id': e.employee_id,
        'name': e.name,
        'designation': e.designation,
        'new_designation': e.new_designation,
        'new_designation_type': e.new_designation_type,
        'cadre': e.cadre,
        'band': e.band,
        'level': e.level,
        'department': e.department,
        'business': e.business,
        'location': e.location,
        'payroll_location': e.payroll_location,
        'new_operational_location': e.new_operational_location,
        'sub_category': e.sub_category,
        'cost_centre': e.cost_centre,
        'category': e.category,
        'hq_location': e.hq_location,
        'gender': e.gender,
        'qualification': e.qualification,
        'date_of_birth': str(e.date_of_birth) if e.date_of_birth else None,
        'date_of_joining': str(e.date_of_joining) if e.date_of_joining else None,
        'age': e.age,
        'tenure_years': e.tenure_years,
        'reporting_manager': e.reporting_manager,
        'reporting_manager_id': e.reporting_manager_id,
        'hod_name': e.hod_name,
        'hod_id': e.hod_id,
        'fiscal_year': e.fiscal_year,
        'fy_2223_ctc': float(e.fy_2223_ctc) if e.fy_2223_ctc else None,
        'fy_2324_ctc': float(e.fy_2324_ctc) if e.fy_2324_ctc else None,
        'fy_2425_ctc': float(e.fy_2425_ctc) if e.fy_2425_ctc else None,
        'current_ctc': float(e.current_ctc),
        'variable_pay': float(e.variable_pay) if e.variable_pay else None,
        'fy_2223_growth_pct': float(e.fy_2223_growth_pct) if e.fy_2223_growth_pct else None,
        'fy_2324_growth_pct': float(e.fy_2324_growth_pct) if e.fy_2324_growth_pct else None,
        'fy_2425_growth_pct': float(e.fy_2425_growth_pct) if e.fy_2425_growth_pct else None,
        'emp_score': float(e.emp_score) if e.emp_score is not None else None,
        'manager_score': float(e.manager_score) if e.manager_score is not None else None,
        'hod_score': float(e.hod_score) if e.hod_score is not None else None,
        'management_score': float(e.management_score) if e.management_score is not None else None,
        'fy_2223_grade': e.fy_2223_grade,
        'fy_2324_grade': e.fy_2324_grade,
        'fy_2425_grade': e.fy_2425_grade,
        'last_promotion_year': e.last_promotion_year,
        'final_score': e.final_score,
        'auto_grade': e.auto_grade,
        'effective_grade': e.effective_grade,
        'override_grade': e.override_grade,
        'grade_label': cfg['label'],
        'grade_color': cfg['color'],
        'inc_min': cfg['inc_min'],
        'inc_max': cfg['inc_max'],
        'promo_pct': cfg['promo_pct'],
        'increment_group': e.increment_group,
        'is_worker': e.is_worker,
        'override_increment_pct': float(e.override_increment_pct) if e.override_increment_pct is not None else None,
        'effective_increment_pct': e.effective_increment_pct,
        'increment_amount': e.increment_amount,
        'promotion_pct': float(e.promotion_pct),
        'effective_promotion_pct': e.effective_promotion_pct,
        'promotion_amount': e.promotion_amount,
        'management_discretion_pct': float(e.management_discretion_pct),
        'management_discretion_amount': e.management_discretion_amount,
        'salary_correction': float(e.salary_correction),
        'salary_correction_amount': e.salary_correction_amount,
        'salary_correction_allowed': e.salary_correction_allowed,
        'merit_eligible': e.merit_eligible,
        'reward_payout': e.reward_payout,
        'special_reward_range': list(e.special_reward_range) if e.special_reward_range else None,
        'total_impact_pct': e.total_impact_pct,
        'new_ctc': e.new_ctc,
        'new_ctc_monthly': e.new_ctc_monthly,
        'promoted': e.promoted,
        'redesignation': e.redesignation,
        'on_time_reward': e.on_time_reward,
        'sustained_performance': e.sustained_performance,
        'sustained_pct': e.sustained_pct,
        'sustained_amount': e.sustained_amount,
        'reward_amount': float(e.reward_amount),
        'promotion_readiness': e.promotion_readiness,
        'manager_remarks': e.manager_remarks,
        'hod_remarks': e.hod_remarks,
        'notes': e.notes,
    }


def build_summary(employees):
    total = len(employees)
    if total == 0:
        return {'total_employees': 0}

    total_ctc     = sum(float(e.current_ctc) for e in employees)
    total_new_ctc = sum(e.new_ctc for e in employees)
    total_inc     = total_new_ctc - total_ctc
    promoted      = sum(1 for e in employees if e.promoted)
    rewarded      = sum(1 for e in employees if e.on_time_reward)
    avg_score     = sum(e.final_score for e in employees) / total

    # Component-wise cost of the total hike
    cost_increment  = round(sum(e.increment_amount for e in employees), 2)
    cost_promotion  = round(sum(e.promotion_amount for e in employees), 2)
    cost_sustained  = round(sum(e.sustained_amount for e in employees), 2)
    cost_correction = round(sum(e.salary_correction_amount for e in employees), 2)
    cost_reward     = round(sum(e.reward_payout for e in employees), 2)
    cost_mgmt_disc  = round(sum(e.management_discretion_amount for e in employees), 2)
    sustained_count = sum(1 for e in employees if e.sustained_performance)
    worker_count    = sum(1 for e in employees if e.is_worker)

    # Workforce demographics (from imported data)
    ages    = [e.age for e in employees if e.age]
    tenures = [e.tenure_years for e in employees if e.tenure_years is not None]
    avg_age    = round(sum(ages) / len(ages), 1) if ages else None
    avg_tenure = round(sum(tenures) / len(tenures), 1) if tenures else None

    def _dist(attr):
        d = {}
        for e in employees:
            k = (getattr(e, attr) or 'Unknown')
            d[k] = d.get(k, 0) + 1
        return dict(sorted(d.items(), key=lambda x: -x[1]))
    location_dist = _dist('location')
    category_dist = _dist('category')
    cadre_dist    = _dist('band')

    grade_dist = {}
    for e in employees:
        g = e.effective_grade
        grade_dist[g] = grade_dist.get(g, 0) + 1

    # Cost-centre-wise breakdown (headcount, increment, promotion, sustained, reward, new CTC)
    cc_map = {}
    for e in employees:
        cc = e.cost_centre or 'Unknown'
        v = cc_map.setdefault(cc, {'count': 0, 'current': 0, 'new': 0, 'inc': 0, 'promo': 0,
                                   'sust': 0, 'reward': 0, 'promoted': 0, 'sustained': 0})
        v['count'] += 1
        v['current'] += float(e.current_ctc)
        v['new'] += e.new_ctc
        v['inc'] += e.increment_amount
        v['promo'] += e.promotion_amount
        v['sust'] += e.sustained_amount
        v['reward'] += e.reward_payout
        if e.promoted: v['promoted'] += 1
        if e.sustained_performance: v['sustained'] += 1
    cost_centre_breakdown = [{
        'cost_centre': cc,
        'count': v['count'],
        'current_ctc': round(v['current'], 2),
        'new_ctc': round(v['new'], 2),
        'increment': round(v['new'] - v['current'], 2),
        'increment_cost': round(v['inc'], 2),
        'promotion_cost': round(v['promo'], 2),
        'sustained_cost': round(v['sust'], 2),
        'reward_cost': round(v['reward'], 2),
        'promoted': v['promoted'],
        'sustained': v['sustained'],
        'hike_pct': round((v['new'] - v['current']) / v['current'] * 100, 2) if v['current'] else 0,
    } for cc, v in sorted(cc_map.items(), key=lambda x: -x[1]['count'])]

    dept_map = {}
    for e in employees:
        d = e.department or 'Unknown'
        if d not in dept_map:
            dept_map[d] = {'current': 0, 'new': 0, 'count': 0, 'scores': [], 'promoted': 0}
        dept_map[d]['current'] += float(e.current_ctc)
        dept_map[d]['new']     += e.new_ctc
        dept_map[d]['count']   += 1
        dept_map[d]['scores'].append(e.final_score)
        if e.promoted:
            dept_map[d]['promoted'] += 1

    dept_breakdown = [{
        'department': d,
        'count': v['count'],
        'current_ctc': round(v['current'], 2),
        'new_ctc': round(v['new'], 2),
        'increment': round(v['new'] - v['current'], 2),
        'avg_score': round(sum(v['scores']) / len(v['scores']), 2),
        'promotion_cases': v['promoted'],
    } for d, v in sorted(dept_map.items())]

    grade_inc = {}
    for e in employees:
        g = e.effective_grade
        if g not in grade_inc:
            grade_inc[g] = {'total_inc': 0, 'count': 0, 'total_ctc': 0}
        grade_inc[g]['total_inc'] += e.increment_amount
        grade_inc[g]['count']     += 1
        grade_inc[g]['total_ctc'] += float(e.current_ctc)

    sorted_emps = sorted(employees, key=lambda e: e.final_score, reverse=True)
    top10 = [serialize_emp(e) for e in sorted_emps[:10]]
    bot10 = [serialize_emp(e) for e in sorted_emps[-10:]]

    readiness = {'ready_now': 0, '1_year': 0, '2_years': 0, 'not_ready': 0}
    for e in employees:
        if e.promotion_readiness in readiness:
            readiness[e.promotion_readiness] += 1

    scores = sorted(e.final_score for e in employees)
    ctcs   = sorted(float(e.current_ctc) for e in employees)
    med_score = scores[total // 2]
    med_ctc   = ctcs[total // 2]
    quadrants = {'high_perf_high_pay': 0, 'high_perf_low_pay': 0, 'low_perf_high_pay': 0, 'low_perf_low_pay': 0}
    for e in employees:
        hp = e.final_score >= med_score
        hs = float(e.current_ctc) >= med_ctc
        key = ('high' if hp else 'low') + '_perf_' + ('high' if hs else 'low') + '_pay'
        quadrants[key] += 1

    gender_map = {}
    for e in employees:
        g = e.gender or 'Not Specified'
        if g not in gender_map:
            gender_map[g] = {'count': 0, 'scores': []}
        gender_map[g]['count'] += 1
        gender_map[g]['scores'].append(e.final_score)
    gender_breakdown = [{'gender': g, 'count': v['count'], 'avg_score': round(sum(v['scores'])/len(v['scores']), 2)} for g, v in gender_map.items()]

    band_map = {}
    for e in employees:
        b = e.band or 'Unknown'
        if b not in band_map:
            band_map[b] = {'current': 0, 'new': 0, 'count': 0}
        band_map[b]['current'] += float(e.current_ctc)
        band_map[b]['new']     += e.new_ctc
        band_map[b]['count']   += 1

    return {
        'total_employees': total,
        'total_current_ctc': round(total_ctc, 2),
        'total_new_ctc': round(total_new_ctc, 2),
        'total_increment': round(total_inc, 2),
        'increment_pct': round((total_inc / total_ctc * 100) if total_ctc else 0, 2),
        'avg_score': round(avg_score, 2),
        'promoted_count': promoted,
        'reward_count': rewarded,
        'sustained_count': sustained_count,
        'worker_count': worker_count,
        'cost_increment': cost_increment,
        'cost_promotion': cost_promotion,
        'cost_sustained': cost_sustained,
        'cost_correction': cost_correction,
        'cost_reward': cost_reward,
        'cost_mgmt_discretion': cost_mgmt_disc,
        'avg_age': avg_age,
        'avg_tenure': avg_tenure,
        'location_distribution': location_dist,
        'category_distribution': category_dist,
        'cadre_distribution': cadre_dist,
        'cost_centre_breakdown': cost_centre_breakdown,
        'grade_distribution': grade_dist,
        'grade_increment_breakdown': {g: {
            'count': v['count'],
            'total_increment': round(v['total_inc'], 2),
            'avg_increment_pct': round((v['total_inc'] / v['total_ctc'] * 100) if v['total_ctc'] else 0, 2),
        } for g, v in grade_inc.items()},
        'department_breakdown': dept_breakdown,
        'band_breakdown': [{'band': b, 'count': v['count'], 'current_ctc': round(v['current'], 2), 'new_ctc': round(v['new'], 2)} for b, v in sorted(band_map.items())],
        'top10_performers': top10,
        'bottom10_performers': bot10,
        'promotion_readiness': readiness,
        'performance_vs_salary': quadrants,
        'gender_breakdown': gender_breakdown,
        'median_score': med_score,
        'median_ctc': med_ctc,
    }


class PMSListView(APIView):
    def get(self, request):
        employees = list(PMSEmployee.objects.all())
        return Response({'employees': [serialize_emp(e) for e in employees], 'summary': build_summary(employees)})

    def delete(self, request):
        # `pms_offerletter` is an orphaned table (no Django model) whose non-cascading
        # FK to pms_pmsemployee blocks employee deletion. Clear those rows first so the
        # employee delete (which cascades to audit logs) can succeed.
        from django.db import connection
        with connection.cursor() as cursor:
            try:
                cursor.execute("DELETE FROM pms_offerletter")
            except Exception:
                pass
        PMSEmployee.objects.all().delete()
        return Response({'message': 'All PMS data cleared.'})


class PMSImportView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {str(e)}'}, status=400)

        HEADER_ALIASES = {
            'sr no': 'sr_no', 'sr. no': 'sr_no', 'sno': 'sr_no',
            'employee id': 'employee_id', 'emp id': 'employee_id', 'empid': 'employee_id', 'er no': 'employee_id',
            'employee name': 'name', 'name': 'name',
            'designation': 'designation', 'current designation': 'designation',
            'new designation': 'new_designation', 're-designation': 'new_designation',
            'new designation type': 'new_designation_type', 'new employee type': 'new_designation_type', 'designation type': 'new_designation_type',
            'department': 'department', 'dept': 'department',
            'new department': 'business', 'new function': 'business',
            'location': 'location', 'zone': 'location',
            'new operational location': 'new_operational_location', 'est': 'new_operational_location', 'sub zone': 'sub_category',
            'sub category': 'sub_category', 'sub cat': 'sub_category',
            'cost centre': 'cost_centre', 'cost center': 'cost_centre', 'cc': 'cost_centre',
            'category': 'category',
            'hq': 'hq_location', 'hq location': 'hq_location',
            'payroll location': 'payroll_location',
            'cadre': 'cadre',
            'band': 'band', 'grade': 'band',
            'level': 'level',
            'gender': 'gender',
            'qualification': 'qualification', 'qualifications': 'qualification',
            'dob': 'date_of_birth', 'date of birth': 'date_of_birth',
            'age': 'age',
            'doj': 'date_of_joining', 'date of joining': 'date_of_joining',
            'tenure': 'tenure_years', 'tenure in apis as on 31-mar-2026': 'tenure_years', 'tenure in apis as on 31st mar-2026': 'tenure_years',
            'reporting manager': 'reporting_manager', 'report manager': 'reporting_manager',
            'report id': 'reporting_manager_id', 'report no id': 'reporting_manager_id', 'reporting id': 'reporting_manager_id',
            'hod name': 'hod_name',
            'hod id': 'hod_id', 'hod': 'hod_id',
            'last promotion': 'last_promotion_year', 'last promotion (yr)': 'last_promotion_year', 'last promotion (year)': 'last_promotion_year',
            'fy 22-23 ctc': 'fy_2223_ctc', 'fy 2223 ctc': 'fy_2223_ctc', 'fy 22-23 (ctc)': 'fy_2223_ctc',
            'fy 23-24 ctc': 'fy_2324_ctc', 'fy 2324 ctc': 'fy_2324_ctc', 'fy 23-24 (ctc)': 'fy_2324_ctc',
            'fy 24-25 ctc': 'fy_2425_ctc', 'fy 2425 ctc': 'fy_2425_ctc', 'fy 24-25 (ctc)': 'fy_2425_ctc',
            'fy 25-26 current ctc': 'current_ctc', 'current ctc': 'current_ctc', 'ctc': 'current_ctc', 'current ctc (annual inr)': 'current_ctc',
            'fy 25-26 (current ctc)': 'current_ctc', 'current ctc - 31-mar-26': 'current_ctc', 'current ctc- 31-mar-26': 'current_ctc',
            'current variable pay 31-mar-26': 'variable_pay', 'current variable pay': 'variable_pay',
            'variable pay 31-mar-26': 'variable_pay', 'variable pay': 'variable_pay',
            'fy 22-23 (%)': 'fy_2223_growth_pct', 'fy 22-23 %': 'fy_2223_growth_pct',
            'fy 23-24 (%)': 'fy_2324_growth_pct', 'fy 23-24 %': 'fy_2324_growth_pct',
            'fy 24-25 (%)': 'fy_2425_growth_pct', 'fy 24-25 %': 'fy_2425_growth_pct',
            'final score': 'final_score_value', 'final score (0-120)': 'final_score_value', 'final score (0-100)': 'final_score_value',
            'score': 'final_score_value', 'overall score': 'final_score_value', 'total score': 'final_score_value',
            'self score': 'emp_score', 'emp score': 'emp_score', 'emp score (0-100)': 'emp_score',
            'manager score': 'manager_score', 'mgr score': 'manager_score', 'manager score (0-100)': 'manager_score',
            'hod score': 'hod_score', 'hod score (0-100)': 'hod_score',
            'fy 22-23 grade': 'fy_2223_grade',
            'fy 23-24 grade': 'fy_2324_grade',
            'fy 24-25 grade': 'fy_2425_grade',
            'score range': 'score_range',
            'final score range': 'final_score_range',
            'rating': 'rating',
            'performance': 'performance',
            'promotion (y/n)': 'promoted', 'promotion': 'promoted',
            'promotion readiness': 'promotion_readiness',
            'salary correction': 'salary_correction', 'salary correction level': 'salary_correction',
            'promotion %': 'promotion_pct',
            'management discretion': 'management_discretion_pct', 'management discretion %': 'management_discretion_pct',
            'one time reward': 'on_time_reward',
            'sustained performance': 'sustained_performance', 'sustained': 'sustained_performance',
            'reward amount': 'reward_amount',
            'redesignation': 'redesignation', 're-designation': 'redesignation',
            'revised ctc': 'revised_ctc',
            'increment nt %': 'increment_nt_pct', 'increment nt': 'increment_nt_pct',
            'increment on %': 'promotion_pct', 'increment on': 'promotion_pct',
            'hike %': 'total_impact_pct', 'total hike %': 'total_impact_pct',
            'fy 26-27 ctc': 'new_ctc', 'fy 26-27': 'new_ctc',
            'management score': 'management_score', 'mgt score': 'management_score', 'management score (0-100)': 'management_score',
            'manager remarks': 'manager_remarks',
            'hod remarks': 'hod_remarks',
        }

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map = {}
        for ci, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace('*', '').strip()
            field = HEADER_ALIASES.get(key)
            if field:
                col_map[field] = ci

        created = updated = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            data = {field: row[ci] for field, ci in col_map.items() if ci < len(row)}
            emp_id = str(data.get('employee_id') or '').strip()
            name   = str(data.get('name') or '').strip()
            if not emp_id or not name:
                errors.append(f'Row {row_idx}: Missing Employee ID or Name — skipped')
                continue

            def sf(val, default=None):
                if val is None or str(val).strip() == '': return default
                try: return round(float(str(val).replace(',', '')), 2)
                except: return default

            def ss(val):
                v = sf(val)
                return None if v is None else max(0.0, min(100.0, v))

            def parse_bool(val):
                if val is None or str(val).strip() == '': return False
                return str(val).lower().strip() in ('y', 'yes', 'true', '1')

            def parse_date(val):
                if val is None or str(val).strip() == '': return None
                try:
                    from datetime import datetime
                    if isinstance(val, str):
                        return datetime.strptime(val, '%Y-%m-%d').date()
                    else:
                        return val.date() if hasattr(val, 'date') else val
                except:
                    return None

            obj, was_created = PMSEmployee.objects.update_or_create(
                employee_id=emp_id,
                defaults={
                    'name': name,
                    'designation': str(data.get('designation') or '').strip(),
                    'new_designation': str(data.get('new_designation') or '').strip(),
                    'new_designation_type': str(data.get('new_designation_type') or '').strip(),
                    'department': str(data.get('department') or '').strip(),
                    'location': str(data.get('location') or '').strip(),
                    'new_operational_location': str(data.get('new_operational_location') or '').strip(),
                    'sub_category': str(data.get('sub_category') or '').strip(),
                    'cost_centre': str(data.get('cost_centre') or '').strip(),
                    'category': str(data.get('category') or '').strip(),
                    'hq_location': str(data.get('hq_location') or '').strip(),
                    'payroll_location': str(data.get('payroll_location') or '').strip(),
                    'cadre': str(data.get('cadre') or '').strip(),
                    'band': str(data.get('band') or '').strip().upper()[:5],
                    'level': str(data.get('level') or '').strip(),
                    'gender': str(data.get('gender') or '').strip(),
                    'qualification': str(data.get('qualification') or '').strip(),
                    'date_of_birth': parse_date(data.get('date_of_birth')),
                    'date_of_joining': parse_date(data.get('date_of_joining')),
                    'reporting_manager': str(data.get('reporting_manager') or '').strip(),
                    'reporting_manager_id': str(data.get('reporting_manager_id') or '').strip(),
                    'hod_name': str(data.get('hod_name') or '').strip(),
                    'hod_id': str(data.get('hod_id') or '').strip(),
                    'fiscal_year': str(data.get('fiscal_year') or '2025-26').strip(),
                    'fy_2223_ctc': sf(data.get('fy_2223_ctc')),
                    'fy_2324_ctc': sf(data.get('fy_2324_ctc')),
                    'fy_2425_ctc': sf(data.get('fy_2425_ctc')),
                    'current_ctc': sf(data.get('current_ctc'), 0) or 0,
                    'variable_pay': sf(data.get('variable_pay')),
                    'fy_2223_growth_pct': sf(data.get('fy_2223_growth_pct')),
                    'fy_2324_growth_pct': sf(data.get('fy_2324_growth_pct')),
                    'fy_2425_growth_pct': sf(data.get('fy_2425_growth_pct')),
                    'final_score_value': sf(data.get('final_score_value')),
                    'emp_score': ss(data.get('emp_score')),
                    'manager_score': ss(data.get('manager_score')),
                    'hod_score': ss(data.get('hod_score')),
                    'fy_2223_grade': str(data.get('fy_2223_grade') or '').strip(),
                    'fy_2324_grade': str(data.get('fy_2324_grade') or '').strip(),
                    'fy_2425_grade': str(data.get('fy_2425_grade') or '').strip(),
                    'last_promotion_year': int(data.get('last_promotion_year')) if data.get('last_promotion_year') else None,
                    # NOTE: promotion %, management discretion, salary correction and special reward
                    # are NOT imported — they come strictly from the policy table (promotion) or are
                    # entered by management in the UI. Imported values would give wrong salaries.
                    # promoted / redesignation / sustained_performance are set ONLY on first import
                    # (below) so re-importing does not wipe management's UI decisions.
                    'promotion_readiness': str(data.get('promotion_readiness') or '').strip(),
                    'manager_remarks': str(data.get('manager_remarks') or '').strip(),
                    'hod_remarks': str(data.get('hod_remarks') or '').strip(),
                }
            )
            if was_created:
                obj.promoted = parse_bool(data.get('promoted'))
                obj.redesignation = parse_bool(data.get('redesignation'))
                obj.sustained_performance = parse_bool(data.get('sustained_performance'))
                obj.save(update_fields=['promoted', 'redesignation', 'sustained_performance'])
            created += was_created
            updated += not was_created

        employees = list(PMSEmployee.objects.all())
        return Response({
            'message': f'✅ Import complete! {created} new employees added, {updated} updated.',
            'created': created, 'updated': updated, 'errors': errors,
            'total': len(employees), 'summary': build_summary(employees),
        })


class PMSEmployeeUpdateView(APIView):
    def patch(self, request, emp_id):
        try:
            emp = PMSEmployee.objects.get(id=emp_id)
        except PMSEmployee.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)

        d = request.data
        simulate = d.get('simulate_only', False)
        logs = []

        fields_to_update = [
            'final_score_value', 'override_increment_pct',
            'override_grade', 'promoted', 'promotion_pct', 'management_discretion_pct',
            'on_time_reward', 'reward_amount', 'promotion_readiness', 'notes',
            'salary_correction', 'redesignation', 'sustained_performance',
        ]

        for field in fields_to_update:
            if field in d:
                old_val = getattr(emp, field)
                if field in ('promoted', 'redesignation', 'on_time_reward', 'sustained_performance'):
                    new_val = bool(d[field])
                elif field == 'final_score_value':
                    v = d[field]
                    new_val = float(v) if v not in (None, '', 'null') else None
                    if new_val is not None:
                        new_val = max(0.0, min(115.0, new_val))  # % of target; A+ up to 115%
                elif field in ('override_increment_pct', 'promotion_pct', 'management_discretion_pct', 'salary_correction', 'reward_amount'):
                    v = d[field]
                    new_val = float(v) if v not in (None, '', 'null') else (0 if field in ('promotion_pct', 'management_discretion_pct', 'salary_correction', 'reward_amount') else None)
                    if new_val is not None and new_val < 0:
                        new_val = 0
                    # ── Policy enforcement (management discretion % stays UNLIMITED) ──
                    if field == 'salary_correction':
                        # Correction only for A+/A/B+/B and not when promoted → else blocked.
                        allowed = (not emp.promoted) and emp.effective_grade in ('A+', 'A', 'B+', 'B')
                        if not allowed:
                            new_val = 0
                    elif field == 'reward_amount':
                        # Cap at the band's max range (M/O/W). C/D = Director/MD discretion (no cap).
                        rng = emp.special_reward_range
                        if rng and new_val and new_val > rng[1]:
                            new_val = float(rng[1])
                else:
                    new_val = d[field]

                if str(old_val) != str(new_val):
                    logs.append({'field': field, 'old_value': str(old_val), 'new_value': str(new_val)})
                setattr(emp, field, new_val)

        if not simulate:
            emp.save()
            for log in logs:
                PMSAuditLog.objects.create(employee=emp, **log)

        return Response(serialize_emp(emp))


class PMSSettingsView(APIView):
    """Get / set company-wide PMS settings (currently the single Management Score)."""
    def get(self, request):
        s = PMSSettings.get_solo()
        return Response({
            'management_score': float(s.management_score) if s.management_score is not None else None,
        })

    def post(self, request):
        s = PMSSettings.get_solo()
        v = request.data.get('management_score')
        if v in (None, '', 'null'):
            s.management_score = None
        else:
            s.management_score = max(0.0, min(100.0, float(v)))
        s.save()
        return Response({
            'management_score': float(s.management_score) if s.management_score is not None else None,
            'message': 'Management score updated for all employees.',
        })


class PMSTemplateView(APIView):
    def get(self, request):
        from django.http import HttpResponse
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PMS Import'

        headers = [
            'SR NO', 'Employee ID *', 'Employee Name *', 'Designation', 'New Designation',
            'Designation Type', 'Department', 'New Department', 'New Function', 'Zone',
            'Sub Zone', 'Cost Centre', 'Category', 'HQ', 'Payroll Location',
            'Cadre', 'Grade', 'Gender', 'Qualification', 'DOB', 'Age', 'DOJ',
            'Tenure in APIS as on 31st Mar-2026', 'Reporting Manager', 'Reporting ID', 'HOD Name', 'HOD ID',
            'Span of Control', 'Current CTC - 31-Mar-26 *', 'Current Variable Pay 31-Mar-26',
            'Last Promotion (YR)', 'Years Not Promoted Before 2022',
            'FY 22-23 (CTC)', 'Increment 22-23', 'FY 23-24 (CTC)', 'Increment 23-24',
            'FY 24-25 (CTC)', 'Increment 24-25', 'FY 22-23 (%)', 'FY 23-24 (%)', 'FY 24-25 (%)',
            'FY 25-26 (Current CTC)', 'Final Score (0-100) *',
            'Score Range', 'Final Score Range', 'Rating', 'Performance',
            'Promotion (Y/N)', 'Level', 'Promotion Readiness', 'Salary Correction Level',
            'Management Discretion', 'One Time Reward', 'Redesignation', 'Revised CTC',
            'Increment %', 'Promotion %', 'Total Hike %', 'Total Salary Increment',
            'CTC Fixed', 'Variable', 'FY 26-27 Total CTC', 'Manager Remarks', 'HOD Remarks',
        ]

        hf = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        hf_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            is_required = '*' in h
            is_new = 'New' in h or 'Operating' in h
            c.fill = hf_yellow if is_new else hf
            c.font = Font(color='FFFFFF' if not is_new else '000000', bold=True, size=9)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        samples = [
            [1, 'EMP001', 'Rahul Sharma', 'Sales Manager', 'Sr Sales Manager',
             'STAT', 'Sales', 'Sales', 'Revenue', 'Mumbai',
             'West', 'CC001', 'CAT01', 'HQ Mumbai', 'Mumbai',
             'M', 'M4', 'Male', 'MBA', '1985-05-15', 38, '2015-06-01',
             9, 'Rajesh Kumar', 'MGR001', 'Priya Singh', 'HOD001',
             5, 720000, 50000,
             2024, 1,
             600000, 50000, 650000, 50000,
             700000, 20000, 8, 7.7, 7.7,
             720000, 108,
             '', '', '', '',
             'Yes', 'L2', '1_year', 0,
             '', 'No', 'No', '',
             '', '', '', '',
             '', '', '', 'Performing well', 'Strong leader'],
            [2, 'EMP002', 'Priya Singh', 'Sr Executive', '',
             'MANAGER', 'Marketing', 'Marketing', 'Brand', 'Delhi',
             'North', 'CC002', 'CAT01', 'HQ Delhi', 'Delhi',
             'O', 'O5', 'Female', 'B.Tech', '1992-08-22', 31, '2018-03-15',
             7, 'Ramesh Gupta', 'MGR002', 'Amit Kumar', 'HOD002',
             3, 560000, 40000,
             2022, 0,
             450000, 30000, 480000, 40000,
             520000, 40000, 6.7, 8.3, 7.7,
             560000, 96,
             '', '', '', '',
             'Yes', 'L1', 'ready_now', 0,
             '', 'Yes', 'Yes', '',
             '', '', '', '',
             '', '', '', 'Excellent performance', 'Promotion ready'],
            [3, 'EMP003', 'Amit Kumar', 'Associate', '',
             '', 'Operations', 'Operations', 'Logistics', 'Chennai',
             'South', 'CC003', 'CAT02', 'HQ Chennai', 'Chennai',
             'W', 'W2', 'Male', 'B.Sc', '1998-12-10', 25, '2022-01-10',
             2, 'Vikram Singh', 'MGR003', 'Deepak Nair', 'HOD003',
             0, 320000, 20000,
             None, 3,
             250000, 30000, 280000, 20000,
             300000, 20000, 12, 7.1, 7.1,
             320000, 68,
             '', '', '', '',
             'No', 'L3', 'not_ready', 0,
             0, 'No', 'No', '',
             '', 0, '', '',
             '', '', '', 'Adequate performance', 'New employee'],
        ]

        for ri, row in enumerate(samples, 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)

        ws2 = wb.create_sheet('Grade Guide')
        guide = [
            ['Grade', 'Score Range', 'Label', 'Increment %', 'Promotion %'],
            ['A+', '≥ 106%', 'Exceptional', '12–15%', '10%'],
            ['A', '95–100%', 'Outstanding', '10–12%', '8%'],
            ['B+', '85–94%', 'Exceeds Target', '7–10%', '6%'],
            ['B', '65–84%', 'Meets Target', '4–7%', '4%'],
            ['C', '51–64%', 'Near Target', '0–4%', '0%'],
            ['D', '<50%', 'Needs Improvement', '2%', '0%'],
        ]
        for ri, row in enumerate(guide, 1):
            for ci, val in enumerate(row, 1):
                c = ws2.cell(row=ri, column=ci, value=val)
                if ri == 1: c.font = Font(bold=True)

        widths = [8, 14, 18, 16, 16, 18, 14, 12, 20, 16, 12, 12, 16, 16, 10, 8, 8, 10, 14, 14, 10, 12, 14, 18, 12, 12, 10, 14, 14, 14, 16, 14, 14, 14, 16, 16, 16, 14, 14, 14, 14, 12, 12, 12, 16, 16, 16, 18, 18, 20, 20]
        for ci, w in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 55

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="PMS_Complete_Template.xlsx"'
        return resp


class PMSExportView(APIView):
    """Complete appraisal-finalization workbook: full per-employee detail + org summary."""
    def get(self, request):
        from django.http import HttpResponse
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        employees = list(PMSEmployee.objects.all())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Appraisal Sheet'

        groups = [
            ('EMPLOYEE & ORGANISATION', 18, '1F4E79'),
            ('PERFORMANCE', 4, '548235'),
            ('CTC HISTORY', 4, '7F6000'),
            ('INCREMENT COMPONENTS', 13, 'C55A11'),
            ('REVISED CTC', 4, '2E75B6'),
            ('DECISIONS & REMARKS', 5, '7030A0'),
        ]
        headers = [
            'Emp ID', 'Name', 'Gender', 'Designation', 'New Designation',
            'Cadre', 'Grade', 'Category', 'Department', 'Cost Centre', 'Location', 'HQ',
            'Reporting Manager', 'HOD Name', 'DOJ', 'Age', 'Tenure (Yrs)', 'Merit Eligible',
            'Final Score', 'Perf Grade', 'Rating Label', 'Increment Category',
            'FY 22-23 CTC', 'FY 23-24 CTC', 'FY 24-25 CTC', 'Current CTC',
            'Merit Increment %', 'Merit Increment Rs',
            'Promoted', 'Promotion %', 'Promotion Rs',
            'Sustained', 'Sustained %', 'Sustained Rs',
            'Salary Correction Rs',
            'Special Reward', 'Special Reward Rs',
            'Mgmt Discretion %', 'Mgmt Discretion Rs',
            'Total Hike %', 'Total Increase Rs', 'New CTC (Annual)', 'New CTC (Monthly)',
            'Redesignation', 'Promotion Readiness', 'Manager Remarks', 'HOD Remarks', 'Notes',
        ]
        thin = Side(style='thin', color='D9D9D9')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        left_cols = {2, 4, 5, 13, 14, 46, 47, 48}

        ci = 1
        for gname, gcount, gcolor in groups:
            ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + gcount - 1)
            c = ws.cell(row=1, column=ci, value=gname)
            c.fill = PatternFill(start_color=gcolor, end_color=gcolor, fill_type='solid')
            c.font = Font(color='FFFFFF', bold=True, size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            ci += gcount

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=i, value=h)
            c.fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            c.font = Font(color='FFFFFF', bold=True, size=9)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

        for ri, e in enumerate(employees, 3):
            row = [
                e.employee_id, e.name, e.gender, e.designation, e.new_designation,
                e.cadre, e.band, e.category, e.department, e.cost_centre, e.location, e.hq_location,
                e.reporting_manager, e.hod_name, str(e.date_of_joining or ''), e.age or '',
                e.tenure_years if e.tenure_years is not None else '', 'Yes' if e.merit_eligible else 'No',
                e.final_score, e.effective_grade, e.grade_config['label'], e.increment_group,
                float(e.fy_2223_ctc or 0), float(e.fy_2324_ctc or 0), float(e.fy_2425_ctc or 0), float(e.current_ctc),
                e.effective_increment_pct, e.increment_amount,
                'Yes' if e.promoted else 'No', e.effective_promotion_pct, e.promotion_amount,
                'Yes' if e.sustained_performance else 'No', e.sustained_pct, e.sustained_amount,
                e.salary_correction_amount,
                'Yes' if e.on_time_reward else 'No', e.reward_payout,
                float(e.management_discretion_pct), e.management_discretion_amount,
                e.total_impact_pct, round(e.new_ctc - float(e.current_ctc), 2), e.new_ctc, e.new_ctc_monthly,
                'Yes' if e.redesignation else 'No', e.promotion_readiness, e.manager_remarks, e.hod_remarks, e.notes,
            ]
            for ci2, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci2, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='left' if ci2 in left_cols else 'center', vertical='center')

        # Totals row
        last = len(employees) + 3
        tcell = ws.cell(row=last, column=1, value='TOTAL')
        tcell.font = Font(bold=True, size=11)
        totals = {
            26: sum(float(e.current_ctc) for e in employees),
            28: sum(e.increment_amount for e in employees),
            31: sum(e.promotion_amount for e in employees),
            34: sum(e.sustained_amount for e in employees),
            35: sum(e.salary_correction_amount for e in employees),
            37: sum(e.reward_payout for e in employees),
            39: sum(e.management_discretion_amount for e in employees),
            41: sum(e.new_ctc - float(e.current_ctc) for e in employees),
            42: sum(e.new_ctc for e in employees),
        }
        for col, v in totals.items():
            cell = ws.cell(row=last, column=col, value=round(v, 2))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            cell.border = border

        ws.freeze_panes = 'C3'
        colw = ([10, 22, 8, 20, 20, 8, 8, 12, 18, 12, 14, 10, 20, 20, 12, 6, 10, 12] +
                [11, 10, 16, 16] + [14, 14, 14, 14] +
                [13, 15, 9, 12, 14, 10, 11, 13, 16, 12, 15, 14, 16] +
                [11, 15, 16, 16] + [12, 18, 26, 26, 26])
        for i, w in enumerate(colw[:len(headers)], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.row_dimensions[2].height = 40

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws2 = wb.create_sheet('Summary')
        ws2.column_dimensions['A'].width = 34
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 14
        total_ctc = sum(float(e.current_ctc) for e in employees)
        total_new = sum(e.new_ctc for e in employees)
        pct = lambda v: (v / total_ctc * 100) if total_ctc else 0

        def _row(ws_, r, a, b, c=None, bold=False, fill=None):
            ca = ws_.cell(row=r, column=1, value=a); cb = ws_.cell(row=r, column=2, value=b)
            if c is not None:
                cc = ws_.cell(row=r, column=3, value=c); cc.font = Font(bold=bold)
            ca.font = Font(bold=bold); cb.font = Font(bold=bold)
            if fill:
                for cc2 in (ca, cb): cc2.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')

        title = ws2.cell(row=1, column=1, value='APPRAISAL — SALARY IMPACT SUMMARY')
        title.font = Font(bold=True, size=13, color='1F4E79')
        _row(ws2, 3, 'Metric', 'Amount (Rs)', '% of Payroll', bold=True, fill='D9E1F2')
        rows = [
            ('Total Employees', len(employees), None),
            ('Current Payroll (Annual)', round(total_ctc, 2), None),
            ('Merit Increment', round(sum(e.increment_amount for e in employees), 2), round(pct(sum(e.increment_amount for e in employees)), 2)),
            ('Promotion', round(sum(e.promotion_amount for e in employees), 2), round(pct(sum(e.promotion_amount for e in employees)), 2)),
            ('Sustained Performance', round(sum(e.sustained_amount for e in employees), 2), round(pct(sum(e.sustained_amount for e in employees)), 2)),
            ('Salary Correction', round(sum(e.salary_correction_amount for e in employees), 2), round(pct(sum(e.salary_correction_amount for e in employees)), 2)),
            ('Management Discretion', round(sum(e.management_discretion_amount for e in employees), 2), round(pct(sum(e.management_discretion_amount for e in employees)), 2)),
            ('Total Hike (recurring)', round(total_new - total_ctc, 2), round(pct(total_new - total_ctc), 2)),
            ('New Payroll (Annual)', round(total_new, 2), None),
            ('One-Time Rewards (not in CTC)', round(sum(e.reward_payout for e in employees), 2), None),
        ]
        for idx, (a, b, c) in enumerate(rows, 4):
            _row(ws2, idx, a, b, c, bold=(a in ('Total Hike (recurring)', 'New Payroll (Annual)')),
                 fill='FCE4D6' if a in ('Total Hike (recurring)', 'New Payroll (Annual)') else None)

        r0 = 4 + len(rows) + 2
        _row(ws2, r0, 'Grade Distribution', 'Count', None, bold=True, fill='D9E1F2')
        gd = {}
        for e in employees:
            gd[e.effective_grade] = gd.get(e.effective_grade, 0) + 1
        for j, g in enumerate(GRADE_ORDER, r0 + 1):
            _row(ws2, j, g, gd.get(g, 0), None)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="PMS_Appraisal_Finalization.xlsx"'
        return resp


class OfferLetterTemplateView(APIView):
    """Generate Excel template for Offer Letter upload."""

    def get(self, request):
        from django.http import HttpResponse
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from .offer_letter import SALARY_COMPONENTS, ANNEXURE_EMP_FIELDS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Offer Letters'

        # Letter-level identity + Annexure employee details + salary-component columns
        identity = [
            'SR NO', 'Employee ID *', 'Title (Mr./Ms.)', 'Employee Name *', 'Email *', 'Department',
            'Function', 'Current Designation', 'New Designation', 'Cadre', 'Grade',
            'Date of Joining', 'Work Location', 'Current CTC *', 'New CTC *',
            'Increment %', 'Promotion %', 'Performance Rating', 'Performance Assessment', 'Grade Label',
            'Special Reward (One-time)', 'Special Reward Note',
            'Effective Date *', 'Remarks',
        ]
        component_headers = [c[3] for c in SALARY_COMPONENTS]
        headers = identity + component_headers

        hf = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        hf_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        hf_green = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        comp_start = len(identity)  # 0-based index where component columns begin

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            is_required = '*' in h
            is_comp = ci > comp_start
            c.fill = hf if is_required else (hf_green if is_comp else hf_yellow)
            c.font = Font(color='FFFFFF' if is_required else '000000', bold=True, size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

        # Sample rows (dict keyed by header; blank components demonstrate the
        # "not eligible → leave blank → hidden in letter" behaviour)
        sample_dicts = [
            {'SR NO': 1, 'Employee ID *': 'EMP001', 'Title (Mr./Ms.)': 'Mr.',
             'Employee Name *': 'Rahul Sharma', 'Email *': 'rahul.sharma@apis.com',
             'Department': 'Sales', 'Function': 'Field Sales', 'Current Designation': 'Sales Manager',
             'New Designation': 'Senior Sales Manager', 'Cadre': 'M2', 'Grade': 'G5',
             'Date of Joining': '01/06/2019', 'Work Location': 'Delhi HO',
             'Current CTC *': 600000, 'New CTC *': 660000, 'Increment %': 10, 'Promotion %': 5,
             'Performance Rating': 'A', 'Performance Assessment': 'Strong Performer',
             'Grade Label': 'Outstanding', 'Special Reward (One-time)': 25000,
             'Special Reward Note': 'For outstanding project delivery',
             'Effective Date *': '2026-04-01', 'Remarks': 'Excellent performer',
             'Basic Salary (Monthly)': 27500, 'HRA (Monthly)': 11000, 'Special Allowance (Monthly)': 8000,
             'Employer PF (Monthly)': 3300, 'Statutory Bonus (Monthly)': 2292, 'Variable Pay (Monthly)': 3333},
            {'SR NO': 2, 'Employee ID *': 'EMP002', 'Title (Mr./Ms.)': 'Ms.',
             'Employee Name *': 'Priya Singh', 'Email *': 'priya.singh@apis.com',
             'Department': 'Operations', 'Function': 'Operations', 'Current Designation': 'Executive',
             'New Designation': 'Senior Executive', 'Cadre': 'M1', 'Grade': 'G4',
             'Date of Joining': '15/03/2021', 'Work Location': 'Mumbai',
             'Current CTC *': 450000, 'New CTC *': 540000, 'Increment %': 12, 'Promotion %': 8,
             'Performance Rating': 'A+', 'Performance Assessment': 'Outstanding Performer',
             'Grade Label': 'Exceptional', 'Effective Date *': '2026-04-01', 'Remarks': 'Ready for promotion',
             'Basic Salary (Monthly)': 22500, 'HRA (Monthly)': 9000, 'Special Allowance (Monthly)': 6000,
             'Employer PF (Monthly)': 2700, 'Statutory Bonus (Monthly)': 1875},
            {'SR NO': 3, 'Employee ID *': 'EMP003', 'Title (Mr./Ms.)': 'Mr.',
             'Employee Name *': 'Amit Kumar', 'Email *': 'amit.kumar@apis.com',
             'Department': 'IT', 'Function': 'Information Technology', 'Current Designation': 'Associate',
             'New Designation': 'Associate', 'Cadre': 'E3', 'Grade': 'G3',
             'Date of Joining': '10/01/2023', 'Work Location': 'Delhi HO',
             'Current CTC *': 280000, 'New CTC *': 340000, 'Increment %': 5, 'Promotion %': 0,
             'Performance Rating': 'B', 'Performance Assessment': 'Solid Performer',
             'Grade Label': 'Meets Target', 'Effective Date *': '2026-04-01', 'Remarks': '',
             'Basic Salary (Monthly)': 14000, 'HRA (Monthly)': 5600, 'Special Allowance (Monthly)': 4000},
        ]

        date_col = headers.index('Effective Date *') + 1
        for ri, d in enumerate(sample_dicts, 2):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=ri, column=ci, value=d.get(h, ''))
                c.border = border
                if ci == date_col:
                    c.number_format = 'yyyy-mm-dd'

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="OfferLetter_Template.xlsx"'
        return resp


def _process_offer_batch(rows, batch_id, send_emails):
    """Background worker: generate PDFs (+ optional emails) and update batch progress.

    Runs in its own thread so a 500-1000 employee run never blocks/times-out the
    HTTP request. Uses ONE shared SMTP connection for the whole batch."""
    import io as _io
    import re
    from django.core.files.base import ContentFile
    from django.utils import timezone
    from django.db import connections
    from django.core.mail import get_connection
    from .offer_letter import generate_offer_letter_pdf, send_offer_letter_email
    from .models import OfferLetterBatch

    batch = None
    mail_conn = None
    try:
        batch = OfferLetterBatch.objects.get(batch_id=batch_id)
        if send_emails:
            try:
                mail_conn = get_connection()
                mail_conn.open()
            except Exception:
                mail_conn = None

        gen = eml = fail = proc = 0
        errs = []
        total = len(rows)
        for r in rows:
            try:
                offer = OfferLetter.objects.create(
                    employee=None, employee_code=r['emp_id'], employee_name=r['name'],
                    letter_type=r['letter_type'],
                    current_ctc=r['current_ctc'], new_ctc=r['new_ctc'],
                    increment_pct=r['increment_pct'], promotion_pct=r['promotion_pct'],
                    effective_date=r['effective_date'],
                    old_designation=r['current_designation'], new_designation=r['new_designation'],
                    performance_rating=r['performance_rating'], grade_label=r['grade_label'],
                    salutation=r['salutation'], assessment=r['assessment'],
                    function=r['function'], cadre=r['cadre'], grade=r['grade'],
                    date_of_joining=r['date_of_joining'], work_location=r['work_location'],
                    salary_breakup=r['salary_breakup'],
                    special_reward=r['special_reward'], special_reward_note=r['special_reward_note'],
                    email_address=r['email'], department=r['department'],
                    batch_id=batch_id, status='pending',
                )
                pdf_bytes = generate_offer_letter_pdf(
                    None, r['current_ctc'], r['new_ctc'], r['increment_pct'], r['promotion_pct'],
                    r['effective_date'],
                    old_designation=r['current_designation'], new_designation=r['new_designation'],
                    performance_rating=r['performance_rating'], grade_label=r['grade_label'],
                    employee_id=r['emp_id'], employee_name=r['name'], department=r['department'],
                    salutation_title=r['salutation'], assessment=r['assessment'],
                    emp_details=r['emp_details'], salary_breakup=r['salary_breakup'],
                    special_reward=r['special_reward'], special_reward_note=r['special_reward_note'],
                ).getvalue()
                safe_id = re.sub(r'[^A-Za-z0-9_.-]', '_', str(r['emp_id']))[:40]
                offer.pdf_file.save(f"offer_{safe_id}_{offer.id}.pdf",
                                    ContentFile(pdf_bytes), save=True)
                gen += 1

                if send_emails and r['email']:
                    try:
                        send_offer_letter_email(r['email'], r['name'], _io.BytesIO(pdf_bytes),
                                                r['effective_date'], offer.id, connection=mail_conn)
                        offer.status = 'sent'
                        offer.email_sent = True
                        offer.email_sent_at = timezone.now()
                        offer.save(update_fields=['status', 'email_sent', 'email_sent_at'])
                        eml += 1
                    except Exception as ee:
                        offer.status = 'failed'
                        offer.save(update_fields=['status'])
                        errs.append(f"{r['emp_id']}: email failed: {ee}")
            except Exception as e:
                fail += 1
                errs.append(f"{r['emp_id']}: {e}")

            proc += 1
            if proc % 5 == 0 or proc == total:  # flush progress periodically
                OfferLetterBatch.objects.filter(batch_id=batch_id).update(
                    processed=proc, generated=gen, emailed=eml, failed=fail, errors=errs[:50])

        OfferLetterBatch.objects.filter(batch_id=batch_id).update(
            processed=proc, generated=gen, emailed=eml, failed=fail,
            errors=errs[:50], status='completed')
    except Exception as e:
        if batch is not None:
            OfferLetterBatch.objects.filter(batch_id=batch_id).update(
                status='error', errors=[str(e)])
    finally:
        if mail_conn is not None:
            try:
                mail_conn.close()
            except Exception:
                pass
        connections.close_all()  # release this thread's DB connections


class OfferLetterUploadView(APIView):
    """Parse the uploaded Excel instantly, then generate letters in a background
    thread. Returns a batch_id the UI polls for progress — scales to 500-1000+."""
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        from datetime import datetime, date

        send_emails = str(request.data.get('send_emails', 'false')).lower() == 'true'

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {str(e)}'}, status=400)

        from .offer_letter import SALARY_COMPONENTS
        HEADER_MAP = {
            'sr no': 'sr_no', 'employee id': 'employee_id',
            'title (mr./ms.)': 'salutation', 'title': 'salutation',
            'employee name': 'name',
            'email': 'email', 'department': 'department', 'function': 'function',
            'current designation': 'current_designation', 'new designation': 'new_designation',
            'cadre': 'cadre', 'grade': 'grade',
            'date of joining': 'date_of_joining', 'work location': 'work_location',
            'current ctc': 'current_ctc', 'new ctc': 'new_ctc',
            'increment %': 'increment_pct', 'promotion %': 'promotion_pct',
            'performance rating': 'performance_rating',
            'performance assessment': 'assessment',
            'grade label': 'grade_label',
            'special reward (one-time)': 'special_reward', 'special reward': 'special_reward',
            'special reward note': 'special_reward_note',
            'effective date': 'effective_date', 'remarks': 'remarks',
        }
        # salary-component columns keyed by their (lower-cased) Excel header → component key
        COMP_MAP = {c[3].strip().lower(): c[0] for c in SALARY_COMPONENTS}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map = {}
        comp_col = {}   # component_key → column index
        for ci, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace('*', '').strip()
            if key in HEADER_MAP:
                col_map[HEADER_MAP[key]] = ci
            elif key in COMP_MAP:
                comp_col[COMP_MAP[key]] = ci

        required = ['employee_id', 'name', 'email', 'current_ctc', 'new_ctc', 'effective_date']
        if not all(f in col_map for f in required):
            return Response({'error': 'Missing required columns',
                             'required': ['Employee ID', 'Employee Name', 'Email', 'Current CTC', 'New CTC', 'Effective Date'],
                             'mapped': list(col_map.keys())}, status=400)

        def get_val(row, field, default=None):
            if field not in col_map:
                return default
            ci = col_map[field]
            return row[ci] if ci < len(row) else default

        def sf(val, default=0):
            if val is None or str(val).strip() == '':
                return default
            try:
                return float(str(val).replace(',', ''))
            except Exception:
                return default

        def format_date(val):
            if val is None or str(val).strip() == '':
                return None
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date):
                return val
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except Exception:
                    pass
            return None

        # ── Fast parse: turn every valid row into a plain dict (no PDF/DB yet) ──
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            emp_id = str(get_val(row, 'employee_id') or '').strip()
            name = str(get_val(row, 'name') or '').strip()
            if not emp_id or not name:
                continue
            increment_pct = sf(get_val(row, 'increment_pct'))
            promotion_pct = sf(get_val(row, 'promotion_pct'))
            current_designation = str(get_val(row, 'current_designation') or '').strip()
            new_designation = str(get_val(row, 'new_designation') or '').strip()
            function = str(get_val(row, 'function') or '').strip()
            cadre = str(get_val(row, 'cadre') or '').strip()
            grade = str(get_val(row, 'grade') or '').strip()
            date_of_joining = str(get_val(row, 'date_of_joining') or '').strip()
            work_location = str(get_val(row, 'work_location') or '').strip()

            salary_breakup = {}
            for ckey, ci in comp_col.items():
                cval = sf(row[ci] if ci < len(row) else None, default=0)
                if cval:
                    salary_breakup[ckey] = cval

            letter_type = 'increment'
            if new_designation and new_designation != current_designation:
                letter_type = 'promotion' if promotion_pct > 0 else 'redesignation'
            if increment_pct > 0 and promotion_pct > 0:
                letter_type = 'combined'

            rows.append({
                'emp_id': emp_id, 'name': name,
                'email': str(get_val(row, 'email') or '').strip(),
                'current_ctc': sf(get_val(row, 'current_ctc')),
                'new_ctc': sf(get_val(row, 'new_ctc')),
                'increment_pct': increment_pct, 'promotion_pct': promotion_pct,
                'current_designation': current_designation, 'new_designation': new_designation,
                'performance_rating': str(get_val(row, 'performance_rating') or '').strip(),
                'assessment': str(get_val(row, 'assessment') or '').strip(),
                'grade_label': str(get_val(row, 'grade_label') or '').strip(),
                'salutation': str(get_val(row, 'salutation') or '').strip(),
                'department': str(get_val(row, 'department') or '').strip(),
                'function': function, 'cadre': cadre, 'grade': grade,
                'date_of_joining': date_of_joining, 'work_location': work_location,
                'special_reward': sf(get_val(row, 'special_reward')),
                'special_reward_note': str(get_val(row, 'special_reward_note') or '').strip(),
                'effective_date': format_date(get_val(row, 'effective_date')) or date.today(),
                'salary_breakup': salary_breakup,
                'emp_details': {'function': function, 'cadre': cadre, 'grade': grade,
                                'date_of_joining': date_of_joining, 'work_location': work_location},
                'letter_type': letter_type,
            })

        if not rows:
            return Response({'error': 'No valid employee rows found in the file.'}, status=400)

        # ── Kick off background generation and return immediately ──
        import threading
        import uuid
        from .models import OfferLetterBatch
        batch_id = uuid.uuid4().hex[:16]
        OfferLetterBatch.objects.create(batch_id=batch_id, total=len(rows),
                                        send_emails=send_emails, status='running')
        threading.Thread(target=_process_offer_batch, args=(rows, batch_id, send_emails),
                         daemon=True).start()

        return Response({
            'message': f'Processing {len(rows)} letter(s) in the background…',
            'batch_id': batch_id, 'total': len(rows), 'send_emails': send_emails,
        })


class OfferLetterPDFView(APIView):
    """Download/view a generated offer-letter PDF."""
    def get(self, request, offer_letter_id):
        from django.http import FileResponse
        try:
            offer = OfferLetter.objects.get(id=offer_letter_id)
        except OfferLetter.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        if not offer.pdf_file:
            return Response({'error': 'No PDF generated for this letter'}, status=404)
        code = offer.employee_code or (offer.employee.employee_id if offer.employee else offer.id)
        return FileResponse(offer.pdf_file.open('rb'), content_type='application/pdf',
                            filename=f'OfferLetter_{code}.pdf')


class OfferLetterBatchStatusView(APIView):
    """Poll bulk-generation progress. Returns live counts while running and the
    per-letter results once completed."""
    def get(self, request, batch_id):
        from .models import OfferLetterBatch
        from django.utils import timezone
        from datetime import timedelta
        try:
            b = OfferLetterBatch.objects.get(batch_id=batch_id)
        except OfferLetterBatch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=404)

        # If a running batch hasn't advanced for 5 min, the worker died — mark it errored
        if b.status == 'running' and b.updated_at < timezone.now() - timedelta(minutes=5):
            b.status = 'error'
            b.errors = (b.errors or []) + ['Batch stalled — generation stopped unexpectedly. '
                                           'Please re-upload the remaining rows.']
            b.save(update_fields=['status', 'errors'])

        data = {
            'batch_id': b.batch_id, 'status': b.status, 'total': b.total,
            'processed': b.processed, 'generated': b.generated, 'emailed': b.emailed,
            'failed': b.failed, 'send_emails': b.send_emails, 'errors': b.errors,
        }
        if b.status == 'completed':
            data['results'] = [
                {'employee_id': o.employee_code, 'name': o.employee_name, 'status': o.status,
                 'pdf_url': f'/api/pms/offer-letter/{o.id}/pdf/'}
                for o in OfferLetter.objects.filter(batch_id=b.batch_id).order_by('id')
            ]
        return Response(data)
