"""Appraisal / compensation-revision letters (the "offer letter" pipeline).

Template -> Excel upload -> background batch generation -> history -> ZIP.
PDF rendering itself lives in pms/offer_letter.py.
"""
import io
import os
import re
import secrets
import openpyxl
from django.core.cache import cache
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from ..models import (PMSEmployee, PMSAuditLog, PMSSettings, OfferLetter,
                      WarningLetter, GRADE_META)

from .common import _clip_to_field

class OfferLetterTemplateView(APIView):
    """Generate Excel template for Offer Letter upload."""

    def get(self, request):
        from django.http import HttpResponse
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from ..offer_letter import SALARY_COMPONENTS, ANNEXURE_EMP_FIELDS

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Offer Letters'

        # Letter-level identity + Annexure employee details + salary-component columns
        identity = [
            'SR NO', 'Employee ID *', 'Title (Mr./Ms.)', 'Employee Name *', 'Email *', 'Department',
            'Function', 'Current Designation', 'New Designation', 'Cadre', 'Grade',
            'Date of Joining', 'Work Location', 'Current CTC *', 'New CTC *',
            'Increment %', 'Promotion %', 'Performance Rating', 'Performance Assessment', 'Grade Label',
            'Special Reward (One-time)', 'Special Reward Note',
            'Effective Date *', 'Remarks',
        ]
        component_headers = [c[3] for c in SALARY_COMPONENTS]
        headers = identity + component_headers

        hf = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        hf_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        hf_green = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        comp_start = len(identity)  # 0-based index where component columns begin

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            is_required = '*' in h
            is_comp = ci > comp_start
            c.fill = hf if is_required else (hf_green if is_comp else hf_yellow)
            c.font = Font(color='FFFFFF' if is_required else '000000', bold=True, size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

        # Sample rows (dict keyed by header; blank components demonstrate the
        # "not eligible → leave blank → hidden in letter" behaviour)
        sample_dicts = [
            {'SR NO': 1, 'Employee ID *': 'EMP001', 'Title (Mr./Ms.)': 'Mr.',
             'Employee Name *': 'Rahul Sharma', 'Email *': 'rahul.sharma@apis.com',
             'Department': 'Sales', 'Function': 'Field Sales', 'Current Designation': 'Sales Manager',
             'New Designation': 'Senior Sales Manager', 'Cadre': 'M2', 'Grade': 'G5',
             'Date of Joining': '01/06/2019', 'Work Location': 'Delhi HO',
             'Current CTC *': 600000, 'New CTC *': 660000, 'Increment %': 10, 'Promotion %': 5,
             'Performance Rating': 'A', 'Performance Assessment': 'Strong Performer',
             'Grade Label': 'Outstanding', 'Special Reward (One-time)': 25000,
             'Special Reward Note': 'For outstanding project delivery',
             'Effective Date *': '2026-04-01', 'Remarks': 'Excellent performer',
             'Basic Salary (Monthly)': 27500, 'HRA (Monthly)': 11000, 'Special Allowance (Monthly)': 8000,
             'Employer PF (Monthly)': 3300, 'Statutory Bonus (Monthly)': 2292, 'Variable Pay (Monthly)': 3333},
            {'SR NO': 2, 'Employee ID *': 'EMP002', 'Title (Mr./Ms.)': 'Ms.',
             'Employee Name *': 'Priya Singh', 'Email *': 'priya.singh@apis.com',
             'Department': 'Operations', 'Function': 'Operations', 'Current Designation': 'Executive',
             'New Designation': 'Senior Executive', 'Cadre': 'M1', 'Grade': 'G4',
             'Date of Joining': '15/03/2021', 'Work Location': 'Mumbai',
             'Current CTC *': 450000, 'New CTC *': 540000, 'Increment %': 12, 'Promotion %': 8,
             'Performance Rating': 'A+', 'Performance Assessment': 'Outstanding Performer',
             'Grade Label': 'Exceptional', 'Effective Date *': '2026-04-01', 'Remarks': 'Ready for promotion',
             'Basic Salary (Monthly)': 22500, 'HRA (Monthly)': 9000, 'Special Allowance (Monthly)': 6000,
             'Employer PF (Monthly)': 2700, 'Statutory Bonus (Monthly)': 1875},
            {'SR NO': 3, 'Employee ID *': 'EMP003', 'Title (Mr./Ms.)': 'Mr.',
             'Employee Name *': 'Amit Kumar', 'Email *': 'amit.kumar@apis.com',
             'Department': 'IT', 'Function': 'Information Technology', 'Current Designation': 'Associate',
             'New Designation': 'Associate', 'Cadre': 'E3', 'Grade': 'G3',
             'Date of Joining': '10/01/2023', 'Work Location': 'Delhi HO',
             'Current CTC *': 280000, 'New CTC *': 340000, 'Increment %': 5, 'Promotion %': 0,
             'Performance Rating': 'B', 'Performance Assessment': 'Solid Performer',
             'Grade Label': 'Meets Target', 'Effective Date *': '2026-04-01', 'Remarks': '',
             'Basic Salary (Monthly)': 14000, 'HRA (Monthly)': 5600, 'Special Allowance (Monthly)': 4000},
        ]

        date_col = headers.index('Effective Date *') + 1
        for ri, d in enumerate(sample_dicts, 2):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=ri, column=ci, value=d.get(h, ''))
                c.border = border
                if ci == date_col:
                    c.number_format = 'yyyy-mm-dd'

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="OfferLetter_Template.xlsx"'
        return resp


def _process_offer_batch(rows, batch_id, send_emails):
    """Background worker: generate PDFs (+ optional emails) and update batch progress.

    Runs in its own thread so a 500-1000 employee run never blocks/times-out the
    HTTP request. Uses ONE shared SMTP connection for the whole batch."""
    import io as _io
    import re
    from django.core.files.base import ContentFile
    from django.utils import timezone
    from django.db import connections
    from django.core.mail import get_connection
    from django.conf import settings
    from ..offer_letter import generate_offer_letter_pdf, send_offer_letter_email
    from ..models import OfferLetterBatch

    def _offer_mail_connection():
        """SMTP connection using the dedicated offer-letter account (separate
        from the account used for PMS login OTPs)."""
        return get_connection(
            host=settings.EMAIL_HOST, port=settings.EMAIL_PORT,
            username=settings.OFFER_LETTER_EMAIL_HOST_USER,
            password=settings.OFFER_LETTER_EMAIL_HOST_PASSWORD,
            use_tls=settings.EMAIL_USE_TLS, timeout=30,
        )

    batch = None
    mail_conn = None
    try:
        batch = OfferLetterBatch.objects.get(batch_id=batch_id)
        # 30s socket timeout so a single hung email can never freeze the whole
        # batch (no EMAIL_TIMEOUT is configured globally).
        if send_emails:
            try:
                mail_conn = _offer_mail_connection()
                mail_conn.open()
            except Exception:
                mail_conn = None

        def _send_with_retry(email, name, pdf_io, eff_date, oid, attach_name):
            """Send via the shared SMTP connection; if it has dropped mid-batch
            (common on 500+ sends), reopen a fresh connection once and retry."""
            nonlocal mail_conn
            try:
                send_offer_letter_email(email, name, pdf_io, eff_date, oid, connection=mail_conn, filename=attach_name)
            except Exception:
                try:
                    if mail_conn is not None:
                        mail_conn.close()
                except Exception:
                    pass
                mail_conn = _offer_mail_connection()
                mail_conn.open()
                pdf_io.seek(0)
                send_offer_letter_email(email, name, pdf_io, eff_date, oid, connection=mail_conn, filename=attach_name)

        gen = eml = fail = proc = 0
        errs = []
        total = len(rows)
        for r in rows:
            offer = None
            try:
                offer = OfferLetter.objects.create(
                    employee=None,
                    employee_code=_clip_to_field(OfferLetter, 'employee_code', r['emp_id']),
                    employee_name=_clip_to_field(OfferLetter, 'employee_name', r['name']),
                    letter_type=r['letter_type'],
                    current_ctc=r['current_ctc'], new_ctc=r['new_ctc'],
                    increment_pct=r['increment_pct'], promotion_pct=r['promotion_pct'],
                    effective_date=r['effective_date'],
                    old_designation=_clip_to_field(OfferLetter, 'old_designation', r['current_designation']),
                    new_designation=_clip_to_field(OfferLetter, 'new_designation', r['new_designation']),
                    performance_rating=_clip_to_field(OfferLetter, 'performance_rating', r['performance_rating']),
                    grade_label=_clip_to_field(OfferLetter, 'grade_label', r['grade_label']),
                    salutation=_clip_to_field(OfferLetter, 'salutation', r['salutation']),
                    assessment=_clip_to_field(OfferLetter, 'assessment', r['assessment']),
                    function=_clip_to_field(OfferLetter, 'function', r['function']),
                    cadre=_clip_to_field(OfferLetter, 'cadre', r['cadre']),
                    grade=_clip_to_field(OfferLetter, 'grade', r['grade']),
                    date_of_joining=_clip_to_field(OfferLetter, 'date_of_joining', r['date_of_joining']),
                    work_location=_clip_to_field(OfferLetter, 'work_location', r['work_location']),
                    salary_breakup=r['salary_breakup'],
                    special_reward=r['special_reward'],
                    special_reward_note=_clip_to_field(OfferLetter, 'special_reward_note', r['special_reward_note']),
                    email_address=r['email'],
                    department=_clip_to_field(OfferLetter, 'department', r['department']),
                    batch_id=batch_id, status='pending',
                )
                pdf_bytes = generate_offer_letter_pdf(
                    None, r['current_ctc'], r['new_ctc'], r['increment_pct'], r['promotion_pct'],
                    r['effective_date'],
                    old_designation=r['current_designation'], new_designation=r['new_designation'],
                    performance_rating=r['performance_rating'], grade_label=r['grade_label'],
                    employee_id=r['emp_id'], employee_name=r['name'], department=r['department'],
                    salutation_title=r['salutation'], assessment=r['assessment'],
                    emp_details=r['emp_details'], salary_breakup=r['salary_breakup'],
                    special_reward=r['special_reward'], special_reward_note=r['special_reward_note'],
                ).getvalue()
                safe_id = re.sub(r'[^A-Za-z0-9_.-]', '_', str(r['emp_id']))[:40]
                offer.pdf_file.save(f"offer_{safe_id}_{offer.id}.pdf",
                                    ContentFile(pdf_bytes), save=True)
                gen += 1

                if send_emails:
                    if not r['email']:
                        # No email in the sheet — record as a failure so it is visible.
                        offer.status = 'failed'
                        offer.save(update_fields=['status'])
                        fail += 1
                        errs.append(f"{r['emp_id']} ({r['name']}): no email address in the sheet")
                    else:
                        try:
                            _send_with_retry(r['email'], r['name'], _io.BytesIO(pdf_bytes),
                                             r['effective_date'], offer.id, _letter_filename(offer))
                            offer.status = 'sent'
                            offer.email_sent = True
                            offer.email_sent_at = timezone.now()
                            offer.save(update_fields=['status', 'email_sent', 'email_sent_at'])
                            eml += 1
                        except Exception as ee:
                            offer.status = 'failed'
                            offer.save(update_fields=['status'])
                            fail += 1
                            errs.append(f"{r['emp_id']} ({r['name']}) <{r['email']}>: email failed: {ee}")
            except Exception as e:
                fail += 1
                errs.append(f"{r['emp_id']} ({r.get('name','')}): {e}")
                # Whenever we count a failure, an OfferLetter row with
                # status='failed' must exist — otherwise the batch summary
                # and the Letters History dashboard disagree (batch says N
                # failed, History shows none) because History is just a
                # filtered view of this same table.
                if offer is not None:
                    # Row was created but something after that (PDF gen,
                    # pdf_file.save) blew up — it's otherwise stuck at the
                    # 'pending' default forever.
                    try:
                        offer.status = 'failed'
                        offer.save(update_fields=['status'])
                    except Exception:
                        pass
                else:
                    # OfferLetter.objects.create() itself failed (bad/oversized
                    # data for some field) — no row exists yet at all. Persist a
                    # minimal one so this employee still shows up as failed
                    # instead of vanishing without a trace.
                    try:
                        OfferLetter.objects.create(
                            employee=None, employee_code=r.get('emp_id', ''),
                            employee_name=r.get('name', ''),
                            current_ctc=0, new_ctc=0,
                            effective_date=r.get('effective_date') or timezone.now().date(),
                            email_address=r.get('email', ''), department=r.get('department', ''),
                            batch_id=batch_id, status='failed',
                        )
                    except Exception:
                        pass  # last resort — if even this can't be saved, nothing more we can do

            proc += 1
            if proc % 5 == 0 or proc == total:  # flush progress periodically
                OfferLetterBatch.objects.filter(batch_id=batch_id).update(
                    processed=proc, generated=gen, emailed=eml, failed=fail,
                    errors=errs[:50], updated_at=timezone.now())

        OfferLetterBatch.objects.filter(batch_id=batch_id).update(
            processed=proc, generated=gen, emailed=eml, failed=fail,
            errors=errs[:500], status='completed', updated_at=timezone.now())
    except Exception as e:
        if batch is not None:
            from django.utils import timezone as _tz
            OfferLetterBatch.objects.filter(batch_id=batch_id).update(
                status='error', errors=[str(e)], updated_at=_tz.now())
    finally:
        if mail_conn is not None:
            try:
                mail_conn.close()
            except Exception:
                pass
        connections.close_all()  # release this thread's DB connections


class OfferLetterUploadView(APIView):
    """Parse the uploaded Excel instantly, then generate letters in a background
    thread. Returns a batch_id the UI polls for progress — scales to 500-1000+."""
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        from datetime import datetime, date

        send_emails = str(request.data.get('send_emails', 'false')).lower() == 'true'

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {str(e)}'}, status=400)

        from ..offer_letter import SALARY_COMPONENTS
        HEADER_MAP = {
            'sr no': 'sr_no', 'employee id': 'employee_id',
            'title (mr./ms.)': 'salutation', 'title': 'salutation',
            'employee name': 'name',
            'email': 'email', 'department': 'department', 'function': 'function',
            'current designation': 'current_designation', 'new designation': 'new_designation',
            'cadre': 'cadre', 'grade': 'grade',
            'date of joining': 'date_of_joining', 'work location': 'work_location',
            'current ctc': 'current_ctc', 'new ctc': 'new_ctc',
            'increment %': 'increment_pct', 'promotion %': 'promotion_pct',
            'performance rating': 'performance_rating',
            'performance assessment': 'assessment',
            'grade label': 'grade_label',
            'special reward (one-time)': 'special_reward', 'special reward': 'special_reward',
            'special reward note': 'special_reward_note',
            'effective date': 'effective_date', 'remarks': 'remarks',
        }
        # salary-component columns keyed by their (lower-cased) Excel header → component key
        COMP_MAP = {c[3].strip().lower(): c[0] for c in SALARY_COMPONENTS}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map = {}
        comp_col = {}   # component_key → column index
        for ci, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace('*', '').strip()
            if key in HEADER_MAP:
                col_map[HEADER_MAP[key]] = ci
            elif key in COMP_MAP:
                comp_col[COMP_MAP[key]] = ci

        required = ['employee_id', 'name', 'email', 'current_ctc', 'new_ctc', 'effective_date']
        if not all(f in col_map for f in required):
            return Response({'error': 'Missing required columns',
                             'required': ['Employee ID', 'Employee Name', 'Email', 'Current CTC', 'New CTC', 'Effective Date'],
                             'mapped': list(col_map.keys())}, status=400)

        def get_val(row, field, default=None):
            if field not in col_map:
                return default
            ci = col_map[field]
            return row[ci] if ci < len(row) else default

        def sf(val, default=0):
            if val is None or str(val).strip() == '':
                return default
            try:
                return float(str(val).replace(',', ''))
            except Exception:
                return default

        def format_date(val):
            if val is None or str(val).strip() == '':
                return None
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date):
                return val
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
                try:
                    return datetime.strptime(str(val).strip(), fmt).date()
                except Exception:
                    pass
            return None

        def format_doj_display(val):
            """Clean "06 May 2024" display string for the Annexure's Date of
            Joining. Avoids the " 00:00:00" artifact that appears when the
            Excel cell is a native datetime object and gets naively str()'d,
            and normalises whatever date format was typed in the sheet."""
            if val is None or str(val).strip() == '':
                return ''
            if isinstance(val, (datetime, date)):
                return val.strftime('%d %B %Y')
            s = str(val).strip()
            try:
                from dateutil import parser as _dateparser
                return _dateparser.parse(s, dayfirst=True).strftime('%d %B %Y')
            except Exception:
                return s  # not a recognisable date — show the original text as-is

        # ── Fast parse: turn every valid row into a plain dict (no PDF/DB yet) ──
        # Every row that is dropped or altered is recorded and reported back —
        # silently skipping rows in a 500-employee run means people simply never
        # get their letter and nobody notices.
        rows = []
        skipped_rows = []        # missing Employee ID or Name
        date_fallback_rows = []  # Effective Date missing/unreadable -> today
        missing_email_rows = []  # no email address (only matters when sending)
        duplicate_codes = []     # same Employee ID appearing twice in the file
        seen_codes = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            emp_id = str(get_val(row, 'employee_id') or '').strip()
            name = str(get_val(row, 'name') or '').strip()
            if not emp_id or not name:
                skipped_rows.append(row_idx)
                continue
            code_key = emp_id.lower()
            if code_key in seen_codes:
                duplicate_codes.append(f'{emp_id} (rows {seen_codes[code_key]} & {row_idx})')
            else:
                seen_codes[code_key] = row_idx
            increment_pct = sf(get_val(row, 'increment_pct'))
            promotion_pct = sf(get_val(row, 'promotion_pct'))
            current_designation = str(get_val(row, 'current_designation') or '').strip()
            new_designation = str(get_val(row, 'new_designation') or '').strip()
            function = str(get_val(row, 'function') or '').strip()
            cadre = str(get_val(row, 'cadre') or '').strip()
            grade = str(get_val(row, 'grade') or '').strip()
            date_of_joining = format_doj_display(get_val(row, 'date_of_joining'))
            work_location = str(get_val(row, 'work_location') or '').strip()

            salary_breakup = {}
            for ckey, ci in comp_col.items():
                cval = sf(row[ci] if ci < len(row) else None, default=0)
                if cval:
                    salary_breakup[ckey] = cval

            letter_type = 'increment'
            if new_designation and new_designation != current_designation:
                letter_type = 'promotion' if promotion_pct > 0 else 'redesignation'
            if increment_pct > 0 and promotion_pct > 0:
                letter_type = 'combined'

            # An appraisal letter states WHEN the revision takes effect, so a
            # silently-wrong date is a serious error — record the fallback.
            eff_date = format_date(get_val(row, 'effective_date'))
            if eff_date is None:
                date_fallback_rows.append(row_idx)
                eff_date = date.today()

            row_email = str(get_val(row, 'email') or '').strip()
            if not row_email:
                missing_email_rows.append(row_idx)

            rows.append({
                'emp_id': emp_id, 'name': name,
                'email': row_email,
                'current_ctc': sf(get_val(row, 'current_ctc')),
                'new_ctc': sf(get_val(row, 'new_ctc')),
                'increment_pct': increment_pct, 'promotion_pct': promotion_pct,
                'current_designation': current_designation, 'new_designation': new_designation,
                'performance_rating': str(get_val(row, 'performance_rating') or '').strip(),
                'assessment': str(get_val(row, 'assessment') or '').strip(),
                'grade_label': str(get_val(row, 'grade_label') or '').strip(),
                'salutation': str(get_val(row, 'salutation') or '').strip(),
                'department': str(get_val(row, 'department') or '').strip(),
                'function': function, 'cadre': cadre, 'grade': grade,
                'date_of_joining': date_of_joining, 'work_location': work_location,
                'special_reward': sf(get_val(row, 'special_reward')),
                'special_reward_note': str(get_val(row, 'special_reward_note') or '').strip(),
                'effective_date': eff_date,
                'salary_breakup': salary_breakup,
                'emp_details': {'function': function, 'cadre': cadre, 'grade': grade,
                                'date_of_joining': date_of_joining, 'work_location': work_location},
                'letter_type': letter_type,
            })

        if not rows:
            return Response({'error': 'No valid employee rows found in the file.'}, status=400)

        # ── Kick off background generation and return immediately ──
        import threading
        import uuid
        from ..models import OfferLetterBatch
        batch_id = uuid.uuid4().hex[:16]
        OfferLetterBatch.objects.create(batch_id=batch_id, total=len(rows),
                                        send_emails=send_emails, status='running')
        threading.Thread(target=_process_offer_batch, args=(rows, batch_id, send_emails),
                         daemon=True).start()

        def _rowlist(nums, limit=15):
            head = ', '.join(str(n) for n in nums[:limit])
            return head + (f' …and {len(nums) - limit} more' if len(nums) > limit else '')

        warnings = []
        if skipped_rows:
            warnings.append(f'{len(skipped_rows)} row(s) SKIPPED — no Employee ID or Name '
                            f'(sheet row {_rowlist(skipped_rows)}). These employees get NO letter.')
        if duplicate_codes:
            warnings.append(f'{len(duplicate_codes)} duplicate Employee ID(s) — a separate letter '
                            f'and email is produced for each occurrence: '
                            f'{"; ".join(duplicate_codes[:10])}'
                            + (' …' if len(duplicate_codes) > 10 else ''))
        if date_fallback_rows:
            warnings.append(f'{len(date_fallback_rows)} row(s) had a missing/unreadable Effective Date '
                            f'— today\'s date was used instead (sheet row {_rowlist(date_fallback_rows)}). '
                            f'Check these before sending.')
        if send_emails and missing_email_rows:
            warnings.append(f'{len(missing_email_rows)} row(s) have no email address — the letter is '
                            f'generated but cannot be emailed (sheet row {_rowlist(missing_email_rows)}).')

        return Response({
            'message': f'Processing {len(rows)} letter(s) in the background…',
            'batch_id': batch_id, 'total': len(rows), 'send_emails': send_emails,
            'warnings': warnings,
        })



def _letter_filename(offer):
    """<EMPCODE>_PMS <FY>.pdf, e.g. "EMP0001_PMS 2025-2026.pdf" — searchable by
    employee code when hundreds of letters are archived together. FY matches
    the same reviewed-year calendar used in the letter body (_fy_context in
    offer_letter.py): effective date's year Y -> reviewed FY "(Y-1)-Y"."""
    code = offer.employee_code or (offer.employee.employee_id if offer.employee else str(offer.id))
    safe_code = re.sub(r'[^A-Za-z0-9_-]', '_', str(code))[:40] or str(offer.id)
    if offer.effective_date:
        ey = offer.effective_date.year
        fy = f"{ey - 1}-{ey}"
    else:
        fy = "PMS"
    return f"{safe_code}_PMS {fy}.pdf"


class OfferLetterPDFView(APIView):
    """Download/view a generated offer-letter PDF."""
    def get(self, request, offer_letter_id):
        from django.http import FileResponse
        try:
            offer = OfferLetter.objects.get(id=offer_letter_id)
        except OfferLetter.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        if not offer.pdf_file:
            return Response({'error': 'No PDF generated for this letter'}, status=404)
        return FileResponse(offer.pdf_file.open('rb'), content_type='application/pdf',
                            filename=_letter_filename(offer))


class OfferLetterDownloadAllView(APIView):
    """Bulk-download every (optionally filtered) stored letter as one ZIP,
    each PDF named via _letter_filename() for easy searching once archived."""
    def get(self, request):
        import zipfile
        from django.http import HttpResponse
        from django.db.models import Q

        qs = OfferLetter.objects.exclude(pdf_file='').exclude(pdf_file__isnull=True).order_by('employee_code')

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(employee_name__icontains=search) | Q(employee_code__icontains=search) |
                           Q(email_address__icontains=search) | Q(department__icontains=search))
        status_filter = (request.query_params.get('status') or '').strip()
        if status_filter:
            qs = qs.filter(status=status_filter)
        batch_id = (request.query_params.get('batch_id') or '').strip()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)

        letters = list(qs)
        if not letters:
            return Response({'error': 'No letters with a stored PDF match this filter.'}, status=404)

        buf = io.BytesIO()
        used_names = {}
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for o in letters:
                name = _letter_filename(o)
                # Disambiguate if the same filename would occur twice (e.g. a
                # re-imported/duplicate employee code) instead of one entry
                # silently overwriting the other inside the zip.
                if name in used_names:
                    used_names[name] += 1
                    base, ext = name.rsplit('.', 1)
                    name = f"{base} ({used_names[name]}).{ext}"
                else:
                    used_names[name] = 0
                try:
                    o.pdf_file.open('rb')
                    zf.writestr(name, o.pdf_file.read())
                finally:
                    o.pdf_file.close()
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/zip')
        resp['Content-Disposition'] = f'attachment; filename="APIS_Offer_Letters_{len(letters)}.zip"'
        return resp


class OfferLetterBatchStatusView(APIView):
    """Poll bulk-generation progress. Returns live counts while running and the
    per-letter results once completed."""
    def get(self, request, batch_id):
        from ..models import OfferLetterBatch
        from django.utils import timezone
        from datetime import timedelta
        try:
            b = OfferLetterBatch.objects.get(batch_id=batch_id)
        except OfferLetterBatch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=404)

        # If a running batch hasn't advanced for 5 min, the worker died — mark it errored
        if b.status == 'running' and b.updated_at < timezone.now() - timedelta(minutes=5):
            b.status = 'error'
            b.errors = (b.errors or []) + ['Batch stalled — generation stopped unexpectedly. '
                                           'Please re-upload the remaining rows.']
            b.save(update_fields=['status', 'errors'])

        data = {
            'batch_id': b.batch_id, 'status': b.status, 'total': b.total,
            'processed': b.processed, 'generated': b.generated, 'emailed': b.emailed,
            'failed': b.failed, 'send_emails': b.send_emails, 'errors': b.errors,
        }
        if b.status == 'completed':
            # Map each failure reason back to its employee code so the UI can show
            # WHY a letter failed, not just that it did. Batch errors look like
            # "EMP001 (Name) <email>: reason".
            reason_map = {}
            for line in (b.errors or []):
                code = str(line).split(' ', 1)[0].split(':', 1)[0].strip()
                if code:
                    reason_map.setdefault(code, str(line))
            data['results'] = [
                {'employee_id': o.employee_code, 'name': o.employee_name, 'status': o.status,
                 'email': o.email_address,
                 'message': reason_map.get(o.employee_code, '') if o.status == 'failed' else '',
                 'pdf_url': f'/api/pms/offer-letter/{o.id}/pdf/'}
                for o in OfferLetter.objects.filter(batch_id=b.batch_id).order_by('id')
            ]
        return Response(data)


class OfferLetterHistoryView(APIView):
    """Persistent dashboard of every generated letter — survives page refresh,
    unlike the in-memory batch-results view which resets once you navigate
    away. Every letter that OfferLetterUploadView ever queued lives in the
    OfferLetter table regardless of email/send outcome, so this is simply
    a browsable window onto that table."""
    def get(self, request):
        from django.db.models import Q

        qs = OfferLetter.objects.all().order_by('-created_at')

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(employee_name__icontains=search) |
                           Q(employee_code__icontains=search) |
                           Q(email_address__icontains=search) |
                           Q(department__icontains=search))

        status_filter = (request.query_params.get('status') or '').strip()
        if status_filter:
            qs = qs.filter(status=status_filter)

        batch_id = (request.query_params.get('batch_id') or '').strip()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)

        total_count = qs.count()
        try:
            limit = max(1, min(500, int(request.query_params.get('limit', 100))))
        except (TypeError, ValueError):
            limit = 100
        try:
            offset = max(0, int(request.query_params.get('offset', 0)))
        except (TypeError, ValueError):
            offset = 0

        page = qs[offset:offset + limit]
        results = [{
            'id': o.id, 'employee_id': o.employee_code, 'name': o.employee_name,
            'department': o.department, 'email': o.email_address,
            'letter_type': o.letter_type, 'status': o.status,
            'email_sent': o.email_sent,
            'email_sent_at': o.email_sent_at.isoformat() if o.email_sent_at else None,
            'batch_id': o.batch_id, 'created_at': o.created_at.isoformat(),
            'pdf_url': f'/api/pms/offer-letter/{o.id}/pdf/' if o.pdf_file else None,
        } for o in page]

        # Summary counts (over the FULL filtered set, not just this page).
        summary = {
            'total': total_count,
            'sent': qs.filter(status='sent').count(),
            'failed': qs.filter(status='failed').count(),
            'pending': qs.filter(status='pending').count(),
        }

        return Response({
            'results': results, 'count': total_count,
            'limit': limit, 'offset': offset, 'summary': summary,
        })

    def delete(self, request):
        """Clear DB: permanently deletes every stored letter record, its PDF
        file on disk, and every batch record. Irreversible — the frontend
        must confirm with the user before calling this."""
        letters = list(OfferLetter.objects.all())
        deleted_files = 0
        for o in letters:
            if o.pdf_file:
                try:
                    o.pdf_file.delete(save=False)
                    deleted_files += 1
                except Exception:
                    pass
        deleted_count = len(letters)
        OfferLetter.objects.all().delete()
        from ..models import OfferLetterBatch
        OfferLetterBatch.objects.all().delete()
        return Response({
            'message': f'Cleared {deleted_count} letter record(s) and {deleted_files} PDF file(s).',
            'deleted': deleted_count,
        })

