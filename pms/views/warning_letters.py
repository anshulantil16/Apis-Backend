"""Warning / disciplinary letters — second Letters Generator component.

Mirrors the appraisal pipeline against the separate WarningLetter table, plus
a single-employee create endpoint. PDF rendering lives in pms/warning_letter.py.
"""
import io
import os
import re
import secrets
import openpyxl
from django.core.cache import cache
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from ..models import (PMSEmployee, PMSAuditLog, PMSSettings, OfferLetter,
                      WarningLetter, GRADE_META)

from .common import _clip_to_field

# Mirrors the appraisal-letter pipeline (template → upload → background batch →
# history → ZIP) against the separate WarningLetter table, plus a
# single-employee create endpoint since warnings are usually issued one at a time.
# ══════════════════════════════════════════════════════════════════════════════

WARNING_TYPE_ALIASES = {
    'verbal': 'verbal', 'verbal warning': 'verbal',
    'first': 'first', 'first written warning': 'first', 'first warning': 'first',
    '1st': 'first', '1st written warning': 'first',
    'second': 'second', 'second written warning': 'second', 'second warning': 'second',
    '2nd': 'second', '2nd written warning': 'second',
    'final': 'final', 'final warning': 'final', 'final written warning': 'final',
    'show cause': 'show_cause', 'show_cause': 'show_cause', 'show cause notice': 'show_cause',
}


def _normalise_warning_type(val):
    """Map whatever HR typed in the sheet to a valid choice. Unrecognised text is
    NOT discarded — it is returned as a custom heading label so the letter still
    carries the wording they intended instead of silently becoming generic."""
    s = str(val or '').strip()
    if not s:
        return 'first', ''
    key = ' '.join(s.lower().replace('-', ' ').split())
    if key in WARNING_TYPE_ALIASES:
        return WARNING_TYPE_ALIASES[key], ''
    return 'first', s


def _split_emails(raw):
    """Free-text CC box -> clean address list (comma / semicolon / newline)."""
    raw = str(raw or '').replace(';', ',').replace('\n', ',')
    return [e.strip() for e in raw.split(',') if e.strip()]


def _warning_filename(w):
    """<EMPCODE>_Warning <YYYY-MM-DD>.pdf — searchable by employee code, and
    date-stamped because one employee can accumulate several warnings."""
    code = w.employee_code or str(w.id)
    safe_code = re.sub(r'[^A-Za-z0-9_-]', '_', str(code))[:40] or str(w.id)
    d = w.letter_date.strftime('%Y-%m-%d') if w.letter_date else 'undated'
    return f"{safe_code}_Warning {d}.pdf"


def _build_warning_pdf(w):
    """Render the PDF for a saved WarningLetter row."""
    from ..warning_letter import generate_warning_letter_pdf
    return generate_warning_letter_pdf(
        employee_name=w.employee_name, employee_code=w.employee_code,
        designation=w.designation, department=w.department,
        salutation=w.salutation or 'Mr./Ms.',
        warning_type=w.warning_type, warning_type_label=w.warning_type_label,
        subject=w.subject, incident_date=w.incident_date,
        incident_description=w.incident_description,
        previous_warning_ref=w.previous_warning_ref,
        corrective_action=w.corrective_action,
        response_due_days=w.response_due_days, letter_date=w.letter_date,
        issued_by=w.issued_by, issued_by_designation=w.issued_by_designation,
        emp_details={'work_location': w.work_location,
                     'reporting_manager': w.reporting_manager,
                     'function': w.function, 'grade': w.grade, 'cadre': w.cadre,
                     'date_of_joining': w.date_of_joining},
        remarks=w.remarks,
    )


def _warning_mail_connection():
    """SMTP connection for warning letters — the HR account (same as offer
    letters unless a dedicated one is configured), never the PMS login OTP
    account."""
    from django.core.mail import get_connection
    from django.conf import settings
    return get_connection(
        host=settings.EMAIL_HOST, port=settings.EMAIL_PORT,
        username=getattr(settings, 'WARNING_LETTER_EMAIL_HOST_USER', None)
        or settings.OFFER_LETTER_EMAIL_HOST_USER,
        password=getattr(settings, 'WARNING_LETTER_EMAIL_HOST_PASSWORD', None)
        or settings.OFFER_LETTER_EMAIL_HOST_PASSWORD,
        use_tls=settings.EMAIL_USE_TLS, timeout=30,
    )


def _generate_and_maybe_send(w, send_email, mail_conn=None):
    """Render + store the PDF for one WarningLetter and optionally email it.
    Returns (generated, emailed, error). Never raises — the caller records the
    outcome, so one bad row can't kill a whole batch."""
    import io as _io
    from django.core.files.base import ContentFile
    from django.utils import timezone
    from ..warning_letter import send_warning_letter_email
    try:
        pdf_bytes = _build_warning_pdf(w).getvalue()
        safe_id = re.sub(r'[^A-Za-z0-9_.-]', '_', str(w.employee_code or w.id))[:40]
        w.pdf_file.save(f"warning_{safe_id}_{w.id}.pdf", ContentFile(pdf_bytes), save=True)
    except Exception as e:
        try:
            w.status = 'failed'
            w.save(update_fields=['status'])
        except Exception:
            pass
        return False, False, f'PDF generation failed: {e}'

    if not send_email:
        return True, False, None
    if not w.email_address:
        w.status = 'failed'
        w.save(update_fields=['status'])
        return True, False, 'no email address'
    try:
        # Never Cc the employee on their own letter — they are already the To,
        # and a duplicate delivers them two copies of their own warning.
        cc = [e for e in w.cc_list
              if e.strip().lower() != (w.email_address or '').strip().lower()]
        send_warning_letter_email(w.email_address, w.employee_name, _io.BytesIO(pdf_bytes),
                                  w.type_label, w.id, connection=mail_conn,
                                  filename=_warning_filename(w), cc=cc)
        w.status = 'sent'
        w.email_sent = True
        w.email_sent_at = timezone.now()
        w.save(update_fields=['status', 'email_sent', 'email_sent_at'])
        return True, True, None
    except Exception as e:
        w.status = 'failed'
        w.save(update_fields=['status'])
        return True, False, f'email failed: {e}'


def _process_warning_batch(rows, batch_id, send_emails):
    """Background worker for bulk warning letters — same shape as
    _process_offer_batch: one shared SMTP connection, periodic progress flushes,
    and a guarantee that every counted failure leaves a status='failed' row so
    the batch summary and the History dashboard can never disagree."""
    from django.utils import timezone
    from django.db import connections
    from ..models import WarningLetterBatch

    batch = None
    mail_conn = None
    try:
        batch = WarningLetterBatch.objects.get(batch_id=batch_id)
        if send_emails:
            try:
                mail_conn = _warning_mail_connection()
                mail_conn.open()
            except Exception:
                mail_conn = None

        gen = eml = fail = proc = 0
        errs = []
        total = len(rows)
        for r in rows:
            w = None
            try:
                w = WarningLetter.objects.create(
                    employee=None,
                    employee_code=_clip_to_field(WarningLetter, 'employee_code', r['emp_id']),
                    employee_name=_clip_to_field(WarningLetter, 'employee_name', r['name']),
                    salutation=_clip_to_field(WarningLetter, 'salutation', r['salutation']),
                    designation=_clip_to_field(WarningLetter, 'designation', r['designation']),
                    department=_clip_to_field(WarningLetter, 'department', r['department']),
                    function=_clip_to_field(WarningLetter, 'function', r['function']),
                    grade=_clip_to_field(WarningLetter, 'grade', r['grade']),
                    cadre=_clip_to_field(WarningLetter, 'cadre', r['cadre']),
                    date_of_joining=_clip_to_field(WarningLetter, 'date_of_joining', r['date_of_joining']),
                    work_location=_clip_to_field(WarningLetter, 'work_location', r['work_location']),
                    reporting_manager=_clip_to_field(WarningLetter, 'reporting_manager', r['reporting_manager']),
                    warning_type=r['warning_type'],
                    warning_type_label=_clip_to_field(WarningLetter, 'warning_type_label', r['warning_type_label']),
                    subject=_clip_to_field(WarningLetter, 'subject', r['subject']),
                    incident_date=_clip_to_field(WarningLetter, 'incident_date', r['incident_date']),
                    incident_description=r['incident_description'],
                    previous_warning_ref=_clip_to_field(WarningLetter, 'previous_warning_ref', r['previous_warning_ref']),
                    corrective_action=r['corrective_action'],
                    response_due_days=r['response_due_days'],
                    letter_date=r['letter_date'],
                    issued_by=_clip_to_field(WarningLetter, 'issued_by', r['issued_by']),
                    issued_by_designation=_clip_to_field(WarningLetter, 'issued_by_designation', r['issued_by_designation']),
                    remarks=r['remarks'],
                    email_address=r['email'],
                    cc_emails=r['cc_emails'],
                    batch_id=batch_id, status='pending',
                )
                generated, emailed, err = _generate_and_maybe_send(w, send_emails, mail_conn)
                if generated:
                    gen += 1
                if emailed:
                    eml += 1
                if err:
                    fail += 1
                    errs.append(f"{r['emp_id']} ({r.get('name', '')}): {err}")
            except Exception as e:
                fail += 1
                errs.append(f"{r['emp_id']} ({r.get('name', '')}): {e}")
                if w is not None:
                    try:
                        w.status = 'failed'
                        w.save(update_fields=['status'])
                    except Exception:
                        pass
                else:
                    # create() itself failed — persist a minimal row so this
                    # employee still shows as failed instead of vanishing.
                    try:
                        WarningLetter.objects.create(
                            employee=None, employee_code=r.get('emp_id', ''),
                            employee_name=r.get('name', ''),
                            letter_date=r.get('letter_date') or timezone.now().date(),
                            email_address=r.get('email', ''),
                            batch_id=batch_id, status='failed',
                        )
                    except Exception:
                        pass

            proc += 1
            if proc % 5 == 0 or proc == total:
                WarningLetterBatch.objects.filter(batch_id=batch_id).update(
                    processed=proc, generated=gen, emailed=eml, failed=fail,
                    errors=errs[:50], updated_at=timezone.now())

        WarningLetterBatch.objects.filter(batch_id=batch_id).update(
            processed=proc, generated=gen, emailed=eml, failed=fail,
            errors=errs[:500], status='completed', updated_at=timezone.now())
    except Exception as e:
        if batch is not None:
            WarningLetterBatch.objects.filter(batch_id=batch_id).update(
                status='error', errors=[str(e)], updated_at=timezone.now())
    finally:
        if mail_conn is not None:
            try:
                mail_conn.close()
            except Exception:
                pass
        connections.close_all()


class WarningLetterTemplateView(APIView):
    """Excel template for bulk warning letters."""

    def get(self, request):
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Warning Letters'
        headers = [
            'SR NO', 'Employee ID *', 'Title (Mr./Ms.)', 'Employee Name *', 'Email *',
            'CC Emails', 'Designation', 'Department', 'Function', 'Grade', 'Cadre',
            'Date of Joining', 'Work Location', 'Reporting Manager',
            'Warning Type *', 'Subject', 'Incident Date', 'Incident Description',
            'Previous Warning Reference', 'Corrective Action', 'Response Due (Days)',
            'Letter Date', 'Issued By', 'Issued By Designation', 'Remarks',
        ]
        hf = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        hf_yellow = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            required = '*' in h
            c.fill = hf if required else hf_yellow
            c.font = Font(color='FFFFFF' if required else '000000', bold=True, size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

        samples = [
            {'SR NO': 1, 'Employee ID *': 'EMP001', 'Title (Mr./Ms.)': 'Mr.',
             'Employee Name *': 'Rahul Sharma', 'Email *': 'rahul.sharma@apis.com',
             'CC Emails': 'manager@apis.com, hr@apis.com',
             'Designation': 'Sales Officer', 'Department': 'Sales', 'Function': 'Field Sales',
             'Grade': 'G4', 'Cadre': 'O2', 'Date of Joining': '01/06/2019',
             'Work Location': 'Delhi HO', 'Reporting Manager': 'Suresh Patel',
             'Warning Type *': 'First Written Warning', 'Subject': 'Habitual Late Attendance',
             'Incident Date': '12/05/2026',
             'Incident Description': 'Reported late on 5 occasions during April 2026 without prior intimation.',
             'Previous Warning Reference': '', 'Corrective Action': 'Report by 9:30 AM daily.',
             'Response Due (Days)': '', 'Letter Date': '2026-08-04',
             'Issued By': '', 'Issued By Designation': '', 'Remarks': ''},
            {'SR NO': 2, 'Employee ID *': 'EMP002', 'Title (Mr./Ms.)': 'Ms.',
             'Employee Name *': 'Priya Singh', 'Email *': 'priya.singh@apis.com',
             'CC Emails': '',
             'Designation': 'Executive', 'Department': 'Operations', 'Function': 'Operations',
             'Grade': 'G3', 'Cadre': 'O1', 'Date of Joining': '15/03/2021',
             'Work Location': 'Mumbai', 'Reporting Manager': 'Anita Desai',
             'Warning Type *': 'Show Cause Notice', 'Subject': 'Unauthorised Absence',
             'Incident Date': '01/07/2026',
             'Incident Description': 'Absent from duty from 1 July to 10 July 2026 without approved leave.',
             'Previous Warning Reference': '', 'Corrective Action': '',
             'Response Due (Days)': 3, 'Letter Date': '2026-08-04',
             'Issued By': '', 'Issued By Designation': '', 'Remarks': ''},
        ]
        for ri, d in enumerate(samples, 2):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=ri, column=ci, value=d.get(h, ''))
                c.border = border

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22

        # Reference sheet so HR knows exactly what "Warning Type" accepts.
        ref = wb.create_sheet('Warning Types')
        ref.append(['Accepted "Warning Type *" values'])
        ref['A1'].font = Font(bold=True, size=11)
        for label in ['Verbal Warning', 'First Written Warning', 'Second Written Warning',
                      'Final Warning', 'Show Cause Notice']:
            ref.append([label])
        ref.append([])
        ref.append(['Any other text is used as-is as the letter heading.'])
        ref.append([])
        ref.append(['CC Emails: separate multiple addresses with a comma.'])
        ref.column_dimensions['A'].width = 52

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="WarningLetter_Template.xlsx"'
        return resp


class WarningLetterCreateView(APIView):
    """Create + generate a SINGLE warning letter from the on-screen form.
    Runs inline (one letter is fast) and returns the outcome immediately."""

    def post(self, request):
        from datetime import datetime as _dt, date as _d
        d = request.data

        name = str(d.get('employee_name') or '').strip()
        code = str(d.get('employee_code') or '').strip()
        if not name or not code:
            return Response({'error': 'Employee Name and Employee Code are required.'}, status=400)

        send_email = str(d.get('send_email', 'false')).lower() == 'true'
        email = str(d.get('email_address') or '').strip()
        if send_email and not email:
            return Response({'error': 'Email address is required when "send email" is enabled.'}, status=400)

        def parse_date(v, default=None):
            if v is None or str(v).strip() == '':
                return default
            if isinstance(v, _dt):
                return v.date()
            if isinstance(v, _d):
                return v
            try:
                from dateutil import parser as _p
                return _p.parse(str(v).strip(), dayfirst=True).date()
            except Exception:
                return default

        wtype_raw = d.get('warning_type') or 'first'
        valid = {c[0] for c in WarningLetter.WARNING_TYPE_CHOICES}
        if str(wtype_raw).strip() in valid:
            wtype, auto_label = str(wtype_raw).strip(), ''
        else:
            wtype, auto_label = _normalise_warning_type(wtype_raw)

        try:
            due_days = int(float(d.get('response_due_days') or 0))
        except (TypeError, ValueError):
            due_days = 0

        w = WarningLetter.objects.create(
            employee=None,
            employee_code=_clip_to_field(WarningLetter, 'employee_code', code),
            employee_name=_clip_to_field(WarningLetter, 'employee_name', name),
            salutation=_clip_to_field(WarningLetter, 'salutation', str(d.get('salutation') or '').strip()),
            designation=_clip_to_field(WarningLetter, 'designation', str(d.get('designation') or '').strip()),
            department=_clip_to_field(WarningLetter, 'department', str(d.get('department') or '').strip()),
            function=_clip_to_field(WarningLetter, 'function', str(d.get('function') or '').strip()),
            grade=_clip_to_field(WarningLetter, 'grade', str(d.get('grade') or '').strip()),
            cadre=_clip_to_field(WarningLetter, 'cadre', str(d.get('cadre') or '').strip()),
            date_of_joining=_clip_to_field(WarningLetter, 'date_of_joining', str(d.get('date_of_joining') or '').strip()),
            work_location=_clip_to_field(WarningLetter, 'work_location', str(d.get('work_location') or '').strip()),
            reporting_manager=_clip_to_field(WarningLetter, 'reporting_manager', str(d.get('reporting_manager') or '').strip()),
            warning_type=wtype,
            warning_type_label=_clip_to_field(WarningLetter, 'warning_type_label',
                                              str(d.get('warning_type_label') or auto_label).strip()),
            subject=_clip_to_field(WarningLetter, 'subject', str(d.get('subject') or '').strip()),
            incident_date=_clip_to_field(WarningLetter, 'incident_date', str(d.get('incident_date') or '').strip()),
            incident_description=str(d.get('incident_description') or '').strip(),
            previous_warning_ref=_clip_to_field(WarningLetter, 'previous_warning_ref',
                                                str(d.get('previous_warning_ref') or '').strip()),
            corrective_action=str(d.get('corrective_action') or '').strip(),
            response_due_days=due_days,
            letter_date=parse_date(d.get('letter_date'), _d.today()),
            issued_by=_clip_to_field(WarningLetter, 'issued_by', str(d.get('issued_by') or '').strip()),
            issued_by_designation=_clip_to_field(WarningLetter, 'issued_by_designation',
                                                 str(d.get('issued_by_designation') or '').strip()),
            remarks=str(d.get('remarks') or '').strip(),
            email_address=email,
            cc_emails=', '.join(_split_emails(d.get('cc_emails'))),
            batch_id='', status='pending',
        )

        conn = None
        if send_email:
            try:
                conn = _warning_mail_connection()
                conn.open()
            except Exception:
                conn = None
        try:
            generated, emailed, err = _generate_and_maybe_send(w, send_email, conn)
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

        return Response({
            'id': w.id, 'status': w.status, 'generated': generated, 'emailed': emailed,
            'error': err, 'cc': w.cc_list,
            'message': ('Letter generated and emailed.' if emailed else
                        ('Letter generated.' if generated and not err else
                         f'Letter could not be completed: {err}')),
            'pdf_url': f'/api/pms/warning-letter/{w.id}/pdf/' if w.pdf_file else None,
        }, status=200 if generated else 500)


class WarningLetterUploadView(APIView):
    """Bulk warning letters from Excel — parses instantly, generates in a
    background thread, returns a batch_id the UI polls."""
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        from datetime import datetime, date

        send_emails = str(request.data.get('send_emails', 'false')).lower() == 'true'
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided.'}, status=400)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Cannot read file: {str(e)}'}, status=400)

        HEADER_MAP = {
            'sr no': 'sr_no', 'employee id': 'employee_id',
            'title (mr./ms.)': 'salutation', 'title': 'salutation',
            'employee name': 'name', 'email': 'email',
            'cc emails': 'cc_emails', 'cc': 'cc_emails',
            'designation': 'designation', 'department': 'department', 'function': 'function',
            'grade': 'grade', 'cadre': 'cadre',
            'date of joining': 'date_of_joining', 'work location': 'work_location',
            'reporting manager': 'reporting_manager',
            'warning type': 'warning_type', 'subject': 'subject',
            'incident date': 'incident_date', 'incident description': 'incident_description',
            'previous warning reference': 'previous_warning_ref',
            'corrective action': 'corrective_action',
            'response due (days)': 'response_due_days', 'response due days': 'response_due_days',
            'letter date': 'letter_date', 'issued by': 'issued_by',
            'issued by designation': 'issued_by_designation', 'remarks': 'remarks',
        }
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        col_map = {}
        for ci, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip().lower().replace('*', '').strip()
            if key in HEADER_MAP:
                col_map[HEADER_MAP[key]] = ci

        if not all(f in col_map for f in ('employee_id', 'name')):
            return Response({'error': 'Missing required columns',
                             'required': ['Employee ID', 'Employee Name'],
                             'mapped': list(col_map.keys())}, status=400)

        def get_val(row, field, default=None):
            if field not in col_map:
                return default
            ci = col_map[field]
            return row[ci] if ci < len(row) else default

        def s_val(row, field):
            return str(get_val(row, field) or '').strip()

        def parse_date(val, default=None):
            if val is None or str(val).strip() == '':
                return default
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date):
                return val
            try:
                from dateutil import parser as _p
                return _p.parse(str(val).strip(), dayfirst=True).date()
            except Exception:
                return default

        def display_date(val):
            """Free-text-tolerant display date (keeps 'week of 12 May' as typed)."""
            if val is None or str(val).strip() == '':
                return ''
            if isinstance(val, (datetime, date)):
                return val.strftime('%d %B %Y')
            s = str(val).strip()
            try:
                from dateutil import parser as _p
                return _p.parse(s, dayfirst=True).strftime('%d %B %Y')
            except Exception:
                return s

        rows = []
        skipped_rows = []
        missing_email_rows = []
        no_detail_rows = []      # no incident description — letter reads generic
        date_fallback_rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue
            emp_id = s_val(row, 'employee_id')
            name = s_val(row, 'name')
            if not emp_id or not name:
                skipped_rows.append(row_idx)
                continue

            wtype, auto_label = _normalise_warning_type(get_val(row, 'warning_type'))
            ldate = parse_date(get_val(row, 'letter_date'))
            if ldate is None:
                if s_val(row, 'letter_date'):
                    date_fallback_rows.append(row_idx)
                ldate = date.today()

            email = s_val(row, 'email')
            if not email:
                missing_email_rows.append(row_idx)
            desc = s_val(row, 'incident_description')
            if not desc:
                no_detail_rows.append(row_idx)

            try:
                due = int(float(s_val(row, 'response_due_days') or 0))
            except (TypeError, ValueError):
                due = 0

            rows.append({
                'emp_id': emp_id, 'name': name, 'email': email,
                'cc_emails': ', '.join(_split_emails(get_val(row, 'cc_emails'))),
                'salutation': s_val(row, 'salutation'),
                'designation': s_val(row, 'designation'),
                'department': s_val(row, 'department'),
                'function': s_val(row, 'function'),
                'grade': s_val(row, 'grade'), 'cadre': s_val(row, 'cadre'),
                'date_of_joining': display_date(get_val(row, 'date_of_joining')),
                'work_location': s_val(row, 'work_location'),
                'reporting_manager': s_val(row, 'reporting_manager'),
                'warning_type': wtype, 'warning_type_label': auto_label,
                'subject': s_val(row, 'subject'),
                'incident_date': display_date(get_val(row, 'incident_date')),
                'incident_description': desc,
                'previous_warning_ref': s_val(row, 'previous_warning_ref'),
                'corrective_action': s_val(row, 'corrective_action'),
                'response_due_days': due,
                'letter_date': ldate,
                'issued_by': s_val(row, 'issued_by'),
                'issued_by_designation': s_val(row, 'issued_by_designation'),
                'remarks': s_val(row, 'remarks'),
            })

        if not rows:
            return Response({'error': 'No valid employee rows found in the file.'}, status=400)

        import threading
        import uuid
        from ..models import WarningLetterBatch
        batch_id = uuid.uuid4().hex[:16]
        WarningLetterBatch.objects.create(batch_id=batch_id, total=len(rows),
                                          send_emails=send_emails, status='running')
        threading.Thread(target=_process_warning_batch, args=(rows, batch_id, send_emails),
                         daemon=True).start()

        def _rowlist(nums, limit=15):
            head = ', '.join(str(n) for n in nums[:limit])
            return head + (f' …and {len(nums) - limit} more' if len(nums) > limit else '')

        warnings = []
        if skipped_rows:
            warnings.append(f'{len(skipped_rows)} row(s) SKIPPED — no Employee ID or Name '
                            f'(sheet row {_rowlist(skipped_rows)}). These employees get NO letter.')
        if date_fallback_rows:
            warnings.append(f'{len(date_fallback_rows)} row(s) had an unreadable Letter Date — '
                            f"today's date was used (sheet row {_rowlist(date_fallback_rows)}).")
        if no_detail_rows:
            warnings.append(f'{len(no_detail_rows)} row(s) have no Incident Description — the letter '
                            f'is generated but reads generically (sheet row {_rowlist(no_detail_rows)}). '
                            f'A disciplinary letter without specifics is hard to defend.')
        if send_emails and missing_email_rows:
            warnings.append(f'{len(missing_email_rows)} row(s) have no email address — the letter is '
                            f'generated but cannot be emailed (sheet row {_rowlist(missing_email_rows)}).')

        return Response({
            'message': f'Processing {len(rows)} warning letter(s) in the background…',
            'batch_id': batch_id, 'total': len(rows), 'send_emails': send_emails,
            'warnings': warnings,
        })


class WarningLetterBatchStatusView(APIView):
    """Poll bulk warning-letter progress."""

    def get(self, request, batch_id):
        from ..models import WarningLetterBatch
        from django.utils import timezone
        from datetime import timedelta
        try:
            b = WarningLetterBatch.objects.get(batch_id=batch_id)
        except WarningLetterBatch.DoesNotExist:
            return Response({'error': 'Batch not found'}, status=404)

        if b.status == 'running' and b.updated_at < timezone.now() - timedelta(minutes=5):
            b.status = 'error'
            b.errors = (b.errors or []) + ['Batch stalled — generation stopped unexpectedly. '
                                           'Please re-upload the remaining rows.']
            b.save(update_fields=['status', 'errors'])

        data = {
            'batch_id': b.batch_id, 'status': b.status, 'total': b.total,
            'processed': b.processed, 'generated': b.generated, 'emailed': b.emailed,
            'failed': b.failed, 'send_emails': b.send_emails, 'errors': b.errors,
        }
        if b.status == 'completed':
            reason_map = {}
            for line in (b.errors or []):
                code = str(line).split(' ', 1)[0].split(':', 1)[0].strip()
                if code:
                    reason_map.setdefault(code, str(line))
            data['results'] = [
                {'employee_id': w.employee_code, 'name': w.employee_name, 'status': w.status,
                 'email': w.email_address, 'cc': w.cc_list, 'warning_type': w.type_label,
                 'message': reason_map.get(w.employee_code, '') if w.status == 'failed' else '',
                 'pdf_url': f'/api/pms/warning-letter/{w.id}/pdf/'}
                for w in WarningLetter.objects.filter(batch_id=b.batch_id).order_by('id')
            ]
        return Response(data)


class WarningLetterPDFView(APIView):
    """Download/view a generated warning-letter PDF."""

    def get(self, request, warning_letter_id):
        from django.http import FileResponse
        try:
            w = WarningLetter.objects.get(id=warning_letter_id)
        except WarningLetter.DoesNotExist:
            return Response({'error': 'Not found'}, status=404)
        if not w.pdf_file:
            return Response({'error': 'No PDF generated for this letter'}, status=404)
        return FileResponse(w.pdf_file.open('rb'), content_type='application/pdf',
                            filename=_warning_filename(w))


class WarningLetterDownloadAllView(APIView):
    """Bulk-download every (optionally filtered) warning letter as one ZIP."""

    def get(self, request):
        import zipfile
        from django.http import HttpResponse
        from django.db.models import Q

        qs = WarningLetter.objects.exclude(pdf_file='').exclude(pdf_file__isnull=True).order_by('employee_code')

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(employee_name__icontains=search) | Q(employee_code__icontains=search) |
                           Q(email_address__icontains=search) | Q(department__icontains=search) |
                           Q(subject__icontains=search))
        status_filter = (request.query_params.get('status') or '').strip()
        if status_filter:
            qs = qs.filter(status=status_filter)
        wtype = (request.query_params.get('warning_type') or '').strip()
        if wtype:
            qs = qs.filter(warning_type=wtype)
        batch_id = (request.query_params.get('batch_id') or '').strip()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)

        letters = list(qs)
        if not letters:
            return Response({'error': 'No letters with a stored PDF match this filter.'}, status=404)

        buf = io.BytesIO()
        used_names = {}
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for w in letters:
                name = _warning_filename(w)
                if name in used_names:
                    used_names[name] += 1
                    base, ext = name.rsplit('.', 1)
                    name = f"{base} ({used_names[name]}).{ext}"
                else:
                    used_names[name] = 0
                try:
                    w.pdf_file.open('rb')
                    zf.writestr(name, w.pdf_file.read())
                finally:
                    w.pdf_file.close()
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type='application/zip')
        resp['Content-Disposition'] = f'attachment; filename="APIS_Warning_Letters_{len(letters)}.zip"'
        return resp


class WarningLetterHistoryView(APIView):
    """Persistent dashboard of every warning letter issued."""

    def get(self, request):
        from django.db.models import Q

        qs = WarningLetter.objects.all().order_by('-created_at')

        search = (request.query_params.get('search') or '').strip()
        if search:
            qs = qs.filter(Q(employee_name__icontains=search) |
                           Q(employee_code__icontains=search) |
                           Q(email_address__icontains=search) |
                           Q(department__icontains=search) |
                           Q(subject__icontains=search))
        status_filter = (request.query_params.get('status') or '').strip()
        if status_filter:
            qs = qs.filter(status=status_filter)
        wtype = (request.query_params.get('warning_type') or '').strip()
        if wtype:
            qs = qs.filter(warning_type=wtype)
        batch_id = (request.query_params.get('batch_id') or '').strip()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)

        total_count = qs.count()
        try:
            limit = max(1, min(500, int(request.query_params.get('limit', 100))))
        except (TypeError, ValueError):
            limit = 100
        try:
            offset = max(0, int(request.query_params.get('offset', 0)))
        except (TypeError, ValueError):
            offset = 0

        page = qs[offset:offset + limit]
        results = [{
            'id': w.id, 'employee_id': w.employee_code, 'name': w.employee_name,
            'department': w.department, 'designation': w.designation,
            'email': w.email_address, 'cc': w.cc_list,
            'warning_type': w.warning_type, 'warning_type_label': w.type_label,
            'subject': w.subject, 'status': w.status,
            'email_sent': w.email_sent,
            'email_sent_at': w.email_sent_at.isoformat() if w.email_sent_at else None,
            'letter_date': w.letter_date.isoformat() if w.letter_date else None,
            'batch_id': w.batch_id, 'created_at': w.created_at.isoformat(),
            'pdf_url': f'/api/pms/warning-letter/{w.id}/pdf/' if w.pdf_file else None,
        } for w in page]

        summary = {
            'total': total_count,
            'sent': qs.filter(status='sent').count(),
            'failed': qs.filter(status='failed').count(),
            'pending': qs.filter(status='pending').count(),
        }
        return Response({
            'results': results, 'count': total_count,
            'limit': limit, 'offset': offset, 'summary': summary,
        })

    def delete(self, request):
        """Clear DB: permanently deletes every warning-letter record, its PDF on
        disk, and every warning batch. Irreversible — the frontend must confirm.
        Scoped to warning letters ONLY; appraisal letters are untouched."""
        letters = list(WarningLetter.objects.all())
        deleted_files = 0
        for w in letters:
            if w.pdf_file:
                try:
                    w.pdf_file.delete(save=False)
                    deleted_files += 1
                except Exception:
                    pass
        deleted_count = len(letters)
        WarningLetter.objects.all().delete()
        from ..models import WarningLetterBatch
        WarningLetterBatch.objects.all().delete()
        return Response({
            'message': f'Cleared {deleted_count} warning letter(s) and {deleted_files} PDF file(s).',
            'deleted': deleted_count,
        })

